#!/usr/bin/python
# -*- coding: UTF-8 -*-
'''
Created on 7/12/2018

@author: JB
# -*- coding: latin-1 -*-
'''
from __future__ import division, print_function

import os
import sys
import pathlib
import time
# import datetime
import types # Ver https://docs.python.org/2/library/types.html
# import csv
import re
import math
import random
import platform
import inspect
# import traceback
# import subprocess
# import argparse
from configparser import RawConfigParser
import logging
# import importlib
# import struct
import shutil
# import gc
import unicodedata
import xml.etree.ElementTree as ET

# Paquetes de terceros
try:
    import psutil
    psutilOk = True
except:
    psutilOk = False
try:
    import dbfread
    dbfreadDisponible = True
except:
    print('clidconfig-> AVISO: error al importar dbfread')
    dbfreadDisponible = False

# ==============================================================================
if __name__ == '__main__':
    print('\nclidconfig-> ATENCION: este modulo no se puede ejecutar de forma autonoma')
    sys.exit(0)
# ==============================================================================
if '--cargadoClidconfig' in sys.argv:
    moduloPreviamenteCargado = True
    # print(f'\nclidconfig->1> moduloPreviamenteCargado: {moduloPreviamenteCargado}; sys.argv: {sys.argv}')
else:
    moduloPreviamenteCargado = False
    # print(f'\nclidconfig->1> moduloPreviamenteCargado: {moduloPreviamenteCargado}; sys.argv: {sys.argv}')
    sys.argv.append('--cargadoClidconfig')
# ==============================================================================
if '--idProceso' in sys.argv and len(sys.argv) > sys.argv.index('--idProceso') + 1:
    ARGS_idProceso = sys.argv[sys.argv.index('--idProceso') + 1]
else:
    ARGS_idProceso = str(random.randint(1, 999998))
    # ARGS_idProceso = '999999'
    sys.argv.append('--idProceso')
    sys.argv.append(ARGS_idProceso)
# ==============================================================================
if type(ARGS_idProceso) == int:
    MAIN_idProceso = ARGS_idProceso
elif type(ARGS_idProceso) == str:
    try:
        MAIN_idProceso = int(ARGS_idProceso)
    except:
        MAIN_idProceso = 0
        sys.stdout.write(f'clidconfig-> ATENCION: revisar asignacion de idProceso.\n')
        sys.stdout.write(f'ARGS_idProceso: {type(ARGS_idProceso)} {ARGS_idProceso}\n')
        sys.stdout.write(f'sys.argv: {sys.argv}\n')
else:
    MAIN_idProceso = 0
    sys.stdout.write(f'clidconfig-> ATENCION: revisar codigo de idProceso.\n')
    sys.stdout.write(f'ARGS_idProceso: {type(ARGS_idProceso)} {ARGS_idProceso}\n')
    sys.stdout.write(f'sys.argv: {sys.argv}\n')
# ==============================================================================
# Verbose provisional para la version alpha
if '-vvv' in sys.argv:
    __verbose__ = 3
elif '-vv' in sys.argv:
    __verbose__ = 2
elif '-v' in sys.argv or '--verbose' in sys.argv:
    __verbose__ = 1
else:
    # En eclipse se adopta el valor indicado en Run Configurations -> Arguments
    __verbose__ = 0
# ==============================================================================
if '-q' in sys.argv:
    __quiet__ = 1
    __verbose__ = 0
else:
    __quiet__ = 0
# ==============================================================================
CONFIGverbose = __verbose__ > 2
CONFIGverbose = True
if CONFIGverbose:
    sys.stdout.write(f'\nclidconfig-> Aviso: CONFIGverbose: {CONFIGverbose} (asignado con codigo en funcion de __verbose__: {__verbose__})\n')
if __verbose__ > 1:
    sys.stdout.write(f'\nclidconfig-> INFO:  __verbose__:   {__verbose__} (leido en linea de comandos: -v)\n')
    sys.stdout.write(f'clidconfig-> sys.argv: {sys.argv}\n')
# ==============================================================================

# ==============================================================================
# ============================ Variables GLOBALES ==============================
# ==============================================================================

# ==============================================================================
# #variablesGlobalesBasicas
# ==============================================================================
# TB = '\t'
TB = ' ' * 13
TV = ' ' * 3
TW = ' ' * 2
# ==============================================================================
# No se importa nada con: from clidconfig import *
__all__ = []
# ==============================================================================

# ==============================================================================
# #variablesGlobalesMAINentorno
# ==============================================================================
# Directorio que depende del entorno:
MAIN_HOME_DIR = str(pathlib.Path.home())
# Dir del modulo que se esta ejecutando (clidconfig.py):
MAIN_FILE_DIR = os.path.dirname(os.path.abspath(__file__))  # En calendula /LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1/cartolidar/cartolidar/clidax
# ==============================================================================
# Unidad de disco si MAIN_ENTORNO = 'windows'
MAIN_DRIVE = os.path.splitdrive(MAIN_FILE_DIR)[0]  # 'D:' o 'C:'
# ==============================================================================
if MAIN_FILE_DIR[:12] == '/LUSTRE/HOME':
    MAIN_ENTORNO = 'calendula'
    MAIN_PC = 'calendula'
elif MAIN_FILE_DIR[:8] == '/content':
    MAIN_ENTORNO = 'colab'
    MAIN_PC = 'colab'
else:
    MAIN_ENTORNO = 'windows'
    try:
        if 'benmarjo' in MAIN_HOME_DIR:
            MAIN_PC = 'JCyL'
        elif 'joseb' in MAIN_HOME_DIR:
            MAIN_PC = 'Casa'
        else:
            MAIN_PC = 'Otro'
    except:
        MAIN_ENTORNO = 'calendula'
        MAIN_PC = 'calendula'
# ==============================================================================

# ==============================================================================
# #variablesGlobalesMAINdirs
# ==============================================================================
# Directorios de la aplicacion:
# Dir de clidbase.py:
MAIN_BASE_DIR = os.path.abspath(os.path.join(MAIN_FILE_DIR, '..'))
# Dir de setup.py:
MAIN_PROJ_DIR = os.path.abspath(os.path.join(MAIN_BASE_DIR, '..')) # Equivale a MAIN_FILE_DIR = pathlib.Path(__file__).parent
# Dir en el que esta el proyecto (D:/_clid o /LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1):
MAIN_RAIZ_DIR = os.path.abspath(os.path.join(MAIN_PROJ_DIR, '..'))
# Directorio desde el que se lanza la app (estos dos coinciden):
MAIN_THIS_DIR = os.path.abspath('.')
MAIN_WORK_DIR = os.getcwd()
# Directorio del ejecutable de python:
MAIN_PYTHON_DIR = os.path.dirname(sys.executable)
# ==============================================================================

# ==============================================================================
# Estos parametros tb se asignan en clidbase.py y guardan en el cfg, pero
# clidbase.py importa clidaux antes, x lo q pasa antes x aqui y tb lo hago aqui.
if MAIN_ENTORNO == 'calendula':
    # MAINrutaRaizHome = MAIN_RAIZ_DIR
    MAINrutaRaizHome =  '/LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1'  # = MAIN_RAIZ_DIR
    MAINrutaRaizData = '/scratch/jcyl_spi_1/jcyl_spi_1_1'
else:
    MAINrutaRaizHome = MAIN_RAIZ_DIR
    MAINrutaRaizData = MAIN_RAIZ_DIR
# ==============================================================================

# ==============================================================================
# #variablesGlobalesVERSIONFILE
# ==============================================================================
# Ver https://peps.python.org/pep-0008/#module-level-dunder-names
# Ver https://stackoverflow.com/questions/458550/standard-way-to-embed-version-into-python-package
# Lectura de VERSIONFILE en clidbase.py, clidflow.py, qlidtwins.py,
#     clidax.config.py, clidax.clidaux.py, clidfr.clidhead.py, clidtools.__init__.py 
# ==============================================================================
VERSIONFILE = os.path.abspath(os.path.join(MAIN_FILE_DIR, '_version.py'))
if not os.path.exists(VERSIONFILE):
    VERSIONFILE = os.path.abspath(os.path.join(MAIN_FILE_DIR, '..', '_version.py'))
    if not os.path.exists(VERSIONFILE):
        VERSIONFILE = os.path.abspath(os.path.join(MAIN_FILE_DIR, '../..', '_version.py'))
        if not os.path.exists(VERSIONFILE):
            VERSIONFILE = os.path.abspath('_version.py')
if os.path.exists(VERSIONFILE):
    try:
        verstrline = open(VERSIONFILE, "rt").read()
        VSRE = r"^__version__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            # __version__ = mo.groups()[0]
            __version__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __version__ = "a.b.c"')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __version__ = "a.b.c"\n')
            __version__ = '0.0a0'
        VSRE = r"^__date__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            __date__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __date__ = "year1-year2"')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __date__ = "year1-year2"\n')
            __date__ = '2016-2022'
        VSRE = r"^__updated__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            __updated__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __updated__ = "date"')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __updated__ = "date"\n')
            __updated__ = '2022-07-01'
        VSRE = r"^__copyright__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            __copyright__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __copyright__ = "..."')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __copyright__ = "..."')
            __copyright__ = '@clid 2016-22'
    except:
        sys.stderr.write(f'clidtools.__init__-> no se ha podido leer {VERSIONFILE}\n')
        __version__ = '0.0a0'
        __date__ = '2016-2022'
        __updated__ = '2022-07-01'
        __copyright__ = '@clid 2016-22'
else:
    __version__ = '0.0a0'
    __date__ = '2016-2022'
    __updated__ = '2022-07-01'
    __copyright__ = '@clid 2016-22'
# ==============================================================================

# ==============================================================================
# ============================== Variables TRNS ================================
# ==============================================================================
# ================== TRNS: Variables globales transitorias =====================
# ======= Si las quiero usar en todos los modulos, pasarlas a Globales =========
TRNSpreguntarMemoria = False
# ==============================================================================


# ==============================================================================
def normalize(c):
    return unicodedata.normalize("NFD", c)[0]


# ==============================================================================
# Duplico esta funcion de clidaux para no importar clidaux
def infoUsuario(verbose=False):
    if psutilOk:
        try:
            esteUsuario = psutil.users()[0].name
            if verbose:
                print('clidconfig-> Usuario:', esteUsuario)
        except:
            esteUsuario = psutil.users()
            if verbose:
                print('clidconfig-> Users:', esteUsuario)
        if not isinstance(esteUsuario, str) or esteUsuario == '':
            esteUsuario = 'local'
    else:
        esteUsuario = 'SinUsuario'
    return esteUsuario


# ==============================================================================
# Duplico esta funcion de clidaux para no importar clidaux
def showCallingModules(inspect_stack=inspect.stack(), verbose=False):
    # print('->->->inspect_stack  ', inspect_stack
    # print('->->->inspect.stack()', inspect.stack())
    if len(inspect_stack) > 1:
        try:
            esteModuloFile0 = inspect_stack[0][1]
            esteModuloNum0 = inspect_stack[0][2]
            esteModuloFile1 = inspect_stack[1][1]
            esteModuloNum1 = inspect_stack[1][2]
            esteModuloName0 = inspect.getmodulename(esteModuloFile0)
            esteModuloName1 = inspect.getmodulename(esteModuloFile1)
        except:
            print(f'{TB}clidconfig-> Error identificando el modulo 1')
            return 'desconocido1', 'desconocido1'
    else:
        if verbose:
            print(f'{TB}clidconfig-> No hay modulos que identificar')
        return 'noHayModuloPrevio', 'esteModulo'

    if not esteModuloName0 is None:
        esteModuloName = esteModuloName0
        esteModuloNum = esteModuloNum0
        stackSiguiente = 1
    else:
        esteModuloName = esteModuloName1
        esteModuloNum = esteModuloNum1
        stackSiguiente = 2

    callingModulePrevio = ''
    callingModuleInicial = ''
    if verbose:
        print(f'{TB}clidconfig-> El modulo {esteModuloName} ({esteModuloNum}) ha sido', end=' ')
    for llamada in inspect_stack[stackSiguiente:]:
        if 'cartolid' in llamada[1] or 'clid' in llamada[1] or 'qlid' in llamada[1]:
            callingModule = inspect.getmodulename(llamada[1])
            if callingModule != esteModuloName and callingModulePrevio == '':
                callingModulePrevio = callingModule
            callingModuleInicial = callingModule
            if verbose:
                print(f'importado desde: {callingModule} ({llamada[2]})', end='; ')
    if verbose:
        print()
    return callingModulePrevio, callingModuleInicial


# ==============================================================================
if CONFIGverbose:
    sys.stdout.write(f'\n{" log - config ":_^80}\n')
    sys.stdout.write(f'clidconfig-> Cargando clidconfig...\n')
    sys.stdout.write(f'{TB}-> Directorio desde el que se lanza la aplicacion-> os.getcwd(): {MAIN_WORK_DIR}\n')
    sys.stdout.write(f'{TB}-> Revisando la pila de llamadas...\n')
callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
if callingModulePrevio == 'clidbase' or callingModuleInicial == 'clidclas' or callingModuleInicial == 'clidtools': # or callingModuleInicial == 'clidtry':
    # Solo se muestra el contenido de configVarsDict la primera vez que
    # se importa clidconfig.py (normalmente desde clidbase.py) y no cuando
    # se importa desde el resto de modulos (controlado con self.LCLverbose).
    LCLverbose = True
    if __verbose__ or CONFIGverbose:
        sys.stdout.write(f'\n{"":_^80}\n')
        sys.stdout.write(f'clidconfig-> Este modulo se importa desde todos los modulos,\n')
        sys.stdout.write(f'{TB}pero solo se muestra la carga en pantalla cuando:\n')
        sys.stdout.write(f'{TB}-> Se importa desde clidbase, clidclas o clidtools.\n')
        sys.stdout.write(f'clidconfig-> Modulo desde el que se importa clidconfig: {callingModulePrevio}\n')
        sys.stdout.write(f'{TB}-> El modulo lanzado inicialmente es {callingModuleInicial}.\n')
        sys.stdout.write(f'{"":=^80}\n')
else:
    __verbose__ = 0
    LCLverbose = False
# ==============================================================================

if (
    False
    # or callingModuleInicial == 'clidtools'
    # or callingModuleInicial == 'clidclas'
):
    usarXLS = False
else:
    # callingModuleInicial == 'clidbase'
    # or callingModuleInicial == 'clidflow'
    # or callingModuleInicial == 'clidtry'
    # or callingModuleInicial == 'clidgis'
    usarXLS = True

# ==============================================================================
if callingModuleInicial == 'clidflow':
    printMsgToFile = False
else:
    printMsgToFile = True
# ==============================================================================o


# ==============================================================================
def iniciaConsLog(myModule='clidconfig', myVerbose=False, myQuiet=False):
    if myVerbose == 3:
        logLevel = logging.DEBUG  # 10
    elif myVerbose == 2:
        logLevel = logging.DEBUG  # 10
    elif myVerbose == 1:
        logLevel = logging.INFO  # 20
    elif myVerbose == 0:
        if myQuiet:
            # logLevel = logging.CRITICAL
            logLevel = logging.ERROR  # 40
        else:
            # logLevel = logging.WARNING  # 30
            logLevel = logging.INFO  # 20
    # ==============================================================================
    # class ContextFilter(logging.Filter):
    #     """
    #     This is a filter which injects contextual information into the log.
    #     """
    #
    #     def filter(self, record):
    #         record.thisUser = myUser
    #         record.thisFile = myModule[:10]
    #         return True
    # myFilter = ContextFilter()
    # ==============================================================================
    # formatter1 = '{asctime}|{name:10s}|{levelname:8s}|{thisUser:8s}|> {message}'
    # formatterFile = logging.Formatter(formatter1, style='{', datefmt='%d-%m-%y %H:%M:%S')
    formatterCons = logging.Formatter('{message}', style='{')
    
    myLog = logging.getLogger(myModule)
    if sys.argv[0].endswith('__main__.py') and 'cartolidar' in sys.argv[0]:
        # qlidtwins.py se ejecuta lanzando el paquete cartolidar desde linea de comandos:
        #  python -m cartolidar
        # En __main__.py ya se ha confiigurado el logging.basicConfig()
        # if myModule == __name__.split('.')[-1]:
        #     print(f'{myModule}-> En __main.py se va a crear el loggin de consola para todos los modulos en __main__.py')
        # else:
        #     print(f'{myModule}-> Ya se ha creado el loggin de consola para todos los modulos en __main__.py')
        pass
    consLog = logging.StreamHandler(stream=sys.stdout)
    consLog.setFormatter(formatterCons)
    consLog.setLevel(logLevel)
    myLog.setLevel(logLevel)
    myLog.addHandler(consLog)
    return myLog


# ==============================================================================
myUser = infoUsuario()
# myModule = __name__.split('.')[-1]
# if not moduloPreviamenteCargado or True:
#     # print('\nclidconfig-> Aviso: creando myLog (ConsLog)')
#     myLog = iniciaConsLog(myModule=myModule, myVerbose=__verbose__)
# ==============================================================================
if CONFIGverbose:
    sys.stdout.write(f'{"":_^80}\n')
    sys.stdout.write(f'clidconfig-> Debug & alpha version info:\n')
    sys.stdout.write(f'{TB}-> ENTORNO:          {MAIN_ENTORNO}\n')
    sys.stdout.write(f'{TB}-> Modulo principal: <{sys.argv[0]}>\n') # = __file__
    sys.stdout.write(f'{TB}-> __package__ :     <{__package__ }>\n')
    sys.stdout.write(f'{TB}-> __name__:         <{__name__}>\n')
    sys.stdout.write(f'{TB}-> __verbose__:      <{__verbose__}>\n')
    sys.stdout.write(f'{TB}-> IdProceso         <{MAIN_idProceso:006}>\n')
    # sys.stdout.write(f'{TB}-> configFile:       <{GLO.configFileNameCfg}>\n')
    sys.stdout.write(f'{TB}-> sys.argv:         <{sys.argv}>\n')
    sys.stdout.write(f'{"":=^80}\n')
# ==============================================================================


# ==============================================================================
def creaLog(consLogYaCreado=False, myModule='module', myUser=myUser, myPath='.', myVerbose=0, myVerboseFile=0, myQuiet=False):
    myVerboseCons = myVerbose

    if myVerboseFile == 3:
        logLevelFile = logging.DEBUG
    elif myVerboseFile == 2:
        logLevelFile = logging.INFO  # 20
    elif myVerboseFile == 1:
        logLevelFile = logging.INFO  # 20
    elif myVerboseFile == 0:
        if myQuiet:
            # logLevelFile = logging.CRITICAL
            # logLevelFile = logging.ERROR  # 40
            logLevelFile = logging.WARNING  # 30
        else:
            logLevelFile = logging.INFO  # 20

    if myVerboseCons == 3:
        logLevelCons = logging.DEBUG  # 10
    elif myVerboseCons == 2:
        logLevelCons = logging.INFO  # 20
    elif myVerboseCons == 1:
        if myQuiet:
            logLevelCons = logging.WARNING  # 30
        else:
            logLevelCons = logging.INFO  # 20
    elif myVerboseCons == 0:
        if myQuiet:
            # logLevelCons = logging.CRITICAL
            logLevelCons = logging.ERROR  # 40
        else:
            # logLevelCons = logging.WARNING  # 30
            logLevelCons = logging.INFO  # 20

    if myPath == '':
        myLogPath = myPath
    else:
        myLogPath = os.path.abspath(myPath)
        if not os.path.isdir(myLogPath):
            try:
                os.makedirs(myLogPath)
            except:
                print(f'No se ha podido crear el directorio {myLogPath}. Revisar derechos de escritura en esa ruta.')
                sys.exit(0)
    if 'GLBLlogFile' in dir(GLO) and GLO.GLBLlogFile:
        mainLogFile = os.path.join(myLogPath, f'clidbase.log')
        thisLogFile = os.path.join(myLogPath, f'{myModule}.log')
        # print(f'clidconfig-> mainLogFile: {mainLogFile}')
        # print(f'clidconfig-> thisLogFile: {thisLogFile}')
        try:
            if not os.path.exists(mainLogFile):
                controlConfigFile = open(mainLogFile, mode='w')
                controlConfigFile.close()
                os.remove(mainLogFile)
            else:
                controlConfigFile = open(mainLogFile, mode='r+')
                controlConfigFile.close()
        except:
            callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
            print(f'\n{"":_^80}')
            print(f'clidconfig-> AVISO:')
            print(f'{TB}-> No se puede guardar el fichero log:  {mainLogFile}')
            print(f'{TB}-> Es posible que no tenga permisos de escritura en: {myLogPath}')
            print(f'{TB}-> O que exista el fichero {mainLogFile} y este bloqueado.')
            try:
                MAIN_HOME_DIR = str(pathlib.Path.home())
                MAIN_LOG_DIR = os.path.join(MAIN_HOME_DIR, 'Documents')
                print(f'\tSe intenta la ruta alternativa: {MAIN_LOG_DIR}')
                mainLogFile = os.path.join(MAIN_LOG_DIR, f'clidbase.log')
                thisLogFile = os.path.join(MAIN_LOG_DIR, f'{myModule}.log')
                if not os.path.exists(mainLogFile):
                    controlConfigFile = open(mainLogFile, mode='w')
                    controlConfigFile.close()
                    os.remove(mainLogFile)
                else:
                    controlConfigFile = open(mainLogFile, mode='r+')
                    controlConfigFile.close()
                print(f'\tOk log file1: {mainLogFile}')
                print(f'\tOk log file2: {thisLogFile}')
                myLogPath = MAIN_LOG_DIR
            except:
                print(f'\tTampoco se puede escribir en la ruta {MAIN_LOG_DIR}')
                MAIN_HOME_DIR = str(pathlib.Path.home())
                mainLogFile = os.path.join(MAIN_HOME_DIR, f'clidbase.log')
                thisLogFile = os.path.join(MAIN_HOME_DIR, f'{myModule}.log')
                print(f'\tSe opta por log file1: {mainLogFile}')
                print(f'\tlog file2: {thisLogFile}')
                myLogPath = MAIN_HOME_DIR
            print(f'{TB}-> callingModulePrevio: {callingModulePrevio}')
            print(f'{"":=^80}')

    # return (myLogPath, mainLogFile, thisLogFile)


    # ==============================================================================
    class ContextFilter(logging.Filter):
        """
        This is a filter which injects contextual information into the log.
        """

        def filter(self, record):
            record.thisUser = myUser
            record.thisFile = os.path.join(myLogPath, myModule[:10])
            return True
    # ==============================================================================
    myFilter = ContextFilter()
    # ==============================================================================
    # formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    # formatter = logging.Formatter('%(name)-12s: %(message)s')
    # formatter = logging.Formatter('{name:16s}: {levelname:8s} {message}', style='{')
    # formatter3 = '%(asctime)s|%(process)d|%(thisFile)-10s|%(levelname)-8s|%(thisUser)-8s|>%(message)s' # tyle='%'
    
    formatter0 = '{message}'
    formatter1 = '{asctime}|{name:10s}|{levelname:8s}|> {message}'
    formatter2 = '{asctime}|{name:10s}|{levelname:8s}|{thisUser:8s}|> {message}'
    formatterFile = logging.Formatter(formatter2, style='{', datefmt='%d-%m-%y %H:%M:%S')
    formatterCons = logging.Formatter(formatter0, style='{')

    if consLogYaCreado:
        # qlidtwins.py se ejecuta lanzando el paquete cartolidar desde linea de comandos:
        #  python -m cartolidar
        # En __main__.py ya se ha confiigurado el logging.basicConfig()
        # print(f'clidconfig-> {myModule}-> Ya se ha creado el loggin de consola para todos los modulos en __main__.py')
        pass
    else:
        # print(f'clidconfig-> {myModule}-> Iniciando logging.basicConfig<> para fileLog & consLog')
        # Este logger se mantiene activo para todos los modulos,
        # pero no puedo utilizar myFilter con todos los modulos porque algunas librerias (matplotlib)
        # lanzan mensajes de debug y no tienen definido el filtro, con lo que da error
        # https://docs.python.org/3/library/logging.html#logging.basicConfig
        if 'GLBLlogFile' in dir(GLO) and GLO.GLBLlogFile:
            logging.basicConfig(
                filename=mainLogFile,
                filemode='w',
                format=formatter1,
                style='{',
                datefmt='%d-%m-%y %H:%M:%S',
                # datefmt='%d-%b-%y %H:%M:%S',
                level=logLevelFile,
                # level=logging.DEBUG,
                # level=logging.INFO,
                # level=logging.WARNING,
                # level=logging.ERROR,
                # level=logging.CRITICAL,
            )
        else:
            logging.basicConfig(
                format=formatter0,
                style='{',
                datefmt='%d-%m-%y %H:%M:%S',
                # datefmt='%d-%b-%y %H:%M:%S',
                level=logLevelCons,
                # level=logging.DEBUG,
                # level=logging.INFO,
                # level=logging.WARNING,
                # level=logging.ERROR,
                # level=logging.CRITICAL,
            )
    myLog = logging.getLogger(myModule)
    # myLog.setLevel(logging.DEBUG)
    myLog.addFilter(myFilter)

    if 'GLBLlogFile' in dir(GLO) and GLO.GLBLlogFile:
        # Este logger solo actua para este modulo:
        # https://docs.python.org/3/library/logging.handlers.html#filehandler
        fileLog = logging.FileHandler(thisLogFile, mode='w')
        # fileLog.terminator = ''
        fileLog.set_name(myModule)
        fileLog.setLevel(logLevelFile)
        fileLog.setFormatter(formatterFile)
        fileLog.addFilter(myFilter)
        # logging.getLogger().addHandler(fileLog)
        myLog.addHandler(fileLog)
    if not consLogYaCreado:
        # https://docs.python.org/3/library/logging.handlers.html#logging.StreamHandler
        consLog = logging.StreamHandler(stream=sys.stdout)
        # https://docs.python.org/3/library/logging.handlers.html#logging.StreamHandler.terminator
        # consLog.terminator = ''  # Sustituye al valor por defecto que es '\n'
        consLog.setFormatter(formatterCons)
        consLog.setLevel(logLevelCons)
        # logging.getLogger().addHandler(consLog)
        myLog.addHandler(consLog)
    # ==============================================================================
    # myLog.debug('clidconfig-> debug')
    # myLog.info('clidconfig-> info')
    # myLog.warning('clidconfig-> warning')
    # myLog.error('clidconfig-> error')
    # myLog.critical('clidconfig-> critical')
    # ==============================================================================

    return myLog


# # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# def quitarContrabarrasAgregarBarraFinal(ruta=''):
#     if not ruta:
#         return None
#     nuevaRuta = ''
#     for letra in ruta:
#         letraSinBackslash = letra if letra != '\\' else '/'
#         nuevaRuta += letraSinBackslash
#     if nuevaRuta[-1:] != '/':
#         nuevaRuta += '/'
#     return nuevaRuta


# ooooooooooo Variable gobal precalculada: GLBLminimoDeMemoriaRAM ooooooooooooooo
def asignarMinimoDeMemoriaRAM(GLBLminimoDeMemoriaRAM):
    if platform.processor()[:3] == 'x86':
        minimoDeMemoriaRAM_recomendada = 0
    else:
        minimoDeMemoriaRAM_recomendada = GLBLminimoDeMemoriaRAM
    if TRNSpreguntarMemoria:
        print('Valor minimo de RAM para interrumpir lectura MB)?')
        print('Sugerencia: Para PC de 64 bits: 100 MB; para PC de 32 bits, 0 MB')
        selec = input(f'{TB}clidconfig-> Pulsa un numero o [enter] directamente para valor por defecto ({minimoDeMemoriaRAM_recomendada} MB) -> ')
        try:
            minimoDeMemoriaRAM = int(selec)
        except:
            minimoDeMemoriaRAM = minimoDeMemoriaRAM_recomendada
        print('    -> La lectura se interrumpe cuando la RAM disponible baja de %i MB' % minimoDeMemoriaRAM)
    else:
        minimoDeMemoriaRAM = minimoDeMemoriaRAM_recomendada
    return minimoDeMemoriaRAM


# ==============================================================================
def leerTablaDBF(elArchivo, GLOBAL_seleccionadosParaRepetir='', GLOBAL_listaRepetir=[]):
    # miTabla = dbf_old.Table(elArchivo)
    # miTabla = dbf.Table(elArchivo)
    if not dbfreadDisponible:
        return 0, None
    miTabla = dbfread.DBF(elArchivo)
    print('\nListado de campos del fichero dbf:')
    for miCampo in miTabla.fields:
        print(miCampo.name, miCampo.type, miCampo.length, miCampo.decimal_count)
    usarPropiedad_meta = False  # No funciona
    if usarPropiedad_meta:
        print('\nListado de campos (con _meta):')
        for campo in miTabla._meta.keys():
            if miTabla._meta[campo][0] == 'N':
                decimales = '; dec: ' + str(miTabla._meta[campo][4])
            else:
                decimales = ''
            print(campo.ljust(10), 'Tipo: ' + miTabla._meta[campo][0] + '; ancho:', str(miTabla._meta[campo][2]) + decimales)
    usarPropiedad_field_layout = False  # No funciona
    if usarPropiedad_field_layout:
        print('\nListado de campos (con _field_layout):')
        for i in range(len(miTabla.field_names)):
            print('miTabla._field_layout({:1}):'.format(i), miTabla._field_layout(i))

    print('Nombre del fichero:', elArchivo)
    print('Datos generales de', elArchivo + ':', miTabla)
    # Este es el string method de la "class Table(_Navigation):"= print( miTabla )
    print('Datos individuales de', elArchivo + ':')
    print('Nombre:              ', miTabla.filename)
    # print( 'Version:             ', miTabla.version ) # No en dfbread
    print('Version:             ', miTabla.dbversion, '=', miTabla.header.dbversion)  # Especifica de dbfread
    # print( 'Ult actualizacion:   ', miTabla.last_update ) # No en dfbread
    print('Ult actualizacion:   ', miTabla.date, '=', miTabla.header.year, miTabla.header.month, miTabla.header.day)  # Especifica de dbfread
    # print( 'codepage:            ', miTabla.codepage ) # No en dfbread
    print('codepage:            ', miTabla.encoding)  # Especifica de dbfread
    # Para que el texto con acentos en el dbf no de error establecer el encoding a 'cp1252' o 'latin-1':
    if miTabla.encoding == 'ascii':
        miTabla.encoding = 'cp1252'
    # print( 'Num de campos:       ', miTabla.field_count, '=', len(miTabla.field_names) ) # No en dfbread
    print('Num de campos:       ', len(miTabla.field_names))
    # print( 'Num de registros:    ', miTabla.__len__(), '=', len(miTabla) ) # No en dfbread
    print('Num de registros:    ', miTabla.header.numrecords, '(', len(miTabla.records), ')')
    # print( 'Ancho de cabecera:   ', miTabla.record_length ) # No en dfbread
    print('Ancho de registro:   ', miTabla.header.recordlen)  # Especifica de dbfread
    print('Ancho de cabecera:   ', miTabla.header.headerlen)  # Especifica de dbfread
    print('Campos (field_names):', miTabla.field_names)  # -> List

    # print( 'Registros:     ', type(miTabla.records))
    nContador = 0
    for registro in miTabla.records:
        nContador += 1
        # print( 'Registros - tipo:', type(registro), dir(registro))
        # print( 'Registros - claves:', registro.keys())
        # print( 'Registro', nContador, '->', registro )
        print('Registro', nContador, '-> Fichero lidar:', registro['Lidar'])
        if nContador > 5:
            print('clidconfig-> Etc.')
            break

    # miTabla.open() # No en dfbread
    # print('clidconfig-> miTabla', type(miTabla), dir(miTabla))
    # Metodos de miTabla: 'add_fields', 'allow_nulls', 'append', 'backup', 'bof', 'bottom', 'close', 'codepage', 'create_backup', 'create_index', 'create_template', 'current', 'current_record', 'delete_fields', 'disallow_nulls', 'eof', 'field_count', 'field_info', 'field_names', 'filename', 'first_record', 'goto', 'index', 'last_record', 'last_update', 'memoname', 'new', 'next_record', 'nullable_field', 'open', 'pack', 'prev_record', 'query', 'record_length', 'reindex', 'rename_field', 'resize_field', 'skip', 'status', 'structure', 'supported_tables', 'top', 'version', 'zap']
    for miRegistr in miTabla:
        # print('clidconfig-> miRegistr', type(miRegistr), dir(miRegistr))
        # Metodos de miRegistr <class 'dbf.Record'> ['__class__', '__contains__', '__delattr__', '__dir__', '__doc__', '__enter__', '__eq__', '__exit__', '__format__', '__ge__', '__getattr__', '__getattribute__', '__getitem__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__len__', '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__setattr__', '__setitem__', '__sizeof__', '__slots__', '__str__', '__subclasshook__', '__weakref__', '_commit_flux', '_create_blank_data', '_data', '_dirty', '_memos', '_meta', '_old_data', '_recnum', '_reindex_record', '_retrieve_field_value', '_rollback_flux', '_start_flux', '_update_disk', '_update_field_value', '_write', '_write_to_disk'
        # repetir = miRegistr._retrieve_field_value(0,'repetir')
        # repetir = miRegistr._retrieve_field_value('repetir') # No en dfbread
        repetir = miRegistr['Repetir']
        # nombreLas = miRegistr._retrieve_field_value(0,'Lidar')
        # nombreLas = miRegistr._retrieve_field_value('Lidar') # No en dfbread
        nombreLas = miRegistr['Lidar']
        if repetir >= GLOBAL_seleccionadosParaRepetir:
            GLOBAL_listaRepetir.append(nombreLas)
            # print( nombreLas, str(repetir) )
    return len(miTabla), GLOBAL_listaRepetir


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def mostrarVariablesGlobales(momento='inicial'):
    print('clidconfig-> Fichero actual (file):', __file__)
    print('\nVariables globales MAIN, TRNS y USER del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if variableGlobal[0:4] == 'MAIN' or variableGlobal[0:4] == 'TRNS' or variableGlobal[0:4] == 'USER':
            print('clidconfig-> {:>50}\t{}'.format(variableGlobal, globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nVariables globales GLBL del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if variableGlobal[0:4] == 'GLBL':
            print('clidconfig-> {:>50}\t{}'.format(variableGlobal, globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nVariables globales no GLBL, MAIN, TRNS o USER del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if (
            variableGlobal[0:3] != 'GLB'
            and variableGlobal[0:4] != 'MAIN'
            and variableGlobal[0:4] != 'TRNS'
            and variableGlobal[0:4] != 'USER'
            and variableGlobal[0:2] != '__'
            and not isinstance(globals()[variableGlobal], type)
        ):
            print('clidconfig-> {:>30}\t{}\t{}'.format(variableGlobal, type(globals()[variableGlobal]), globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nVariables globales __nativas__ del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if variableGlobal[0:2] == '__':
            print('clidconfig-> {:>15}\t{}\t{}'.format(variableGlobal, type(globals()[variableGlobal]), globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nModulos y clases del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if type(globals()[variableGlobal]) == types.ModuleType or isinstance(globals()[variableGlobal], type):
            print('clidconfig-> {:>12}\t{}\t{}'.format(variableGlobal, type(globals()[variableGlobal]), globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print(
        '\nVariables globales almacenadas en el momento {} en el array GLOBALconfigDict del modulo {}:'.format(
            momento, os.path.basename(__file__)
        )
    )
    for variableGlobalAlmacenada in globals()['GLOBALconfigDict']:
        print('clidconfig-> {:>50}\t{}'.format(variableGlobalAlmacenada, globals()['GLOBALconfigDict'][variableGlobalAlmacenada]))
        # print(f'{TB} {variableGlobalAlmacenada} {globals()["GLOBALconfigDict"][variableGlobalAlmacenada]}')


# ==============================================================================
# Una vez importado este clidconfig desde otro modulo,
# tengo varias alternativas para crear GLO en dichos modulos:
#    from cartolidar.clidax.clidconfig import GLOBALconfigDict
#    variable global -> GLOBALconfigDict['miVariable']
# o bien:
#    from cartolidar.clidax import clidconfig as CG
#    variable global -> CG.configVarsDict['miVariable']
#    #o bien:
#    variable global -> CG.GC.miVariable
#    #o bien:
#    VG = CG.GLO_CLASS()
#    variable global -> VG.miVariable
# o bien:
#    from cartolidar.clidax.clidconfig import GLO
#    variable global -> GLO.miVariable
# o bien:
#    GLOBALconfigDict = clidconfig.leerCambiarVariablesGlobales(
# Opto por esta ultima

# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def initConfigDicts(idProceso=MAIN_idProceso, LOCL_verbose=__verbose__):
    global GLOBALconfigDict

    # ==========================================================================
    # Opcion 1: lectura del config (cfg) generado 
    # en un carga anterior de clidconfig o en una ejecucion anterior
    configFileNameCfg = getConfigFileNameCfg(idProceso, LOCL_verbose=LOCL_verbose)

    if CONFIGverbose:
        print(f'\n{"":_^80}')
        print('clidconfig-> Ejecutando initConfigDicts<> para leer el fichero de configuracion cfg (si existe).')
        print(f'{TB}-> Verifico si hay un cfg con idProceso={idProceso} generado')
        print(f'{TB}   en una carga anterior de clidconfig o en una ejecucion anterior:')
        print(f'{TB}{TV}{configFileNameCfg}')
        print(f'{TB}-> En caso negativo uso clidbase.xlsx')
        print(f'{"":=^80}')

    if os.path.exists(configFileNameCfg):
        if CONFIGverbose:
            print(f'{TB}clidconfig-> Se intenta leer los parametros del cfg: {os.path.basename(configFileNameCfg)}')
        config = RawConfigParser()
        config.optionxform = str  # Avoid change to lowercase
        numObjetivosExtraMax = 0
        LOCALconfigDict = {}
        try:
            config.read(configFileNameCfg)
            for grupoParametroConfiguracion in config.sections():
                for nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion):
                    strParametroConfiguracion = config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion)
                    listaParametroConfiguracion = strParametroConfiguracion.split('|+|')
                    valorParametroConfiguracion = valorConfig(
                        listaParametroConfiguracion[0],
                        nombreParametro=nombreParametroDeConfiguracion,
                        tipoVariable=listaParametroConfiguracion[1]
                    )
                    if len(listaParametroConfiguracion) > 1:
                        tipoParametroConfiguracion = listaParametroConfiguracion[1]
                    else:
                        tipoParametroConfiguracion = 'str'
                    if len(listaParametroConfiguracion) > 2:
                        descripcionParametroConfiguracion = listaParametroConfiguracion[2]
                    else:
                        descripcionParametroConfiguracion = ''
                    if nombreParametroDeConfiguracion[:1] == '_':
                        grupoParametroConfiguracion_new = '_%s' % grupoParametroConfiguracion
                    else:
                        grupoParametroConfiguracion_new = grupoParametroConfiguracion
                    LOCALconfigDict[nombreParametroDeConfiguracion] = [
                        valorParametroConfiguracion,
                        grupoParametroConfiguracion_new,
                        descripcionParametroConfiguracion,
                        tipoParametroConfiguracion,
                    ]
                    numObjetivosExtra = max(len(listaParametroConfiguracion) - 3, 0)
                    numObjetivosExtraMax = max(numObjetivosExtra, numObjetivosExtraMax)
                    if numObjetivosExtraMax:
                        valObjetivosExtra = []
                        for nObjetivoExtra in range(numObjetivosExtraMax):
                            if 3 + nObjetivoExtra < len(listaParametroConfiguracion):
                                valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[3 + nObjetivoExtra], tipoVariable=listaParametroConfiguracion[1]))
                            else:
                                valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[0], tipoVariable=listaParametroConfiguracion[1]))
                        LOCALconfigDict[nombreParametroDeConfiguracion].extend(valObjetivosExtra)
    
            # Lectura de cfg SI ok:
            LOCALconfigDict['configFileNameCfg'] = [
                configFileNameCfg,
                'GrupoMAIN',
                'Nombre del fichero de configuracion de donde se lee la configuacion (este fichero).',
                'str',
            ]
            LOCALconfigDict['configFileNameCfg'].extend([configFileNameCfg] * numObjetivosExtra)
            LOCALconfigDict['configLeidoDelCfgOk'] = [
                True,
                'GrupoMAIN',
                'Los parametros de configuracion se leen del cfg en vez del xlsx.',
                'str',
            ]
            LOCALconfigDict['configLeidoDelCfgOk'].extend([True] * numObjetivosExtra)
            configLeidoDelCfgOk = True

            # Lectura de xlsx NO necesaria
            LOCALconfigDict['configFileNameXls'] = [
                '',
                'GrupoMAIN',
                'Nombre del fichero de configuracion xlsx de donde NO se ha leido la configuracion.',
                'str',
            ]
            LOCALconfigDict['configFileNameXls'].extend([''] * numObjetivosExtra)
            LOCALconfigDict['configLeidoDelXlsOk'] = [
                False,
                'GrupoMAIN',
                'Indica si los parametros de configuracion se han leido del xlsx por no haber cfg disponible.',
                'str',
            ]
            LOCALconfigDict['configLeidoDelXlsOk'].extend([False] * numObjetivosExtra)

            GLOBALconfigDict = LOCALconfigDict
        except:
            configLeidoDelCfgOk = False
    else:
        configLeidoDelCfgOk = False

    if configLeidoDelCfgOk:
        if LOCL_verbose:
            print(f'{TB}-> Parametros leidos ok del fichero cfg:')
            print(f'{TB}-> {GLOBALconfigDict["configFileNameCfg"][0]}')
            print(f'{"":=^80}')
        return GLOBALconfigDict
    else:
        if LOCL_verbose:
            print(f'{TB}-> No se han leido los parametros del fichero cfg:')
            print(f'{TB}-> {configFileNameCfg}')
            print(f'{TB}-> Se van a leer del fichero xlsx.')
            print(f'{"":=^80}')
    # ==========================================================================

    # ==========================================================================
    # Opcion 2: si no hay cfg generado en un carga anterior de clidconfig
    # o en una ejecucion anterior se busca el fichero de configuracion xlsx.
    (
        filenameXlsCfg1Clidbase,
        filenameXlsCwdClidbase,
        filenameXlsCwdModulo,
        filenameXlsCfg2Clidbase,
        filenameXlsProjClidbase,
        MAIN_CFG_DIR,
    ) = getConfigFileNameXls(
        configFileNameCfg,
        LOCL_verbose=LOCL_verbose,
    )

    directorioActual = (os.getcwd()).replace(os.sep, '/')
    directorioProyecto = os.path.dirname(sys.argv[0])
    moduloInicialXls = (os.path.basename(sys.argv[0])).replace('.py', '.xlsx')

    if CONFIGverbose:
        print('clidconfig-> Ejecutando initConfigDicts<> para leer el fichero de configuracion xlsx')
        print(f'{TB}-> Fichero cfg no disponible, se leen los parametros del xls:')
        # print(f'{TB}{TV}-> {filenameXlsCfg1Clidbase}')
        print(f'{TB}-> Aqui leo las columnas correspondientes a todos los objetivoEjecucion.')
        print(f'{TB}-> Mas adelante, segun el valor de MAINobjetivoEjecucion usare uno u otro valor.')

    if os.path.exists(filenameXlsCfg1Clidbase):
        filenameXLS = filenameXlsCfg1Clidbase
        baseNameXLS = os.path.basename(filenameXlsCfg1Clidbase)
        pathNameXls = os.path.dirname(filenameXlsCfg1Clidbase)
        if baseNameXLS == 'clidbase.xlsx':
            if (pathNameXls.replace('/','').replace('\\','')).endswith('cartolidar'):
                if CONFIGverbose:
                    print(f'clidconfig-> Se usa un fichero de configuracion principal: {baseNameXLS}')
                    print(f'{TB}-> Ubicado en la ruta base de cartolidar: {pathNameXls}')
            elif (directorioActual.replace('/','').replace('\\','')).endswith('datacfg'):
                if CONFIGverbose:
                    print(f'clidconfig-> Se usa un fichero de configuracion principal: {baseNameXLS}')
                    print(f'{TB}-> Ubicado en la ruta donde se ubican los ficheros cfg: {pathNameXls}')
            else:
                # Esto no esta previsto (en principio)
                if CONFIGverbose:
                    print(f'clidconfig-> Se usa un fichero de configuracion principal: {baseNameXLS}')
                    print(f'{TB}-> Ubicado en la ruta no prevista: {pathNameXls}')
        else:
            # Esto no esta previsto (en principio)
            if CONFIGverbose:
                print(f'clidconfig-> Se usa un fichero de configuracion especifico de modulo o directorio ({baseNameXLS})')
                print(f'{TB}-> Ubicado en la ruta: {pathNameXls}')
    elif os.path.exists(filenameXlsCwdClidbase):
        filenameXLS = filenameXlsCwdClidbase
        baseNameXLS = os.path.basename(filenameXlsCwdClidbase)
        pathNameXls = os.path.dirname(filenameXlsCwdClidbase)
        if baseNameXLS == 'clidbase.xlsx':
            if CONFIGverbose:
                print(f'clidconfig-> Se usa un fichero de configuracion principal: {baseNameXLS}')
                print(f'{TB}-> Ubicado en el directorio desde el que se lanza la aplicacion: {pathNameXls} = {directorioActual}')
        else:
            if CONFIGverbose:
                # filenameXLSModulo = os.path.basename(filenameXlsCwdModulo)
                # print(f'clidconfig-> Se usa un fichero de configuracion para el modulo inicial: {baseNameXLS} =? {filenameXLSModulo}')
                print(f'clidconfig-> Se usa un fichero de configuracion especifico de modulo o directorio: {baseNameXLS}')
                print(f'{TB}-> Ubicado en el directorio desde el que se lanza la aplicacion: {pathNameXls} = {directorioActual}')
    elif os.path.exists(filenameXlsCwdModulo):
        filenameXLS = filenameXlsCwdModulo
        if CONFIGverbose:
            print(
                'clidconfig-> Se usa un fichero de configuracion especifico de modulo ({}) o directorio ({}):'.format(
                    moduloInicialXls,
                    directorioActual,
                )
            )
            print(f'{TB}-> {filenameXlsCwdModulo}')
    elif os.path.exists(filenameXlsCfg2Clidbase):
        filenameXLS = filenameXlsCfg2Clidbase
        if CONFIGverbose:
            print(f'clidconfig-> Se usa un fichero de configuracion general ubicado en el directorio elegido finalmente para el cfg: {filenameXlsCfg2Clidbase}')
    elif os.path.exists(filenameXlsProjClidbase):
        filenameXLS = filenameXlsProjClidbase
        if CONFIGverbose:
            print(f'clidconfig-> Se usa un fichero de configuracion general ubicado en el directorio de la aplicacion cartolidar: {filenameXlsProjClidbase}')
    else:
        if 'cartolidar' in directorioActual:
            # Por si ejecuto directamente un modulo que no esta en cartolid/ 
            # sino en cartolid/package/ y no tengo un xml propio para ese package.
            if not (directorioActual.replace('/','')).endswith('cartolidar'):
                directorioActual = os.path.abspath(os.path.join(directorioActual, '..'))
        else:
            # Por si llamo a cartolid desde su directorio raiz (en calendula)
            directorioActual = os.path.join(directorioActual, 'cartolidar')
        filenameXLSotroDirCartolid = os.path.join(directorioActual, 'clidbase.xlsx')
        # filenameXML = os.path.join(directorioActual, 'clidbase.xml')
        if os.path.exists(filenameXLSotroDirCartolid):
            print(f'clidconfig-> Se usa un fichero de configuracion general del directorio: {directorioActual}')
            filenameXLS = filenameXLSotroDirCartolid
            print(f'{TB}-> Fichero de configuracion: {filenameXLS}')
        else:
            print(f'clidconfig-> ATENCION no se ha encontrando un fichero de configuracion. Se han intentado (por este orden):')
            print(f'{TB}-> {filenameXlsCfg1Clidbase}')
            print(f'{TB}-> {filenameXlsCwdClidbase}')
            print(f'{TB}-> {filenameXlsCwdModulo}')
            print(f'{TB}-> {filenameXlsCfg2Clidbase}')
            print(f'{TB}-> {filenameXlsProjClidbase}')
            sys.exit(0)

    # https://openpyxl.readthedocs.io/en/stable/
    # https://realpython.com/openpyxl-excel-spreadsheets-python/
    # https://www.pythonexcel.com/openpyxl.php
    if not usarXLS:
        GLOBALconfigDict = {}
        return GLOBALconfigDict

    try:
        if os.path.exists(filenameXLS):
            from openpyxl import load_workbook
            xlsOk = True
        else:
            print(f'clidconfig-> ATENCION: no se encuentra {filenameXLS}')
            xlsOk = False
    except:
        print('clidconfig-> ATENCION: error al importar openpyxl')
        xlsOk = False

    if not xlsOk:
        print(f'clidconfig-> Comprobar que {moduloInicialXls} existe y openpyxl instalado (para que {callingModuleInicial} pueda leerlo).')
        print(f'{TB}-> callingModulePrevio: {callingModulePrevio}')
        sys.exit(0)

    try:
        intIdProceso = int(idProceso)
    except:
        intIdProceso = 999999

    try:
        if len(sys.argv) == 0 or sys.argv[0] == '' or sys.argv[0] == '-m':
            # print('Se esta ejecutando fuera de un modulo, en el interprete interactivo')
            filenameXLSactualSinPath = 'clidbase888888.xlsx'
        else:
            # print('Se esta ejecutando desde un modulo')
            if intIdProceso:
                filenameXLSactualSinPath = os.path.basename(sys.argv[0]).replace('.py', '{:006}.xlsx'.format(intIdProceso))
            else:
                filenameXLSactualSinPath = os.path.basename(sys.argv[0]).replace('.py', '000000.xlsx')

        filenameXLSactualConPath = os.path.join(MAIN_CFG_DIR, filenameXLSactualSinPath)
        if CONFIGverbose:
            print(f'{TB}clidconfig-> Hago una copia del fichero de configuracion xlsx: {filenameXLSactualConPath}')
        shutil.copy(filenameXLS, filenameXLSactualConPath)
    except:
        print('clidconfig-> No se ha podido copiar {} a {} -> revisar.'.format(
            filenameXLS,
            filenameXLSactualConPath,
            )
        )

    myWorkbookOrigin = load_workbook(filename=filenameXLS, read_only=True, data_only=True)
    # myWorkbookActual = load_workbook(filename=filenameXLSactualConPath)
    hojas = myWorkbookOrigin.sheetnames
    nombreHojaCartolidConfig = 'cartolidConfig'
    if nombreHojaCartolidConfig in hojas:
        #hojaActiva = myWorkbookOrigin.active
        hojaCartolidConfigOrigin = myWorkbookOrigin[nombreHojaCartolidConfig]
        # hojaCartolidConfigActual = myWorkbookActual[nombreHojaCartolidConfig]
        verbose = False
        if verbose:
            print('clidconfig-> Mostrando variables globales del fichero de configuracion xlsx:')
            for nRow in range(1, 1000):
                if (
                    hojaCartolidConfigOrigin.cell(row=nRow, column=3).value is None
                    or str(hojaCartolidConfigOrigin.cell(row=nRow, column=3).value).startswith('#')
                    or str(hojaCartolidConfigOrigin.cell(row=nRow, column=3).value).startswith('_')
                ):
                    continue
                print('{:>50}: {:<30} ({})'.format(
                    str(hojaCartolidConfigOrigin.cell(row=nRow, column=3).value),
                    str(hojaCartolidConfigOrigin.cell(row=nRow, column=4).value),
                    str(type(hojaCartolidConfigOrigin.cell(row=nRow, column=4).value)),
                    )
                )
            print()

        configTextDict = {}
        # La funcion iter_rows() considera la columna A como 1, pero dentro de la lista generada (tupleRow), esa columna tiene el index 0.
        for nRow, tupleRow in enumerate(hojaCartolidConfigOrigin.iter_rows(min_col= 1, max_col=25, values_only=True)):
            try:
                nombreParametroDeConfiguracion = tupleRow[2]
                if (
                    nombreParametroDeConfiguracion is None or 
                    nombreParametroDeConfiguracion == 'name'
                ):
                    continue
                if (
                    not nombreParametroDeConfiguracion.startswith('MAIN')
                    and not nombreParametroDeConfiguracion.startswith('GLBL')
                    and not nombreParametroDeConfiguracion.startswith('#')
                ):
                    if verbose:
                        print(
                            'clidconfig-> ATENCION: se ignoran parametros que no empiecen por MAIN o GLBL: {}'.format(
                                nombreParametroDeConfiguracion
                            )
                        )
                    continue
                tipoVariable = tupleRow[5] # Columna F
                valorPrincipal = tupleRow[3] # Columna D
                valorAlternativo = valorPrincipal
                valorCREA_TILES = tupleRow[8] # Columna I
                if valorCREA_TILES is None or valorCREA_TILES == '':
                    valorCREA_TILES = valorPrincipal
                valorPREPROCESADO_EN_CALENDULA = tupleRow[9] # Columna J
                if valorPREPROCESADO_EN_CALENDULA is None or valorPREPROCESADO_EN_CALENDULA == '':
                    valorPREPROCESADO_EN_CALENDULA = valorPrincipal
                valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = tupleRow[10] # Columna K
                if valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ is None or valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ == '':
                    valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = valorPrincipal
                valorCREAR_LAZ = tupleRow[11] # Columna L
                if valorCREAR_LAZ is None or valorCREAR_LAZ == '':
                    valorCREAR_LAZ = valorPrincipal
                valorMALLA_DESFASADA = tupleRow[12] # Columna M
                if valorMALLA_DESFASADA is None or valorMALLA_DESFASADA == '':
                    valorMALLA_DESFASADA = valorPrincipal
                valorEXTRA3 = tupleRow[13] # Columna N
                if valorEXTRA3 is None or valorEXTRA3 == '':
                    valorEXTRA3 = valorPrincipal

                grupoParametros = tupleRow[1] # Columna B
                descripcionParametro = tupleRow[6]
                usoParametro = tupleRow[4]
                configTextDict[nombreParametroDeConfiguracion] = [
                    tipoVariable,
                    valorPrincipal,
                    valorAlternativo,
                    grupoParametros,
                    descripcionParametro,
                    usoParametro,
                    valorCREA_TILES,
                    valorPREPROCESADO_EN_CALENDULA,
                    valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ,
                    valorCREAR_LAZ,
                    valorMALLA_DESFASADA,
                    valorEXTRA3,
                ]
                # hojaCartolidConfigActual.cell(row=nRow+1, column=1).value = nRow
            except:
                print('clidconfig-> Ha habido un problema al leer la hoja', nombreHojaCartolidConfig, 'del fichero de configuracion:', filenameXLS)
                print(f'{TB}-> Fila: {nRow} Columnas leidas (tupleRow): {type(tupleRow)} {tupleRow}')
        # myWorkbookActual.save(filename=filenameXLSactualConPath)
    else:
        xlsOk = False
        print('clidconfig-> ATENCION: Revisar contenido del fichero de configuracion', filenameXLS)
        print(f'{TB}No tiene la hoja "cartolidConfig"')
        sys.exit(0)

    # Creo otro dict, pero:
    #    -> Solo con los valores de los parametros de configuracion
    #    -> Guardados en su formato especifico (str, int, float o bool)
    #    -> Eligiendo el valor principal o alternativo
    GLOBALconfigDict = {}
    if CONFIGverbose and __verbose__ > 3 and not moduloPreviamenteCargado:
        print('clidconfig-> Mostrando parametros de configuracion del xls:')
    for nombreParametroDeConfiguracion in configTextDict.keys():
        # configTextDict[nombreParametroDeConfiguracion] -> [tipoVariable, valorPrincipal, valorAlternativo, grupoParametros, descripcionParametro, usoParametro]
        tipoVariable = configTextDict[nombreParametroDeConfiguracion][0]
        valorPrincipal = configTextDict[nombreParametroDeConfiguracion][1]
        valorAlternativo = configTextDict[nombreParametroDeConfiguracion][2] # El valor alternativo lo he abandonada al pasar de xml a xlsx
        grupoParametros = configTextDict[nombreParametroDeConfiguracion][3]
        descripcionParametro = configTextDict[nombreParametroDeConfiguracion][4]
        usoParametro = configTextDict[nombreParametroDeConfiguracion][5]
        valorCREA_TILES = configTextDict[nombreParametroDeConfiguracion][6]
        valorPREPROCESADO_EN_CALENDULA = configTextDict[nombreParametroDeConfiguracion][7]
        valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = configTextDict[nombreParametroDeConfiguracion][8]
        valorCREAR_LAZ = configTextDict[nombreParametroDeConfiguracion][9]
        valorMALLA_DESFASADA = configTextDict[nombreParametroDeConfiguracion][10]
        valorEXTRA3 = configTextDict[nombreParametroDeConfiguracion][11]

        valorParametroDeConfiguracionPrincipal = valorConfig(
            valorPrincipal,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionCREA_TILES = valorConfig(
            valorCREA_TILES,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionPREPROCESADO_EN_CALENDULA = valorConfig(
            valorPREPROCESADO_EN_CALENDULA,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = valorConfig(
            valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionCREAR_LAZ = valorConfig(
            valorCREAR_LAZ,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionMALLA_DESFASADA = valorConfig(
            valorMALLA_DESFASADA,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionEXTRA3 = valorConfig(
            valorEXTRA3,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )

        GLOBALconfigDict[nombreParametroDeConfiguracion] = [
            valorParametroDeConfiguracionPrincipal,
            grupoParametros,
            descripcionParametro,
            tipoVariable,
            valorParametroDeConfiguracionCREA_TILES,
            valorParametroDeConfiguracionPREPROCESADO_EN_CALENDULA,
            valorParametroDeConfiguracionCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ,
            valorParametroDeConfiguracionCREAR_LAZ,
            valorParametroDeConfiguracionMALLA_DESFASADA,
            valorParametroDeConfiguracionEXTRA3,
        ]
        if CONFIGverbose and __verbose__ > 3 and not moduloPreviamenteCargado:
            print(f'{TB}{nombreParametroDeConfiguracion}-> {GLOBALconfigDict[nombreParametroDeConfiguracion]}')

    numObjetivosExtra = max(len(GLOBALconfigDict['GLBLverbose']) - 3, 0)
    # Lectura de cfg NO ok
    GLOBALconfigDict['configFileNameCfg'] = [
        configFileNameCfg,
        'GrupoMAIN',
        'Nombre del fichero de configuracion donde se guarda la configuracion temporal (se crea nuevo).',
        'str',
    ]
    GLOBALconfigDict['configFileNameCfg'].extend([configFileNameCfg] * numObjetivosExtra)
    GLOBALconfigDict['configLeidoDelCfgOk'] = [
        False,
        'GrupoMAIN',
        'Los parametros de configuracion se leen del cfg en vez del xlsx.',
        'str',
    ]
    GLOBALconfigDict['configLeidoDelCfgOk'].extend([False] * numObjetivosExtra)

    # Lectura de xlsx SI ok
    GLOBALconfigDict['configFileNameXls'] = [
        filenameXLS,
        'GrupoMAIN',
        'Nombre del fichero de configuracion xlsx de donde se ha leido la configuracion.',
        'str',
    ]
    GLOBALconfigDict['configFileNameXls'].extend([filenameXLS] * numObjetivosExtra)
    GLOBALconfigDict['configLeidoDelXlsOk'] = [
        True,
        'GrupoMAIN',
        'Indica si los parametros de configuracion se han leido del xlsx por no haber cfg disponible.',
        'str',
    ]
    GLOBALconfigDict['configLeidoDelXlsOk'].extend([True] * numObjetivosExtra)

    if LOCL_verbose:
        print(f'clidconfig-> Parametros leidos ok del fichero xlsx:')
        print(f'{TB}-> {GLOBALconfigDict["configFileNameXls"][0]}')
        print(f'{"":=^80}')

    return GLOBALconfigDict


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
class GLO_CLASS(object):
    def __init__(self, LOCALconfigDict={}, LCLverbose=__verbose__):
        """
        Clase cuyas propiedades son las variables de configuracion
        que utilizo en otros modulos como variables globales
        Utiliza un dict (LOCALconfigDict) con los nombres y valores de esas variables
        En definitiva, esta clase sirve para convertir valores de un dict
        en propiedades de un objeto
        Ademas crea algunas variables adicionales que se crean en tiempo de ejecucion:
            -> Variables derivadas
            -> Nombres de los ficheros de control
        #TODO: Lo siguiente seria guardar todas esas variables en un fichero de configuracion
        """

        # sys.argv tiene como primer valor el nombre del modulo que se ejecuta inicialmente sys.argv[0]
        # Ademas he agregado dos valores adicionales (van despues de los argumentos en linea de comandos):
        #  sys.argv.append('--idProceso')  -> sys.argv[-2]
        #  sys.argv.append(MAIN_idProceso) -> sys.argv[-1]

        # if len(sys.argv) > 3:
        #     self.ARGSobjetivoEjecucion = sys.argv[1]
        # else:
        #     self.ARGSobjetivoEjecucion = ''

        self.configVarsDict = LOCALconfigDict

        self.LCLverbose = LCLverbose
        if (CONFIGverbose or self.LCLverbose):
            if (
                callingModulePrevio == 'clidbase'
                and callingModuleInicial != 'runpy'
                and callingModuleInicial != '__init__'
                and callingModuleInicial != '__main__'
                and callingModuleInicial != 'clidtwins' and callingModuleInicial != 'qlidtwins'
                and callingModuleInicial != 'clidmerge' and callingModuleInicial != 'qlidmerge'
                and not callingModuleInicial.startswith('test_')
            ):
                print(f'\n{"":_^80}')
                print(f'clidconfig-> Creando la clase GLO_CLASS')
                print(f'{TB}-> Se asignan los parametros leidos del .cfg o .xlsx')
                print(f'{TB}{TV}como propiedades del objeto GLO (que es una instancia de la clase GLO_CLASS).')
                if 'MAINobjetivoEjecucion' in self.configVarsDict:
                    print(f'{TB}{TV}El parametro MAINobjetivoEjecucion ({self.configVarsDict["MAINobjetivoEjecucion"][0]})')
                else:
                    configFileNameCfg = getConfigFileNameCfg(MAIN_idProceso)
                    print(f'\nclidconfig-> ATENCION: revisar el fichero de configuracion ({configFileNameCfg})')
                    print(f'{TB}-> No tiene el parametro MAINobjetivoEjecucion')
                    sys.exit(0)
                print(f'{TB}{TV}se usa para seleccionar el valor correspondiente a ese self.numObjetivoEjecucion')
                print(f'{TB}-> Se genera un cfg que incluye todos los objetivos de ejecucion para poder seleccionar el que proceda.')
                print(f'{"":=^80}')

        if usarXLS:
            self.seleccionarColumnaCorrespondienteAObjetivoEjecucion()
        else:
            self.configVarsDict['GLBLverbose'] = [LCLverbose, 'Main', 'bool', 'Herencias de cartolidar']
            self.numObjetivoEjecucion = 0

        self.confirmarExistenciaValidezDeAlgunosParametrosDeConfiguracion()

        self.cargaVariablesDelConfigVarsDict()

        self.paramConfigAdicionalesGLBL = {}


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def seleccionarColumnaCorrespondienteAObjetivoEjecucion(self):
        if CONFIGverbose and self.LCLverbose and callingModulePrevio == 'clidbase':
            print(f'{TB}-> *Se elige la columna correspondiente al MAINobjetivoEjecucion: {self.configVarsDict["MAINobjetivoEjecucion"][0]}')
            print(f'{TB}-> *callingModulePrevio: {callingModulePrevio}; callingModuleInicial: {callingModuleInicial}')
            print(f'{"":=^80}')

        # Tengo objetivos de ejecucion preconfigurados, distinto del objetivo generico:
        if self.configVarsDict['MAINobjetivoEjecucion'][0] == 'GENERAL':
            # Usa la columna D del fichero de configuracion
            self.numObjetivoEjecucion = 0
        elif self.configVarsDict['MAINobjetivoEjecucion'][0] == 'CREAR_TILES_TRAIN':
            # Para crear tiles destinados al entrenamiento convolucional
            self.numObjetivoEjecucion = 4
        elif self.configVarsDict['MAINobjetivoEjecucion'][0] == 'PREPROCESADO_EN_CALENDULA':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            self.numObjetivoEjecucion = 5
        elif self.configVarsDict['MAINobjetivoEjecucion'][0] == 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            self.numObjetivoEjecucion = 6
        elif self.configVarsDict['MAINobjetivoEjecucion'][0] == 'CREAR_LAZ':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            self.numObjetivoEjecucion = 7
        elif self.configVarsDict['MAINobjetivoEjecucion'][0] == 'MALLA_DESFASADA':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            self.numObjetivoEjecucion = 8
        elif self.configVarsDict['MAINobjetivoEjecucion'][0] == 'EXTRA3':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            self.numObjetivoEjecucion = 9
        else:
            # Objetivos de ejecucion general:
            self.numObjetivoEjecucion = 0


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def confirmarExistenciaValidezDeAlgunosParametrosDeConfiguracion(self):
        # Por si acaso...
        # GLBLmostrarVariablesDeConfiguracion ==================================
        if not 'GLBLmostrarVariablesDeConfiguracion' in self.configVarsDict.keys():
            self.configVarsDict['GLBLmostrarVariablesDeConfiguracion'] = [self.configVarsDict['GLBLverbose'][0], 'Main', 'str', 'Herencias de cartolidar']
            print(f'\n{"":_^80}')
            print(f'clidconfig-> ATENCION: falta el parametro GLBLmostrarVariablesDeConfiguracion')
            print(f'{TB}Se asigna el mismo valor que GLBLverbose')
            print(f'{"":=^80}')
        # ======================================================================

        # MAINobjetivoEjecucion ================================================
        if not 'MAINobjetivoEjecucion' in self.configVarsDict.keys():
            self.configVarsDict['MAINobjetivoEjecucion'] = [
                'GENERAL', 'Main', 'str', 'Herencias de cartolidar', 'GENERAL',
                'GENERAL', 'GENERAL', 'GENERAL',
                'GENERAL', 'GENERAL', 'GENERAL',
            ]
            print(f'\n{"":_^80}')
            print(f'clidconfig-> ATENCION: falta el parametro MAINobjetivoEjecucion')
            print(f'{TB}Se asigna el valor: GENERAL.')
            print(f'{"":=^80}')
        self.listaObjetivosEjecucionSiReglados = (
            'GENERAL',
            'CREAR_TILES_TRAIN',
            'PREPROCESADO_EN_CALENDULA',
            'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ',
            'CREAR_PUNTOS_TRAIN_ROQUEDOS',
            'CREAR_LAZ',
            'MALLA_DESFASADA',
            'EXTRA3',
        )
        self.listaObjetivosEjecucionNoReglados = (
            'CREAR_PUNTOS_TRAIN_ROQUEDOS'
        )
        if (
            not self.configVarsDict['MAINobjetivoEjecucion'][self.numObjetivoEjecucion] in self.listaObjetivosEjecucionSiReglados
            and not self.configVarsDict['MAINobjetivoEjecucion'][self.numObjetivoEjecucion] in self.listaObjetivosEjecucionNoReglados
        ):
            print(f'\n{"":_^80}')
            print(f'clidconfig-> ATENCION: MAINobjetivoEjecucion no contemplado: {self.configVarsDict["MAINobjetivoEjecucion"]}')
            print(f'{TB}-> Lista de MAINobjetivoEjecucion admitidos:')
            print(f'{TB}{TV}-> listaObjetivosEjecucionSiReglados: {self.listaObjetivosEjecucionSiReglados}')
            print(f'{TB}{TV}-> listaObjetivosEjecucionNoReglados: {self.listaObjetivosEjecucionSiReglados}')
            print(f'{TB}-> Corregirlo en el fichero de configuracion .cfg o .xlsx')
            sys.exit(0)
        # ======================================================================


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def cargaVariablesDelConfigVarsDict(self):

        if CONFIGverbose:
            print(f'\n{"":_^80}')
            print(f'clidconfig-> Los parametros globales de configuracion, guardados en configVarsDict se cargan como propiedades de GLO')
            # print(f'{TB}{TV}-> self.configVarsDict: {self.configVarsDict}')
        # ======================================================================
        for nombreParametroDeConfiguracion in self.configVarsDict.keys():
            if self.configVarsDict['GLBLmostrarVariablesDeConfiguracion'][0]:
                if callingModulePrevio == 'clidbase':
                    if self.numObjetivoEjecucion < len(self.configVarsDict[nombreParametroDeConfiguracion]):
                        print(
                            'clidconfig->-------------------------->',
                            nombreParametroDeConfiguracion,
                            type(nombreParametroDeConfiguracion),
                            self.configVarsDict[nombreParametroDeConfiguracion][self.numObjetivoEjecucion]
                        )
                    else:
                        print(
                            'clidconfig->-------------------------->',
                            nombreParametroDeConfiguracion,
                            type(nombreParametroDeConfiguracion),
                            self.configVarsDict[nombreParametroDeConfiguracion][0],
                            '(sin valor especifico para el self.numObjetivoEjecucion',
                            self.numObjetivoEjecucion, ')'
                        )
            # print('clidconfig-> Parametro', nombreParametroDeConfiguracion, '>> self.configVarsDict:', len(self.configVarsDict[nombreParametroDeConfiguracion]), self.configVarsDict[nombreParametroDeConfiguracion])
            if self.numObjetivoEjecucion < len(self.configVarsDict[nombreParametroDeConfiguracion]):
                setattr(self, nombreParametroDeConfiguracion, self.configVarsDict[nombreParametroDeConfiguracion][self.numObjetivoEjecucion])
            else:
                if callingModulePrevio == 'clidbase' and nombreParametroDeConfiguracion != 'configFileNameCfg':
                    print(f'{TB}-> El parametro {nombreParametroDeConfiguracion} no tiene valor especifico para el objetivo de ejecucion {self.numObjetivoEjecucion} (>= {len(self.configVarsDict[nombreParametroDeConfiguracion])})')
                    print(f'{TB}-> Se adopta el valor correspondiente al objetivo general: {self.configVarsDict[nombreParametroDeConfiguracion][0]}') 
                setattr(self, nombreParametroDeConfiguracion, self.configVarsDict[nombreParametroDeConfiguracion][0])
        # ======================================================================

        if usarXLS and self.configVarsDict['GLBLmostrarVariablesDeConfiguracion'][0]:
            if callingModulePrevio == 'clidbase':
                print(f'{"":=^80}')


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def revisarCompletarVariablesMAINdelConfigVarsDict(self):
        # ======================================================================
        # En esta funcion se establecen algunas variables globales (si es necesario modificarlas):
        #    MAINprocedimiento  -> Se retoca si el procedimiento menciona calendula pero lo ejecuto en Windows
        # ======================================================================

        # ======================================================================
        # Adaptacion a/de MAINprocedimiento si se ejecuta fuera de calendula
        if (
            (self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion]).startswith('AUTOMATICO_EN_CALENDULA_SCRATCH')
            or (self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion]).startswith('AUTOMATICO_EN_CALENDULA_SELECT')
        ):
            if MAIN_ENTORNO != 'calendula':
                self.MAINprocedimiento = (self.MAINprocedimiento).replace(
                    'AUTOMATICO_EN_CALENDULA_SCRATCH',
                    'AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA'
                ).replace(
                    'AUTOMATICO_EN_CALENDULA_SELECT',
                    'AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA'
                )
                self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion] = self.MAINprocedimiento
        return True


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def revisarCompletarVariablesGLBLdelConfigVarsDict(self):
        # Se recalculan a la vista del entorno de trabajo
        if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
            print(f'\n{"":_^80}')
            print(f'clidconfig-> Chequeando la configuracion...')


        # Si mi MAINprocedimiento es PASADAS:
        if (
            'LAS_INFO' in self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion]
            or 'LAS_EDIT' in self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion]
            or 'PASADAS' in self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion]
        ) and GLO.GLBLforzarExtensionDeBloqueAvalorNominal:
            GLO.GLBLforzarExtensionDeBloqueAvalorNominal = False
            self.configVarsDict['GLBLforzarExtensionDeBloqueAvalorNominal'] = [False, 'GrupoDimensionCeldasBloques', '', 'bool']
            print(f'clidconfig-> Aviso: si mi procedimiento incluye LAS_INFO, LAS_EDIT o PASADAS entonces cambio GLBLforzarExtensionDeBloqueAvalorNominal a False.')

        # Si mi objetivo es CREAR_LAZ y mi modelo no usa las hiperformas, me ahorro ese paso.
        # Para saber si el modelo usa las hiperformas, me baso en:
        #   1. El numero de inputs del nombre del modelo: GLBLnombreFicheroConModeloParaInferencia
        #   2. En su defecto, en GLBLincluirVarDeHiperformasEnElModeloAcumulativo
        if self.configVarsDict['MAINobjetivoEjecucion'][self.numObjetivoEjecucion] == 'CREAR_LAZ':
            if (
                not self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][self.numObjetivoEjecucion] is None
                and self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][self.numObjetivoEjecucion] != ''
                and '_i' in self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][self.numObjetivoEjecucion]
            ):
                posicionNumInputVars = (self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][self.numObjetivoEjecucion]).index('_i')
                # Asumo un maximo de 999 variables input
                nInputVars = int((self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][self.numObjetivoEjecucion])[posicionNumInputVars + 2 : posicionNumInputVars + 5])
                if self.LCLverbose:
                    print(f'clidconfig-> AVISO: si mi objetivo es CREAR_LAZ y mi modelo no usa las hiperformas, me ahorro ese paso.')
                    print(f'{TB}-> Saco el numero de inputs del nombre del modelo:')
                    print(f'{TB}{TV}-> GLBLnombreFicheroConModeloParaInferencia: {self.configVarsDict["GLBLnombreFicheroConModeloParaInferencia"][0]}')
                    print(f'{TB}{TV}-> nInputVars: {nInputVars}')
            else:
                nInputVars = 0
            if (
                (nInputVars > 0 and not nInputVars in [71, 72, 73, 83, 84, 85])
                or (nInputVars == 0 and not self.configVarsDict['GLBLincluirVarDeHiperformasEnElModeloAcumulativo'][self.numObjetivoEjecucion])
            ):
                self.configVarsDict['GLBLcalcularHiperFormas'][self.numObjetivoEjecucion] = False
                if self.LCLverbose:
                    print(f'{TB}-> Como mi objetivo es CREAR_LAZ y mi modelo no usa las hiperformas')
                    print(f'{TB}{TV}Cambio GLBLcalcularHiperFormas a False')
            else:
                if self.LCLverbose:
                    print(f'{TB}-> No necesito cambiar GLBLcalcularHiperFormas a False')

        if GLO.GLBLmetrosCelda != 10:
            if GLO.GLBLcalcularMdkConClasificacionInferida or GLO.GLBLhacerInferenciaParaTodosLosPuntos: 
                if True:
                    print(f'clidconfig-> AVISO: GLBLmetrosCelda != 10 y:')
                    print(f'{TB}-> Se hace inferencia para todos los puntos con el modelo entrenado para celdas de 10 m-> GLBLhacerInferenciaParaTodosLosPuntos: {GLO.GLBLhacerInferenciaParaTodosLosPuntos}).')
                    print(f'{TB}-> Se calcula Mdk con la clasificacion inferida-> GLBLcalcularMdkConClasificacionInferida: {GLO.GLBLcalcularMdkConClasificacionInferida}.')
                    print(f'{TB}-> Se calcula Mdk con la clasificacion original-> GLBLcalcularMdkConClasificacionOriginal: {GLO.GLBLcalcularMdkConClasificacionOriginal}.')
                else:
                    GLO.GLBLcalcularMdkConClasificacionOriginal = True
                    GLO.GLBLhacerInferenciaParaTodosLosPuntos = False
                    GLO.GLBLgenerarNuevoLaxReclasificado = False
                    print(f'clidconfig-> AVISO: no se hace inferencia ni se genera nuevo lasFile porque GLBLmetrosCelda es {GLO.GLBLmetrosCelda} != 10.')
        if (
            GLO.GLBLpredecirClasificaMiniSubCelConvolucional
            or GLO.GLBLpredecirCubiertasSingularesConvolucional
        ):
            GLO.GLBLcrearTilesTargetMiniSubCelSoloSiHayNoSueloSuficientes = False
            GLO.GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes = False

        if (
            not GLO.GLBLhacerInferenciaEntrenandoModeloAcumulativoTree
            and not GLO.GLBLentrenarNeuronalNetworkConTF
            and not GLO.GLBLcalcularAciertoEntrenandoModeloAcumulativoNeuronal
        ):
            if self.configVarsDict['GLBLcalcularMdkConClasificacionInferida'][self.numObjetivoEjecucion]:
                print('clidconfig-> ATENCION: Para generar plano Mdk se requieren puntos reclasificados (hay que hacer inferencia).')
                sys.exit(0)
            if self.configVarsDict['GLBLgenerarNuevoLaxReclasificado'][self.numObjetivoEjecucion]:
                print('clidconfig-> ATENCION: Para guardar puntos reclasificados las hay que hacer inferencia (tree o nln).')
                sys.exit(0)
            if self.configVarsDict['GLBLguardarTrainPointsEnShape'][self.numObjetivoEjecucion]:
                print('clidconfig-> ATENCION: Para guardar puntos reclasificados shape hay que hacer inferencia (tree o nln).')
                sys.exit(0)

        if not self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][self.numObjetivoEjecucion]:
            print(
                'clidconfig-> AVISO: he quitado la opcion de guardar como string.',
                '\nSe cambia GLBLalmacenarPuntosComoNumpyDtype a True (GLBLalmacenarPuntosComoByteString = False).'
            )
            self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][self.numObjetivoEjecucion] = True
            self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion] = False


        if self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][self.numObjetivoEjecucion]:
            print(
                'clidconfig-> AVISO: he quitado la opcion de guardar version mini string.',
                '\nSe cambia GLBLalmacenarPuntosComoCompactNpDtype a False.'
            )
            self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][self.numObjetivoEjecucion] = False

        if (
            self.configVarsDict['GLBLsoloCrearTilesNoGuardarAsc'][self.numObjetivoEjecucion]
            or self.configVarsDict['GLBLsoloGuardarArraysNpzSinCrearOutputFiles'][self.numObjetivoEjecucion]
        ) and (
            self.configVarsDict['GLBLcrearTilesPostVuelta2'][self.numObjetivoEjecucion]
            or self.configVarsDict['GLBLguardarArraysVuelta2a9EnNpz'][self.numObjetivoEjecucion]
            or self.configVarsDict['GLBLreDepurarMiniSubCelEnVueltaAjustesMdp'][self.numObjetivoEjecucion]
        ):
            if self.LCLverbose:
                print(
                    'clidconfig-> AVISO: Si no se crean output files (asc), no tienen validez estos parametros:',
                    '\nGLBLcrearTilesPostVuelta2',
                    '\nGLBLguardarArraysVuelta2a9EnNpz',
                    '\nGLBLreDepurarMiniSubCelEnVueltaAjustesMdp'
                )

        if not self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][self.numObjetivoEjecucion]:
            if self.LCLverbose:
                print('clidconfig-> AVISO: revisar si se quiere usar formato de punto de texto (en desuso)')
            if (
                GLO.GLBLacumularPuntosEnNpzParaEntrenamientoFuturo
                or GLO.GLBLentrenarNeuronalNetworkConTF
                or GLO.GLBLhacerInferenciaParaTodosLosPuntos
            ):
                print('\nclidconfig-> ATENCION: no se puede haer entrenamiento o inferencia con formato de punto que no sea dtype')
                print(f'{TB}-> Cambiar la configuracion')
                sys.exit(0)

        if self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][self.numObjetivoEjecucion] != 0:
            if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                if self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][self.numObjetivoEjecucion] in [1, 7, 12, 18] and (
                    self.configVarsDict['GLBLentrenarConTarget_1lasOrig_2LasAsig_3lasPred'][self.numObjetivoEjecucion] == 2
                ):
                    if self.LCLverbose:
                        print(
                            'clidconfig-> AVISO: Se va a entrenar con puntos acumulativos con una sola clase seleccionada',
                            '\nLa clase seleccionada es {} (no interesa): Revisar GLBLentrenarBinarioCategoriaSeleccionada'.format(
                                self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][self.numObjetivoEjecucion]
                            )
                        )
                elif self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][self.numObjetivoEjecucion] > 100:
                    if self.LCLverbose:
                        print(
                            'clidconfig-> AVISO: No hay categorias de lasClass superiores a 100. GLBLentrenarBinarioCategoriaSeleccionada: {}'.format(
                                self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][self.numObjetivoEjecucion]
                            )
                        )

        if self.configVarsDict['GLBLcalcularHiperFormas'][self.numObjetivoEjecucion]:
            if (
                not self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion]
                and not self.configVarsDict['GLBLcalcularMdb'][self.numObjetivoEjecucion]
                and (
                    not self.configVarsDict['GLBLcalcularMdp'][self.numObjetivoEjecucion] or not (
                        self.configVarsDict['GLBLcalcularMdfConMiniSubCelValidadosConMetodoManualPuro'][self.numObjetivoEjecucion]
                        or self.configVarsDict['GLBLcalcularMdfConMiniSubCelValidadosConModeloConvolucional'][self.numObjetivoEjecucion]
                        or self.configVarsDict['GLBLcalcularMdfConMiniSubCelValidadosConModConvoManualizado'][self.numObjetivoEjecucion]
                    )
                )
            ):
                self.configVarsDict['GLBLcalcularHiperFormas'][self.numObjetivoEjecucion] = False
                if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                    if self.LCLverbose:
                        print(
                            'clidconfig-> AVISO: Se desactiva el calculo de hiperformas porque no se calculan el plano pleno ni basal ni suelo.',
                            '\n  Revisar GLBLcalcularHiperFormas, GLBLcalcularMdp, GLBLcalcularMdfConMiniSubCelValidados***',
                            '\n  Revisar configuracion.'
                        )

        if self.configVarsDict['GLBLacumularPuntosEnNpzParaEntrenamientoFuturo'][self.numObjetivoEjecucion]:
            if (
                not self.configVarsDict['GLBLcrearTilesPostVuelta1'][self.numObjetivoEjecucion]
                or not self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][self.numObjetivoEjecucion]
                or not self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][self.numObjetivoEjecucion]
            ):
                print('\nclidconfig-> ATENCION: Para acumular puntos para entrenamiento, es necesario predecir convolucionalmente la clase de los miniSubCel y los usos singulares')
                print('\t-> GLBLcrearTilesPostVuelta1:                    {}'.format(self.configVarsDict['GLBLcrearTilesPostVuelta1'][self.numObjetivoEjecucion]))
                print('\t-> GLBLpredecirCubiertasSingularesConvolucional: {}'.format(self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][self.numObjetivoEjecucion]))
                print('\t-> GLBLpredecirClasificaMiniSubCelConvolucional: {}'.format(self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][self.numObjetivoEjecucion]))
                print('Revisar configuracion.')
                sys.exit(0)


        if self.configVarsDict['GLBLsoloGuardarArraysNpzSinCrearOutputFiles'][self.numObjetivoEjecucion] and not self.configVarsDict['GLBLcrearTilesPostVuelta2'][self.numObjetivoEjecucion] and self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
            if self.LCLverbose:
                print(
                    'clidconfig-> AVISO: si GLBLsoloGuardarArraysNpzSinCrearOutputFiles,',
                    '\n  no tiene sentido que este activado GLBLcrearTilesPostVuelta2.',
                    '\n  Se puede cambiar en el fichero de configuracion.'
                )
        
        if (
            (
                self.configVarsDict['GLBLcrearTilesTargetDeCartoRefMdt'][self.numObjetivoEjecucion]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefNucleos'][self.numObjetivoEjecucion]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefLandCover'][self.numObjetivoEjecucion]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefSingUse'][self.numObjetivoEjecucion]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefPixel1m'][self.numObjetivoEjecucion]
            )
            and not self.configVarsDict['GLBLformatoTilesAscRasterRef'][self.numObjetivoEjecucion]
            and self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]
        ):
            if self.LCLverbose:
                print(f'clidconfig-> AVISO: si GLBLcrearTilesTargetDeCartoRefSingUse u otros TargetDeCartoRef,')
                print(f'{TB}-> es recomendable GLBLformatoTilesAscRasterRef para visualizar en Qgis los tiles que se generan.')
                print(f'{TB}-> Se puede cambiar en el fichero de configuracion.')

        if (
            self.configVarsDict['GLBLeliminarTilesTrasProcesado'][self.numObjetivoEjecucion]
            and (
                not self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][self.numObjetivoEjecucion]
                or not self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][self.numObjetivoEjecucion]
                or self.configVarsDict['GLBLsoloCrearTilesNoGuardarAsc'][self.numObjetivoEjecucion]
            )
            and self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]
        ):
            print(
                '\nclidconfig-> ATENCION: si GLBLeliminarTilesTrasProcesado,',
                '\n  es porque se usan para hacer predicciones, (GLBLpredecirCubiertasSingularesConvolucional or GLBLpredecirClasificaMiniSubCelConvolucional).',
                '\n  y porque no es una ejecucion para GLBLsoloCrearTilesNoGuardarAsc.'
                '\n  Revisar configuracion.'
            )
            return False

        if (
            self.configVarsDict['GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes'][self.numObjetivoEjecucion]
            and (
                not self.configVarsDict['GLBLsoloCrearTilesNoGuardarAsc'][self.numObjetivoEjecucion]
                or self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][self.numObjetivoEjecucion]
                or self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][self.numObjetivoEjecucion]
            )
            and self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]
        ):
            print(
                '\nclidconfig-> ATENCION: si GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes,',
                '\n  es porque se trata de una ejecucion destinada solo a crear tiles (GLBLsoloCrearTilesNoGuardarAsc)',
                '\n  y no se va a hacer predicciones, (not GLBLpredecirCubiertasSingularesConvolucional and not GLBLpredecirClasificaMiniSubCelConvolucional).',
                '\n  Para hacer predicciones se deberian generar todos los tiles.'
                '\n  Se puede cambiar en el fichero de configuracion.'
            )
            return False

        minimoDeMemoriaRAMrecomendable = asignarMinimoDeMemoriaRAM(self.configVarsDict['GLBLminimoDeMemoriaRAM'][self.numObjetivoEjecucion])
        if minimoDeMemoriaRAMrecomendable != self.configVarsDict['GLBLminimoDeMemoriaRAM'][self.numObjetivoEjecucion]:
            self.configVarsDict['GLBLminimoDeMemoriaRAM'] = [minimoDeMemoriaRAMrecomendable, 'GrupoManejoMemoria', '', 'int']
        if self.configVarsDict['GLBLusarNumba'][self.numObjetivoEjecucion] and self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][self.numObjetivoEjecucion] < 100:
            self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][self.numObjetivoEjecucion] = 100
        if not self.configVarsDict['GLBLcalcularMdb'][self.numObjetivoEjecucion]:
            self.configVarsDict['GLBLproyectarPuntosSobreMdb'][self.numObjetivoEjecucion] = False
        if self.configVarsDict['GLBLmetrosBloc'][self.numObjetivoEjecucion] < self.configVarsDict['GLBLmetrosCelda'][self.numObjetivoEjecucion]:
            self.configVarsDict['GLBLmetrosBloc'][self.numObjetivoEjecucion] = self.configVarsDict['GLBLmetrosCelda'][self.numObjetivoEjecucion]
        if self.configVarsDict['GLBLtipoIndice'][self.numObjetivoEjecucion] == 1:
            self.configVarsDict['GLBLtipoIndice'][self.numObjetivoEjecucion] = int(
                math.ceil(1.0 * self.configVarsDict['GLBLmetrosBloque'][self.numObjetivoEjecucion] / self.configVarsDict['GLBLmetrosBloc'][self.numObjetivoEjecucion])
            )  # malla regular de blocs
            self.configVarsDict['GLBLtipoIndice'][self.numObjetivoEjecucion] = 100 if self.configVarsDict['GLBLtipoIndice'][self.numObjetivoEjecucion] > 100 else self.configVarsDict['GLBLtipoIndice'][self.numObjetivoEjecucion]
        if self.configVarsDict['GLBLmetrosCelda'][self.numObjetivoEjecucion] < 5:
            self.configVarsDict['GLBLminimoDePuntosTodosRetornosPasadaSeleccionada'][self.numObjetivoEjecucion] = int(
                self.configVarsDict['GLBLminimoDePuntosTodosRetornosPasadaSeleccionada'][self.numObjetivoEjecucion] / 2
            )
            self.configVarsDict['GLBLminimoDePuntosTotales'][self.numObjetivoEjecucion] = int(self.configVarsDict['GLBLminimoDePuntosTotales'][self.numObjetivoEjecucion] / 2)
            self.configVarsDict['GLBLminimoDePuntosSueloParaAjustarPlano'][self.numObjetivoEjecucion] = int(self.configVarsDict['GLBLminimoDePuntosSueloParaAjustarPlano'][self.numObjetivoEjecucion] / 2)
            self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][self.numObjetivoEjecucion] = int(self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][self.numObjetivoEjecucion] / 2)
        if self.configVarsDict['GLBLminDePtosParaAjustarPlanoBasalCielo'][self.numObjetivoEjecucion] < 3 or self.configVarsDict['GLBLminDePtosParaAjustarPlanoMajor'][self.numObjetivoEjecucion] < 3:
            if self.LCLverbose:
                print(f'clidconfig-> Corregir GLBLminDePtosParaAjustarPlanoBasalCielo o GLBLminDePtosParaAjustarPlanoMajor')
                print(f'clidconfig-> El numero minimo de puntos para ajustar debe ser mayor de 3')
            return False

        if (
            self.configVarsDict['GLBLusarNumba'][self.numObjetivoEjecucion]
            and not self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][self.numObjetivoEjecucion]
            and not self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion]
        ):
            if self.LCLverbose:
                print(f'clidconfig-> Si se usa numba y no se usa Dtype, solo se puede guardar con GLBLalmacenarPuntosComoByteString = True -> Se cambia a True')
            try:
                selec = input('\tclidconfig-> Confirmar el cambio de GLBLalmacenarPuntosComoByteString a True (S/n)')
                self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion] = False if selec.upper() == 'N' else True
            except:
                self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion] = True
            if self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion] == False:
                if self.LCLverbose:
                    print(
                        '\tclidconfig-> self.configVarsDict["GLBLalmacenarPuntosComoByteString"] = False -> Opcion no permitida con GLBLusarNumba=True y self.configVarsDict["GLBLalmacenarPuntosComoNumpyDtype"]=False'
                    )
                return False
            if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                print('GLBLalmacenarPuntosComoByteString:', self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion])

        if self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][self.numObjetivoEjecucion] > 0 and not self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion]:
            if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                if self.LCLverbose:
                    print(f'clidconfig-> -> la pasada elegida debe tener puntos suelo -> se cambia la opcion GLBLselecPasadasConClasificacion a True')
            try:
                selec = input('\tclidconfig-> Confirmar el cambio de GLBLselecPasadasConClasificacion a True (S/n)')
                self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion] = False if selec.upper() == 'N' else True
            except:
                self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion] = True
            if self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion] == False:
                if self.LCLverbose:
                    print(
                        '\tclidconfig-> self.configVarsDict["GLBLselecPasadasConClasificacion"][0] = False -> Opcion no permitida con GLBLminimoDePuntosSueloParaElegirPasada > 0'
                    )
                return False
            if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                if self.LCLverbose:
                    print('GLBLselecPasadasConClasificacion:', self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion])

        if self.configVarsDict['GLBLusarNumba'][self.numObjetivoEjecucion] and (self.configVarsDict['GLBLusarSklearn'][self.numObjetivoEjecucion] or self.configVarsDict['GLBLusarStatsmodels'][self.numObjetivoEjecucion]):
            if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                if self.LCLverbose:
                    print(f'clidconfig-> Cuando se usa Numba los ajustes se hacen con algebra matricial y no con SkLearn o Statsmodels.')
            try:
                selec = input('\tclidconfig-> Confirmar el cambio de GLBLusarSklearn y GLBLusarStatsmodels a False (S/n)')
                rpta = True if selec.upper() == 'N' else False
                self.configVarsDict['GLBLusarSklearn'][self.numObjetivoEjecucion] = rpta
                self.configVarsDict['GLBLusarStatsmodels'][self.numObjetivoEjecucion] = rpta
            except:
                self.configVarsDict['GLBLusarSklearn'][self.numObjetivoEjecucion] = False
                self.configVarsDict['GLBLusarStatsmodels'][self.numObjetivoEjecucion] = False
            if rpta == False:
                if self.LCLverbose:
                    print(f'clidconfig-> GLBLusarSklearn o GLBLusarStatsmodels = True -> Opcion no permitida con GLBLusarNumba True')
                return False
            if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                if self.LCLverbose:
                    print('GLBLusarSklearn:    ', self.configVarsDict['GLBLusarSklearn'][self.numObjetivoEjecucion])
                    print('GLBLusarStatsmodels:', self.configVarsDict['GLBLusarStatsmodels'][self.numObjetivoEjecucion])


        if (
            self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][self.numObjetivoEjecucion]
            or self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][self.numObjetivoEjecucion]
        ):
            if self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion]:
                if (
                    self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][self.numObjetivoEjecucion]
                    and self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][self.numObjetivoEjecucion]
                ):
                    print(f'clidconfig-> AVISO:')
                    print(f'{TB}-> GLBLguardarPuntosSueloEnArrayPredimensionada = True')
                    print(f'{TB}-> GLBLguardarPuntosSueloAlFinalDelArrayPralAll = True')
                    print(f'{TB}-> Esto guarda los puntos duplicados y consume extra de RAM')
                    print(f'{TB}-> Se desactiva la opcion GLBLguardarPuntosSueloAlFinalDelArrayPralAll')
                    selec = 'N'
                    self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][self.numObjetivoEjecucion] = False if selec.upper() == 'N' else True
            else:
                print(f'clidconfig-> AVISO:')
                print(f'{TB}-> GLBLguardarPuntosSueloEnArrayPredimensionada = {self.configVarsDict["GLBLguardarPuntosSueloEnArrayPredimensionada"][0]}')
                print(f'{TB}-> GLBLguardarPuntosSueloAlFinalDelArrayPralAll = {self.configVarsDict["GLBLguardarPuntosSueloAlFinalDelArrayPralAll"][0]}')
                print(f'{TB}-> GLBLcalcularMds = False')
                print(f'{TB}-> Esto guarda los puntos suelo pero no se usan para Mds con lo que consume RAM innecesariamente.')
                print(f'{TB}-> Se desactivan las opciones GLBLguardarPuntosSueloEnArrayPredimensionada y GLBLguardarPuntosSueloAlFinalDelArrayPralAll.')
                selec = 'N'
                self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][self.numObjetivoEjecucion] = False if selec.upper() == 'N' else True
                self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][self.numObjetivoEjecucion] = False if selec.upper() == 'N' else True
            if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion]:
                print(f'{TB}-> GLBLguardarPuntosSueloEnArrayPredimensionada: {self.configVarsDict["GLBLguardarPuntosSueloEnArrayPredimensionada"][0]}')
                print(f'{TB}-> GLBLguardarPuntosSueloAlFinalDelArrayPralAll: {self.configVarsDict["GLBLguardarPuntosSueloAlFinalDelArrayPralAll"][0]}')
        else:
            if self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion]:
                try:
                    print(f'clidconfig-> AVISO:')
                    print(f'{TB}-> Se quiere calcular el plano suelo; pero eso requiere ')
                    print(f'{TB}{TV}almacenar puntos suelo en array predimensionada.')
                    print(f'{TB}{TV}-> Se activa esa opcion.')
                    selec = 'S'
                    # selec = input(f'{TB}{TV}almacenar puntos suelo en array predimensionada. Activar esa opcion? (S/n)')
                    confirmar = False if selec.upper() == 'N' else True
                except:
                    confirmar = True
                if confirmar:
                    self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][self.numObjetivoEjecucion] = True
                    print(f'{TB}-> Se usa el array self_aCeldasListaDePtosSuePral. ')
                    print(f'{TB}-> Si se quiere GLBLguardarPuntosSueloAlFinalDelArrayPralAll, cambiar clidbase.xlsx')
                else:
                    self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion] = False
                    print(f'{TB}-> No se calcula en plano suelo.')

        # Atencion: para calcular el punto suelo fuerzo que se use GLBLguardarPuntosSueloEnArrayPredimensionada
        # por lo que lo siguiente ya no tiene vigencia porque
        # no se si se puede calcular plano suelo usando la Psel,
        # sin almacenar los puntos suelo en su array/lugar especifico
        if (
            self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion]
            and not self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][self.numObjetivoEjecucion]
            and not self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion]
            and self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][self.numObjetivoEjecucion] == 0
        ):
            if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                print(f'clidconfig-> -> Se ajusta plano a los puntos suelo pero:')
                print(f'{TB}No se guardan en array los puntos de la pasada con puntos suelo porque')
                print(f'{TB}ocupa demasiada memoria, por lo que se usan siempre los puntos de la pasada Psel.')
                print(f'{TB}La pasada Psel se selecciona por angulo de incidencia sin requerir que tenga puntos clasificados suelo')
                print(f'{TB}GLBLcalcularMds = True pero GLBLguardarPuntosSueloEnArrayPredimensionada = False ')
                print(f'{TB}y GLBLselecPasadasConClasificacion = False y GLBLminimoDePuntosSueloParaElegirPasada = 0')
            try:
                print(f'{TB}Confirmar que se quiere calcular el plano suelo usando la pasada seleccionada')
                selec = input('solo por angulo y que, por lo tanto, puede no tener puntos suelo (S/n)')
                # if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                #     print(
                #         f'{TB}Confirmar que se quiere calcular el plano suelo usando la pasada seleccionada solo por angulo y que, por lo tanto, puede no tener puntos suelo (S/n)'
                #     )
                #     selec = 'S'
                confirmar = False if selec.upper() == 'N' else True
            except:
                confirmar = True
            if not confirmar:
                try:
                    selec = input('\tclidconfig->     Calcular el plano suelo? (S/n)')
                    self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion] = False if selec.upper() == 'N' else True
                except:
                    self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion] = True
                if self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion]:
                    try:
                        selec = input('\tclidconfig->     Requerir que la pasada seleccionada tenga puntos clasificados? (S/n)')
                        requerirPuntosClasificados = False if selec.upper() == 'N' else True
                    except:
                        requerirPuntosClasificados = True
                    if requerirPuntosClasificados:
                        self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion] = True
                        try:
                            selec = input('\tclidconfig->     Requerir que la pasada seleccionada tenga al menos un punto clasificado suelo? (S/n)')
                            requerirPuntoClasificadoSuelo = False if selec.upper() == 'N' else True
                        except:
                            requerirPuntoClasificadoSuelo = True
                        if requerirPuntoClasificadoSuelo:
                            self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion] = True
                            self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][self.numObjetivoEjecucion] = 1
            if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                print(
                    '\tclidconfig->     GLBLcalcularMds =',
                    self.configVarsDict['GLBLcalcularMds'][self.numObjetivoEjecucion],
                    'GLBLguardarPuntosSueloEnArrayPredimensionada =',
                    self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][self.numObjetivoEjecucion],
                    'GLBLminimoDePuntosSueloParaElegirPasada =',
                    self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][self.numObjetivoEjecucion],
                )

            if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                print(
                    '\tclidconfig-> GLBLselecPasadasConClasificacion:',
                    self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion],
                    'GLBLselecPasadaConMasPuntosSuelo:',
                    self.configVarsDict['GLBLselecPasadaConMasPuntosSuelo'][self.numObjetivoEjecucion],
                )
            if self.configVarsDict['GLBLselecPasadaConMasPuntosSuelo'][self.numObjetivoEjecucion]:
                if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                    print(f'clidconfig-> Se selecciona la misma pasada para puntos suelo y para puntos basales')
            elif self.configVarsDict['GLBLselecPasadasConClasificacion'][self.numObjetivoEjecucion]:
                if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                    print(f'clidconfig-> Si solo hay una pasada con puntos clasificados, se selecciona la misma pasada para puntos suelo y para puntos basales')
                    print(
                        f'clidconfig-> Si hay varias, en las celdas con mas de una pasada, la pasada seleccionada (para puntos basales) puede ser distinta de la seleccionada para puntos suelo:'
                    )
                    print(f'clidconfig->     Para puntos suelo: la que teniendo puntos clasificados tenga mas puntos suelo')
                    print(f'clidconfig->     Para puntos basales: la que teniendo puntos clasificados tenga menor angulo medio')
            else:
                if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                    print(
                        f'clidconfig-> En las celdas con mas de una pasada, la pasada seleccionada para puntos basales puede ser distinta de la seleccionada para puntos suelo'
                    )
                    print(f'clidconfig->     Para puntos suelo: la que tenga mas puntos suelo')
                    print(f'clidconfig->     Para puntos basales: la que tenga menor angulo medio')
                    print(f'clidconfig-> AVISO: se trabaja solo con celdas seleccionadas para puntos basales: algunas pueden no tener puntos suelo')

        if self.configVarsDict['GLBLgrabarPropiedadTime'][self.numObjetivoEjecucion] and self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][self.numObjetivoEjecucion]:
            if self.configVarsDict['GLBLmostrarAvisos'][self.numObjetivoEjecucion]:
                print(
                    '\tclidconfig-> GLBLgrabarPropiedadTime = True-> Si se quiere guardar la propiedad RawTime no se puede usar GLBLalmacenarPuntosComoCompactNpDtype'
                )
            try:
                selec = input('\tclidconfig-> Quieres mantener GLBLalmacenarPuntosComoCompactNpDtype = True (no se graba RawTime)? (S/n)')
                self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][self.numObjetivoEjecucion] = False if selec.upper() == 'N' else True
            except:
                self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][self.numObjetivoEjecucion] = True
            if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
                print('GLBLguardarPuntosSueloEnArrayPredimensionada:', self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][self.numObjetivoEjecucion])

        if self.configVarsDict['GLBLgeigerMode'][self.numObjetivoEjecucion]:
            if self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion]:
                self.configVarsDict['GLBLalmacenarPuntosComoByteString'][self.numObjetivoEjecucion] = False
                self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][self.numObjetivoEjecucion] = True
            # Esto siguiente no hace falta porque mas adelante reviso el numero medio de puntos por celda y amplio GLBLnMaxPtosCeldaArrayPredimensionadaTodos si es necesario
            # if self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][self.numObjetivoEjecucion] < 20 * (self.configVarsDict['GLBLmetrosCelda'][self.numObjetivoEjecucion] ** 2):
            #    self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][self.numObjetivoEjecucion] = 20 * (self.configVarsDict['GLBLmetrosCelda'][self.numObjetivoEjecucion] ** 2)

        if self.configVarsDict['GLBLverbose'][self.numObjetivoEjecucion] and self.LCLverbose:
            print(f'{"":=^80}')

        return True


    # ==========================================================================
    def revisarSiHayObjetivoEjecucionNoReglado(self):
        print(f'clidconfig-> GLO.MAINobjetivoEjecucion (bb): {self.MAINobjetivoEjecucion}')
        # A este metodo se le llama desde clidbase.py despues de leer los argumentos en linea de comandos
        # ======================================================================
        if self.configVarsDict['MAINobjetivoEjecucion'][self.numObjetivoEjecucion] in self.listaObjetivosEjecucionSiReglados:
            # Casos habtuales (objetivos de ejecucion si reglados):
            self.MAINobjetivoSiReglado = self.MAINobjetivoEjecucion
            self.configVarsDict['MAINobjetivoSiReglado'] = self.configVarsDict['MAINobjetivoEjecucion']
            self.MAINobjetivoNoReglado = 'NINGUNO'
            self.configVarsDict['MAINobjetivoNoReglado'] = [
                'NINGUNO', 'GrupoMAIN', '', 'str', 'NINGUNO',
                'NINGUNO', 'NINGUNO', 'NINGUNO',
                'NINGUNO', 'NINGUNO', 'NINGUNO',
            ]
        elif self.configVarsDict['MAINobjetivoEjecucion'][self.numObjetivoEjecucion] in self.listaObjetivosEjecucionNoReglados:
            if self.LCLverbose:
                print(
                    f'clidconfig-> Adaptando algunos parametros al objetivoEjecucion no reglado:',
                    f'{self.configVarsDict["MAINobjetivoEjecucion"][self.numObjetivoEjecucion]}'
                )
            # Casos espaciales (objetivos de ejecucion no reglados):
            if self.configVarsDict['MAINobjetivoEjecucion'][self.numObjetivoEjecucion] == 'CREAR_PUNTOS_TRAIN_ROQUEDOS':
                if self.LCLverbose:
                    print(f'clidconfig-> Se adapta la ejecucion a la recopilacion de puntos roquedo para entrenamiento.')
                self.MAINobjetivoSiReglado = 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ'
                self.configVarsDict['MAINobjetivoSiReglado'] = [
                    'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'GrupoMAIN', '', 'str', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ',
                    'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ',
                    'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ',
                ]
                self.MAINobjetivoNoReglado = 'CREAR_PUNTOS_TRAIN_ROQUEDOS'
                self.configVarsDict['MAINobjetivoNoReglado'] = [
                    'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'GrupoMAIN', '', 'str', 'CREAR_PUNTOS_TRAIN_ROQUEDOS',
                    'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS',
                    'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS',
                ]
                self.MAINprocedimiento = 'AUTOMATICO_EN_CALENDULA_SELECT'
                # self.GLBLusarVectorNucleosUrbanosVector = True
                # self.GLBLusarVectorNucleosUrbanosRaster = True
                self.GLBLusarVectorGeologicoVector = True
                self.GLBLusarVectorGeologicoRaster = True
                self.GLBLsoloRoquedosParaEntrenamiento = True
                self.GLBLfraccionDeMuestreoGeneral = 1
                self.GLBLmuestreoEspecificoDeClase = False
                self.GLBLsobreMuestrearClasesSubRepresentadas = False
                self.GLBLsubMuestrearClasesSobreRepresentadas = False
                # Esto condiciona la lectura de una u otra columna de parece que no tiene efecto:
                self.configVarsDict['MAINprocedimiento'] = ['AUTOMATICO_EN_CALENDULA_SELECT', 'GrupoMAIN', '', 'str', 'AUTOMATICO_EN_CALENDULA_SELECT']
                self.configVarsDict['MAINcuadrante'][self.numObjetivoEjecucion] = 'sE'
                self.configVarsDict['GLBLusarVectorGeologicoVector'][self.numObjetivoEjecucion] = True
                self.configVarsDict['GLBLusarVectorGeologicoRaster'][self.numObjetivoEjecucion] = True
                self.configVarsDict['GLBLsoloRoquedosParaEntrenamiento'][self.numObjetivoEjecucion] = True
                self.configVarsDict['GLBLfraccionDeMuestreoGeneral'][self.numObjetivoEjecucion] = 1
                self.configVarsDict['GLBLmuestreoEspecificoDeClase'][self.numObjetivoEjecucion] = False
                self.configVarsDict['GLBLsobreMuestrearClasesSubRepresentadas'][self.numObjetivoEjecucion] = False
                self.configVarsDict['GLBLsubMuestrearClasesSobreRepresentadas'][self.numObjetivoEjecucion] = False
        else:
            print(f'\n{"":_^80}')
            print(f'clidconfig-> ATENCION: MAINobjetivoEjecucion no reglado: {self.configVarsDict["MAINobjetivoEjecucion"]}')
            print(f'{TB}-> Corregirlo en el fichero de configuracion .cfg o .xlsx')
            sys.exit(0)

        print(f'clidconfig-> MAINobjetivoEjecucion all: {self.configVarsDict["MAINobjetivoEjecucion"]}')
        print(f'clidconfig-> numObjetivoEjecucion: {self.numObjetivoEjecucion}')
        print(f'clidconfig-> MAINobjetivoEjecucion:')
        print(f'{TB}-> GLO.MAINobjetivoEjecucion:  {GLO.MAINobjetivoEjecucion}')
        print(f'{TB}-> self.MAINobjetivoEjecucion: {self.MAINobjetivoEjecucion}')
        print(f'{TB}-> self.configVarsDict:        {self.configVarsDict["MAINobjetivoEjecucion"][self.numObjetivoEjecucion]}')
        print(f'clidconfig-> MAINobjetivoSiReglado (a): {self.MAINobjetivoSiReglado}')
        print(f'clidconfig-> MAINobjetivoNoReglado (a): {self.MAINobjetivoNoReglado}')
        print(f'clidconfig-> listaObjetivosEjecucionSiReglados:: {self.listaObjetivosEjecucionSiReglados}')

    # ==========================================================================
    def simultanearEjecucionesEnCalendula_SELECT_SCRATCH(self):
        # A este metodo se le llama desde clidbase.py despues de leer los argumentos en linea de comandos
        # Solo se ejecuta si:
        #    MAINobjetivoEjecucion llega en linea de comandos como ARGSobjetivoEjecucion 
        #    MAINcuadrante llega en linea de comandos como ARGScodCuadrante
        if __verbose__:
            print(f'\n{"":_^80}')
            print(f'clidconfig-> ATENCION: ESTO ES PROVISIONAL Y SOLO ACTUA CUANDO SE USAN EL ARGUMENTO -o (objetivoEjecucion) EN LINEA DE COMANDOS')
            print(f'{TB}-> Esta pensado solo para CALENDULA, pero funciona tb en windows.')
            print(f'{TB}-> Es para permitir la coexistencia de dos ejecuciones con un mismo clidbase.xls (en calendula).')
            print(f'{TB}-> Se fuerza un MAINprocedimiento diferente en funcion de las mayus/minusc del cuadrante:')
            print(f'{TB}{TV}-> Ejecucion completa de cuadrante SE/CE/NE y Se/Ce/Ne-> AUTOMATICO_EN_CALENDULA_SCRATCH')
            print(f'{TB}{TV}-> Ejecucion de chequeo de cuadrante se/ce/ne y sE/cE/nE-> AUTOMATICO_EN_CALENDULA_SELECT')
            print(f'{TB}-> Parametros de configuracion actuales:')
            print(f'{TB}{TV}-> ARGScodCuadrante en linea de comandos?: {self.MAINcuadrante}')
            print(f'{TB}{TV}-> self.MAINprocedimiento antes del retoque: {self.MAINprocedimiento}')
            print(f'{TB}-> Esto no tiene efecto para la ejecucion destinada a generar puntos de entrenamiento')
            print(f'{TB}{TV}-> MAINobjetivoEjecucion = CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ')
            print(f'{TB}{TV}-> Esa ejecucion la hago normalmente con Se/Ce/Ne -antes SE/CE/**- y AUTOMATICO_EN_CALENDULA_SELECT.')
        if self.MAINprocedimiento.startswith('AUTOMATICO_EN_CALENDULA'):
            if self.MAINcuadrante == 'SE' or self.MAINcuadrante == 'CE' or self.MAINcuadrante == 'NE':
                # Desde 10/2021 esto no conlleva seleccionar lasFiles, sino normalmente procesar todos los lasFiles (no solo _SELECT)
                self.MAINprocedimiento = 'AUTOMATICO_EN_CALENDULA_SCRATCH'
                self.configVarsDict['MAINprocedimiento'][0] = 'AUTOMATICO_EN_CALENDULA_SCRATCH'
                self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion] = 'AUTOMATICO_EN_CALENDULA_SCRATCH'
            elif (
                self.MAINcuadrante == 'Se' or self.MAINcuadrante == 'Ce' or self.MAINcuadrante == 'Ne'
                or self.MAINcuadrante == 'se' or self.MAINcuadrante == 'ce' or self.MAINcuadrante == 'ne'
                or self.MAINcuadrante == 'sE' or self.MAINcuadrante == 'cE' or self.MAINcuadrante == 'nE'
            ):
                self.MAINprocedimiento = 'AUTOMATICO_EN_CALENDULA_SELECT'
                self.configVarsDict['MAINprocedimiento'][0] = 'AUTOMATICO_EN_CALENDULA_SELECT'
                self.configVarsDict['MAINprocedimiento'][self.numObjetivoEjecucion] = 'AUTOMATICO_EN_CALENDULA_SELECT'
        if __verbose__:
            print(f'{TB}-> self.MAINprocedimiento desp. del retoque: {self.MAINprocedimiento}')
            print(f'{"":=^80}')


    # ==========================================================================
    def buscarPendientesSeleccionadosParaRepetir(self):
        # A este metodo se le llama desde clidbase.py despues de leer los argumentos en linea de comandos
        TRNSseleccionadosParaRepetir = 2
        if not 'MAINbuscarPendientes' in self.configVarsDict.keys():
            self.configVarsDict['MAINbuscarPendientes'] = [
                False, 'GrupoMAIN', '', 'bool', False,
                False, False, False, False, False, False,
            ]
        self.MAINbuscarPendientes = self.configVarsDict['MAINbuscarPendientes'][self.numObjetivoEjecucion]
        MAINlistaRepetir = []
        if self.MAINbuscarPendientes:
            if self.configVarsDict['MAINusuario'][self.numObjetivoEjecucion] == 'JB':
                DBF_DIR = r'../data/dbf/'
                elArchivo = DBF_DIR + 'MallaLidar2x2km.dbf'
            elif self.configVarsDict['MAINusuario'][self.numObjetivoEjecucion] == 'benmarjo':
                DBF_DIR = r'O:/Sigmena/usuarios/COMUNES/Bengoa/Lidar/Lidas/'
                elArchivo = DBF_DIR + 'MallaLidar2x2km.dbf'
            else:
                self.MAINbuscarPendientes = False
                self.configVarsDict['MAINbuscarPendientes'][self.numObjetivoEjecucion] = False
        if self.MAINbuscarPendientes:
            # MAINlistaRepetir = ['216-4656', '262-4510', '264-4510', '266-4510', '266_4510', '268-4510','268_4510']
            # MAINlistaRepetir = ['262-4510', '264-4510', '266-4510', '268-4510']
            # nfilas = len(MAINlistaRepetir)
            nfilas, MAINlistaRepetir = leerTablaDBF(elArchivo, TRNSseleccionadosParaRepetir, MAINlistaRepetir)
            if MAINlistaRepetir is None:
                print(f'clidconfig-> ATENCION: instalar dbfread para que clidaux pueda leer los dbf (p. ej. con la lista de ficheros pendientes)')
                sys.exit(0)
            elif MAINlistaRepetir != []:
                if self.LCLverbose:
                    print(
                        'Se van a procesar %i ficheros laz de un total de %i que hay que repetir (ver TRNSseleccionadosParaRepetir)' % len(MAINlistaRepetir, nfilas)
                    )
                    print(MAINlistaRepetir)
        self.MAINlistaRepetir = MAINlistaRepetir
        self.configVarsDict['MAINlistaRepetir'] = [MAINlistaRepetir, 'GrupoMAIN', '', 'list', MAINlistaRepetir]


    # ==========================================================================
    def completarVariablesGlobalesRutas(self, rutaLazCompleta=''):
        # A este metodo se le llama desde clidbase.py despues de leer los argumentos en linea de comandos

        # ======================================================================
        # En esta funcion se establecen algunas variables globales (si es necesario modificarlas):
        #    MAINrutaCarto      -> Se adapta para que cuelgue de MAIN_RAIZ_DIR o se adapta segun MAINprocedimiento
        #    MAINrutaOutput     -> Se vincula a MAIN_RAIZ_DIR (windows) o se establece especificamente (calendula) y se adapta segun MAINprocedimiento
        #    MAINrutaLaz        -> Se vincula a MAIN_RAIZ_DIR (windows) o se establece especificamente (calendula) y se adapta segun MAINprocedimiento 
        #    MAINprocedimiento  -> Se retoca si el procedimiento menciona calendula pero lo ejecuto en Windows
        #    GLBLshapeNumPoints, GLBLtipoLectura, GLBLshapeFilter, GLBLprocesarComprimidosLaz, etc.
        # Y se devuelven variables (sin utilidad para determinados procedimiento)
        #    listaDirsLaz, listaSubDirsLaz, coordenadasDeMarcos
        # ======================================================================
    
        # ======================================================================
        # Estos parametros los obtengo al importar clidaux y quedan guardados en el .cfg
        # de forma que clidbase, clidflow y demas modulos los tienen en su GLO
        # Ademas, estan disponibles en este metodo que ha sido llamado desde clidbase:
        #     MAIN_copyright, MAIN_version, MAIN_idProceso, MAINusuario,
        #     MAINmiRutaProyecto, MAIN_ENTORNO, MAIN_PC, MAIN_DRIVE,
        #     MAIN_HOME_DIR, MAIN_FILE_DIR, MAIN_PROJ_DIR,
        #     MAIN_RAIZ_DIR, MAIN_THIS_DIR, MAIN_WORK_DIR,
        #     MAINrutaRaizHome (=MAIN_RAIZ_DIR), MAINrutaRaizData
        #       En calendula MAIN_RAIZ_DIR es HOME (=MAINrutaRaizHome)
        #       En calendula MAINrutaRaizData es SCRATCH
        # ======================================================================
        # Ruta raiz del proyecto:    MAIN_RAIZ_DIR
        #  En windows:
        #    Si cartolid NO esta instalado -> D:/_clid (depende de donde este clidbase.py)
        #    Si cartolid SI esta instalado -> C:\conda\py37\envs\clid\lib (depende de donde este site-packages)
        #  En calendula:
        #    /LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1
        # ======================================================================
        # Ruta raiz para resultados: MAINrutaRaizData
        #  -> Esta variable solo se usa desde clidaux.py, para
        #     En windows: normalmente sera D:/_clid (o lo que diga clidbase.xlsx)
        #       Tiro de self.MAINrutaRaizHome (de clidbase.xlsx), que lo he creado para 
        #       cuando se ejecuta el cartolidar de site-packages, ya que no
        #       quiero que los log, cfg, y resultados vayan a site-packages  
        #     En calendula:
        #       /scratch/jcyl_spi_1/jcyl_spi_1_1
        # ======================================================================
    
        if __verbose__:
            print(f'\n{"":_^80}')
            print(f'{" clidconfig-> Revisando rutas MAIN 2 ":=^80}')
            print(f'{" y guardandolas en el fichero de configuracion cfg ":=^80}')
            print(f'{"":=^80}')
            print(f'{TB}-> MAINrutaLaz: se basa en:')
            print(f'{TB}{TV}-> MAINrutaRaizData (asignado en clidaux.py)')
            print(f'{TB}{TV}-> MAINrutaLaz de configuracion (.cfg o .xlsx)')
            print(f'{TB}-> MAINlistaDirsLaz: se basa en:')
            print(f'{TB}{TV}-> MAINincluirRaizEnListaDirsLaz')
            print(f'{TB}{TV}-> MAINlistaDirsLaz')
            print(f'{TB}{TV}-> MAINcuadrante si MAINlistaDirsLaz="CUADRANTE"')
            print(f'{TB}-> MAINlistaSubDirsLaz se basa en:')
            print(f'{TB}{TV}-> MAINincluirRaizEnListaSubdirsLaz')
            print(f'{TB}{TV}-> MAINlistaDirsLaz')
            print(f'{TB}-> Si MAINusarCasosEspecialesParaMAINrutaLaz')
            print(f'{TB}{TV}-> MAINlistaDirsLaz')
            print(f'{TB}-> MAINrutaCarto se basa en:')
            print(f'{TB}{TV}-> MAINrutaCarto')
            print(f'{TB}{TV}-> asignarMAINrutaCarto')
            print(f'{TB}->  MAINrutaOutput se basa en:')
            print(f'{TB}{TV}-> MAINrutaOutput')
            print(f'{TB}{TV}-> asignarMAINrutaOutput')
            print(f'{TB}-> MAIN_MDLS_DIR se basa en:')
            print(f'{TB}{TV}-> MAIN_ENTORNO')
            print(f'{TB}{TV}-> MAINrutaRaizHome (o, si no existe, MAIN_RAIZ_DIR)')
            print(f'{TB}-> GLBL_TRAIN_DIR se basa en:')
            print(f'{TB}{TV}-> MAINrutaOutput')
            print(f'{TB}{TV}-> subDirTrain (se basa en GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes y GLBLcrearTilesTargetMiniSubCelSoloSiHayNoSueloSuficientes)')
            print(f'{"":=^80}')

        # ==========================================================================
        # =========================== MAINrutaLaz ==================================
        # ==========================================================================
        # Sitio por defecto para MAINrutaLaz
        # No se recorren subcarpetas de MAINrutaLaz salvo que lo establezca el MAINprocedimiento
        if rutaLazCompleta != '':
            # Solo cuando se inicia con clidflow
            self.MAINrutaLaz = os.path.abspath(rutaLazCompleta)
            if __verbose__:
                print(f'\n{"":_^80}')
                print(f'clidaux-> Identificando la ruta de los ficheros laz con completarVariablesGlobales<>')
                print(f'{TB}-> Valor de inicial de rutaLazCompleta: <{rutaLazCompleta}>')
                print(f'{TB}-> Para self.MAINrutaLaz se adopta el valor de rutaLazCompleta: {self.MAINrutaLaz}')
        else:
            if __verbose__:
                print(f'\n{"":_^80}')
                print(f'clidaux-> Identificando la ruta de los ficheros laz con completarVariablesGlobales<>')
                print(f'{TB}-> Valor de inicial de self.MAINrutaLaz: <{self.MAINrutaLaz}>')
            if (not self.MAINrutaLaz is None
                and self.MAINrutaLaz != 'None'
                and  self.MAINrutaLaz != ''
            ):
                if MAIN_ENTORNO == 'calendula':
                    if ':' in self.MAINrutaLaz:
                        # La MAINrutaLaz de configuracion (.cfg, .xlsx) no esta pensada para calendula
                        # Para evitar errores se prueba con la ruta por defecto (laz2):
                        print(f'clidconfig: ATENCION: MAINrutaLaz ({self.MAINrutaLaz}) no valida para calendula; se adopta laz2')
                        self.MAINrutaLaz = os.path.join(self.MAINrutaRaizData, 'laz2')
                        print(f'{TB}-> Se adopta laz2: {self.MAINrutaLaz}')
                    else:
                        self.MAINrutaLaz = os.path.join(self.MAINrutaRaizData, self.MAINrutaLaz)
                        if __verbose__:
                            print(f'{TB}{TV}MAINrutaLaz-> Se integra MAINrutaRaizData y MAINrutaLaz: {self.MAINrutaLaz}')
                else:
                    if ':' in self.MAINrutaLaz:
                        # self.MAINrutaLaz = 'E:/incendioZamora'
                        if __verbose__:
                            print(f'{TB}{TV}MAINrutaLaz-> Se usa directamente self.MAINrutaLaz: {self.MAINrutaLaz}')
                    else:
                        self.MAINrutaLaz = os.path.join(self.MAINrutaRaizData, self.MAINrutaLaz)
                        if __verbose__:
                            print(f'{TB}{TV}MAINrutaLaz-> Se integra MAINrutaRaizData y self.MAINrutaLaz: {self.MAINrutaLaz}')
            else:
                if MAIN_ENTORNO == 'calendula':
                    self.MAINrutaLaz = os.path.join(self.MAINrutaRaizData, 'laz')
                    if __verbose__:
                        print(f'{TB}{TV}MAINrutaLaz-> Valor por defecto basado en MAINrutaRaizData: {self.MAINrutaLaz}')
                else:
                    self.MAINrutaLaz = os.path.join(self.MAINrutaRaizData, 'laz')
                    if __verbose__:
                        print(f'{TB}{TV}MAINrutaLaz-> Valor por defecto basado en rutaRaiz: {self.MAINrutaLaz}')
        if __verbose__:
            print(f'{"":=^80}')
        # ==========================================================================
        if self.MAINincluirRaizEnListaDirsLaz:
            listaDirsLaz = ['']
        else:
            listaDirsLaz = []
        if self.MAINlistaDirsLaz == 'CUADRANTE':
            if (self.MAINcuadrante)[:2].upper() == 'CE':
                listaDirsLaz.extend(['lasfile-ce', 'CE', 'ce'])
            elif (self.MAINcuadrante)[:2].upper() == 'NE':
                listaDirsLaz.extend(['lasfile-ne', 'NE', 'ne'])
            elif (self.MAINcuadrante)[:2].upper() == 'NW':
                listaDirsLaz.extend(['lasfile-nw', 'NW', 'nw'])
            elif (self.MAINcuadrante)[:2].upper() == 'SE':
                listaDirsLaz.extend(['lasfile-se', 'SE', 'se'])
            elif (self.MAINcuadrante)[:2].upper() == 'SW':
                listaDirsLaz.extend(['lasfile-sw', 'SW', 'sw'])
            elif (self.MAINcuadrante)[:2].upper() == 'XX':
                listaDirsLaz.extend(['lasfile-ce', 'lasfile-ne', 'lasfile-nw', 'lasfile-se', 'lasfile-sw'])
                listaDirsLaz.extend(['ce', 'ne', 'nw', 'se', 'sw', 'CE', 'NE', 'NW', 'SE', 'SW'])
                listaDirsLaz.extend(['lasfile-yy'])
                listaDirsLaz.extend(['lasfiles'])
            else:
                listaDirsLaz.extend([
                    'lasfile-ce', 'lasfile-nw', 'lasfile-ne', 'lasfile-se', 'lasfile-sw', 'roquedos',
                    'CE', 'NE', 'NW', 'SE', 'SW',
                    'ce', 'ne', 'nw', 'se', 'sw',
                ])
        elif (
            self.MAINlistaDirsLaz is None
            or self.MAINlistaDirsLaz == 'None'
            or self.MAINlistaDirsLaz == ''
        ):
            if listaDirsLaz == []:
                listaDirsLaz = ['']
        else:
            listaDirsLaz += [i for i in (self.MAINlistaDirsLaz).split('+')]
        # ======================================================================
        if self.MAINincluirRaizEnListaSubdirsLaz:
            listaSubDirsLaz = ['']
        else:
            listaSubDirsLaz = []
        if (
            self.MAINlistaSubDirsLaz is None
            or self.MAINlistaSubDirsLaz == 'None'
            or self.MAINlistaSubDirsLaz == ''
        ):
            if listaSubDirsLaz == []:
                listaSubDirsLaz = ['']
        else:
            listaSubDirsLaz += [i for i in (self.MAINlistaSubDirsLaz).split('+')]
        print(f'clidconfig-> Chequeando MAINincluirRaizEnListaSubdirsLaz: {self.MAINincluirRaizEnListaSubdirsLaz}')
        print(f'clidconfig-> Chequeando MAINlistaSubDirsLaz: {self.MAINlistaSubDirsLaz}')
        print(f'clidconfig-> Chequeando listaSubDirsLaz (a): {listaSubDirsLaz}')
        # ======================================================================
        if self.MAINusarCasosEspecialesParaMAINrutaLaz:
            if __verbose__:
                print(f'{"":=^80}')
                print(f'clidaux-> Se lanza casosEspecialesParaMAINrutaLaz<>')
            # Los casos especiales son:
            # if LCLprocedimiento == 'PRECONFIGURADO_SINRUTA':
            # elif LCLprocedimiento.startswith('AUTOMATICO_DISCOEXTERNO'):
            # elif LCLprocedimiento.startswith('AUTOMATICO_SIGMENA'):
            # elif LCLprocedimiento[:20] == 'AUTOMATICO_CUADRANTE':
            # elif (LCLprocedimiento == 'PRECONFIGURADO_CONRUTA_4CUADRANTES_SIGMENA_CONOSINMARCO'):
            # elif (LCLprocedimiento == 'PRECONFIGURADO_CONRUTA_4CUADRANTES_DISCOEXTERNO_CONOSINMARCO'):
            # elif LCLprocedimiento == 'MANUAL':
            # elif LCLprocedimiento.startswith('CREAR_CAPA_CON_UNA_PROPIEDAD_DE_LOS_FICHEROS_LIDAR'):
            # elif LCLprocedimiento == 'CREAR_SHAPE':
            # elif LCLprocedimiento == 'RENOMBRAR_FICHEROS':
            # elif LCLprocedimiento == 'MERGEAR':
            # elif LCLprocedimiento == 'GEOINTEGRAR':
            self.MAINrutaLaz, listaDirsLaz, listaSubDirsLaz = casosEspecialesParaMAINrutaLaz(
                self.MAINprocedimiento,
                self.MAINrutaLaz,
                listaDirsLaz,
                listaSubDirsLaz,
                self.MAINcuadrante,
                LCLverbose=__verbose__,
            )
        print(f'clidconfig-> Chequeando listaSubDirsLaz (b): {listaSubDirsLaz}')
        # ======================================================================
        if __verbose__ or True:
            print(f'clidconfig-> self.MAINrutaLaz: {self.MAINrutaLaz}')
            print(f'{TB}-> Parametros de configuracion (rutas):')
            print(f'{TB}{TV}-> self.MAINobjetivoEjecucion: {self.MAINobjetivoEjecucion}')
            print(f'{TB}{TV}-> self.MAINprocedimiento:     {self.MAINprocedimiento}')
            print(f'{TB}{TV}-> self.MAINcuadrante:         {self.MAINcuadrante}')

            print(f'{TB}{TV}-> listaDirsLaz:     {listaDirsLaz} {type(listaDirsLaz)}')
            print(f'{TB}{TV}-> MAINlistaDirsLaz: {self.MAINlistaDirsLaz} {type(self.MAINlistaDirsLaz)}')
            print(f'{TB}{TV}-> listaSubDirsLaz:     {type(listaSubDirsLaz)} {listaSubDirsLaz}')
            print(f'{TB}{TV}-> MAINlistaSubDirsLaz: {type(self.MAINlistaSubDirsLaz)} {self.MAINlistaSubDirsLaz}')
            print(f'{TB}{TV}-> MAINincluirRaizEnListaDirsLaz:    {self.MAINincluirRaizEnListaDirsLaz}')
            print(f'{TB}{TV}-> MAINincluirRaizEnListaSubdirsLaz: {self.MAINincluirRaizEnListaSubdirsLaz}')
            print(f'{TB}{TV}-> MAINusarCasosEspecialesParaMAINrutaLaz: {self.MAINusarCasosEspecialesParaMAINrutaLaz}')
            print(f'{"":=^80}')
        # ======================================================================
    
        # ======================================================================
        # ========================== MAINrutaCarto =============================
        # ======================================================================
        if __verbose__:
            print(f'clidconfig-> Asignacion de MAINrutaCarto')
            print(f'{TB}Antes:')
            print(f'{TB}-> self.MAINrutaCarto: {self.MAINrutaCarto}')
        if (
            not self.MAINrutaCarto is None
            and self.MAINrutaCarto != 'None'
            and self.MAINrutaCarto != ''
        ):
            if __verbose__:
                print(f'{TB}{TV}-> self.MAINrutaCarto leida de fichero de configuracion NO es nula: {type(self.MAINrutaCarto)}')
            self.MAINrutaCarto = os.path.abspath(self.MAINrutaCarto)
        else:
            if __verbose__:
                print(f'{TB}{TV}-> Como self.MAINrutaCarto SI es nulo; se asigna a partir de MAINrutaRaizHome.')
            self.MAINrutaCarto = self.asignarMAINrutaCarto()
        if __verbose__:
            print(f'{TB}Desp:')
            print(f'{TB}-> self.MAINrutaCarto: {self.MAINrutaCarto}')
        # ==========================================================================
    
        # ==========================================================================
        # ========================== MAINrutaOutput ================================
        # ==========================================================================
        if __verbose__:
            print(f'clidaux-> Asignacion de MAINrutaOutput')
        if (
            not self.MAINrutaOutput is None
            and self.MAINrutaOutput != 'None'
            and self.MAINrutaOutput != ''
        ):
            if __verbose__:
                print(f'{TB}-> MAINrutaOutput (a1): {self.MAINrutaOutput}')
            self.MAINrutaOutput = os.path.abspath(self.MAINrutaOutput)
            if __verbose__:
                print(f'{TB}-> MAINrutaOutput (a2): {self.MAINrutaOutput}')
        else:
            if __verbose__:
                print(f'{TB}-> MAINrutaOutput (b1): {self.MAINrutaOutput}')
                print(f'{TB}-> MAINobjetivoSiReglado (b): {self.MAINobjetivoSiReglado}')
            self.MAINrutaOutput = self.asignarMAINrutaOutput(
                self.MAINprocedimiento,
                self.MAINrutaOutput,
                self.MAINrutaRaizData,
                self.MAINobjetivoSiReglado,
                self.MAINcuadrante,
            )
            if __verbose__:
                print(f'{TB}-> MAINrutaOutput (b2): {self.MAINrutaOutput}')
        # ==========================================================================
        if self.MAINrutaOutput is None:
            print(f'clidaux-> ATENCION: no se ha asignado correctamente MAINrutaOutput: {self.MAINrutaOutput}')
            print(f'{TB}-> Revisar codigo')
            sys.exit(0)
        elif not os.path.exists(self.MAINrutaOutput):
            print(f'clidaux-> No existe el directorio {self.MAINrutaOutput} -> Se crea automaticamente')
            try:
                os.makedirs(self.MAINrutaOutput)
            except:
                print(f'{TB}-> No se ha podido crear el directorio {self.MAINrutaOutput}. Revisar MAINprocedimiento')
                sys.exit(0)
        # ==========================================================================

        # ==========================================================================
        # ================== MAIN_MDLS_DIR, GLBL_TRAIN_DIR =========================
        # ==========================================================================
        if (
            self.GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes
            or self.GLBLcrearTilesTargetMiniSubCelSoloSiHayNoSueloSuficientes
        ):
            subDirTrain = 'trainSel'
        else:
            subDirTrain = 'trainAll'
        # ==========================================================================
        if MAIN_ENTORNO == 'calendula':
            self.MAIN_MDLS_DIR = '/LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1/data'
            # self.MAIN_MDLS_DIR = os.path.abspath(os.path.join(MAIN_RAIZ_DIR, 'cartolidar/cartolidar/data'))
            self.GLBL_TRAIN_DIR = os.path.join(self.MAINrutaOutput, subDirTrain)
        elif MAIN_ENTORNO == 'colab':
            self.MAIN_MDLS_DIR = os.path.join(MAIN_RAIZ_DIR, 'data')
            self.GLBL_TRAIN_DIR = os.path.join(MAIN_RAIZ_DIR, 'data/datasets/cartolid/trainImg')
        elif 'MAINrutaRaizHome' in dir(GLO):
            if 'cartolidar' in MAIN_RAIZ_DIR:
                self.MAIN_MDLS_DIR = os.path.abspath(os.path.join(self.MAINrutaRaizHome, '../data'))
            else:
                self.MAIN_MDLS_DIR = os.path.join(self.MAINrutaRaizHome, 'data')
            self.GLBL_TRAIN_DIR = os.path.join(self.MAINrutaOutput, subDirTrain)
        else:
            if 'cartolidar' in MAIN_RAIZ_DIR:
                self.MAIN_MDLS_DIR = os.path.abspath(os.path.join(MAIN_RAIZ_DIR, '../data'))
            else:
                self.MAIN_MDLS_DIR = os.path.abspath(os.path.join(MAIN_RAIZ_DIR, 'data'))
            if MAIN_PC == 'JCyL':
                # Imagenes de entrenamiento en disco externo
                # self.GLBL_TRAIN_DIR = 'D:/trainImg'
                # self.GLBL_TRAIN_DIR = os.path.join(RAIZ_DIR, 'cartolidout/train')
                # self.GLBL_TRAIN_DIR = os.path.join(self.MAINrutaOutput, subDirTrain)
                # Imagenes de entrenamiento en disco duro
                # self.GLBL_TRAIN_DIR = os.path.join(RAIZ_DIR, 'data/datasets/cartolid/trainImg')
                # self.GLBL_TRAIN_DIR = 'C:/_ws/cartolidout/train'
                self.GLBL_TRAIN_DIR = os.path.join(self.MAINrutaOutput, subDirTrain)
            else:
                # En casa
                # self.GLBL_TRAIN_DIR = os.path.join(RAIZ_DIR, 'data/trainImg')
                self.GLBL_TRAIN_DIR = os.path.join(self.MAINrutaOutput, subDirTrain)
        # ==========================================================================

        self.paramConfigAdicionalesGLBL['MAIN_MDLS_DIR'] = [
            self.MAIN_MDLS_DIR, 'str', '', 'GrupoDirsFiles', self.MAIN_MDLS_DIR,
            self.MAIN_MDLS_DIR, self.MAIN_MDLS_DIR, self.MAIN_MDLS_DIR,
            self.MAIN_MDLS_DIR, self.MAIN_MDLS_DIR, self.MAIN_MDLS_DIR,
        ]
        self.paramConfigAdicionalesGLBL['GLBL_TRAIN_DIR'] = [
            self.GLBL_TRAIN_DIR, 'str', '', 'GrupoDirsFiles', self.GLBL_TRAIN_DIR,
            self.GLBL_TRAIN_DIR, self.GLBL_TRAIN_DIR, self.GLBL_TRAIN_DIR,
            self.GLBL_TRAIN_DIR, self.GLBL_TRAIN_DIR, self.GLBL_TRAIN_DIR,
        ]
        
        coordenadasDeMarcos = {}
        return listaDirsLaz, listaSubDirsLaz, coordenadasDeMarcos


    # ==========================================================================
    def completarVariablesGlobalesFicherosDeInfoControlGral(self):
        # ==========================================================================
        # ==================== GLBLficheroDeControlGral ============================
        # ==========================================================================
        # self.MAINusuario se asigna al importar clidaux.py desde clidbase.py
        # y se guarda en el fichero .cfg previamente creado en clidconfig.py
        if __verbose__:
            print(f'\n{"":_^80}')
            print(f'{" clidconfig-> Revisando rutas MAIN 3 ":=^80}')
            print(f'{" y guardandolas en el fichero de configuracion cfg ":=^80}')
            print(f'{"":=^80}')
            print(f'{TB}-> GLBLficheroDeControlGral, GLBLficheroDeInfoLas, GLBLficheroMuestraPuntos y GLBLficheroListaBloques se basan en:')
            print(f'{TB}-> MAINrutaOutput: {self.MAINrutaOutput}')
            print(f'{TB}-> MAINusuario:    {self.MAINusuario}')

        # ======================================================================
        self.GLBLficheroDeControlGral = os.path.join(
            self.MAINrutaOutput,
            'GlobalControl_{}.txt'.format(self.MAINusuario)
        )
        creaRutaDeFichero(self.GLBLficheroDeControlGral)
        self.paramConfigAdicionalesGLBL['GLBLficheroDeControlGral'] = [
            self.GLBLficheroDeControlGral, 'str', '', 'GrupoDirsFiles', self.GLBLficheroDeControlGral,
            self.GLBLficheroDeControlGral, self.GLBLficheroDeControlGral, self.GLBLficheroDeControlGral,
            self.GLBLficheroDeControlGral, self.GLBLficheroDeControlGral, self.GLBLficheroDeControlGral,
        ]
        # ======================================================================
        if (
            'LAS_INFO' in self.MAINprocedimiento
            or 'LAS_EDIT' in self.MAINprocedimiento
            or 'PASADAS' in self.MAINprocedimiento
            or True
        ):
            self.GLBLficheroDeInfoLasH30 = os.path.join(
                self.MAINrutaOutput,
                'InfoLasDatabaseH30_{}.csv'.format(self.MAINusuario)
            )
            self.GLBLficheroDeInfoLasH29 = os.path.join(
                self.MAINrutaOutput,
                'InfoLasDatabaseH29_{}.csv'.format(self.MAINusuario)
            )
            creaRutaDeFichero(self.GLBLficheroDeInfoLasH30)
            self.paramConfigAdicionalesGLBL['GLBLficheroDeInfoLasH30'] = [
                self.GLBLficheroDeInfoLasH30, 'str', '', 'GrupoDirsFiles', self.GLBLficheroDeInfoLasH30,
                self.GLBLficheroDeInfoLasH30, self.GLBLficheroDeInfoLasH30, self.GLBLficheroDeInfoLasH30,
                self.GLBLficheroDeInfoLasH30, self.GLBLficheroDeInfoLasH30, self.GLBLficheroDeInfoLasH30,
            ]
            self.paramConfigAdicionalesGLBL['GLBLficheroDeInfoLasH29'] = [
                self.GLBLficheroDeInfoLasH29, 'str', '', 'GrupoDirsFiles', self.GLBLficheroDeInfoLasH29,
                self.GLBLficheroDeInfoLasH29, self.GLBLficheroDeInfoLasH29, self.GLBLficheroDeInfoLasH29,
                self.GLBLficheroDeInfoLasH29, self.GLBLficheroDeInfoLasH29, self.GLBLficheroDeInfoLasH29,
            ]

            self.GLBLficheroMuestraPuntosH30 = os.path.join(
                self.MAINrutaOutput,
                'MuestraPuntosH30_{}.csv'.format(self.MAINusuario)
            )
            self.GLBLficheroMuestraPuntosH29 = os.path.join(
                self.MAINrutaOutput,
                'MuestraPuntosH29_{}.csv'.format(self.MAINusuario)
            )
            creaRutaDeFichero(self.GLBLficheroMuestraPuntosH30)
            self.paramConfigAdicionalesGLBL['GLBLficheroMuestraPuntosH30'] = [
                self.GLBLficheroMuestraPuntosH30, 'str', '', 'GrupoDirsFiles', self.GLBLficheroMuestraPuntosH30,
                self.GLBLficheroMuestraPuntosH30, self.GLBLficheroMuestraPuntosH30, self.GLBLficheroMuestraPuntosH30,
                self.GLBLficheroMuestraPuntosH30, self.GLBLficheroMuestraPuntosH30, self.GLBLficheroMuestraPuntosH30,
            ]
            self.paramConfigAdicionalesGLBL['GLBLficheroMuestraPuntosH29'] = [
                self.GLBLficheroMuestraPuntosH29, 'str', '', 'GrupoDirsFiles', self.GLBLficheroMuestraPuntosH29,
                self.GLBLficheroMuestraPuntosH29, self.GLBLficheroMuestraPuntosH29, self.GLBLficheroMuestraPuntosH29,
                self.GLBLficheroMuestraPuntosH29, self.GLBLficheroMuestraPuntosH29, self.GLBLficheroMuestraPuntosH29,
            ]

            self.GLBLficheroListaBloquesH30 = os.path.join(
                self.MAINrutaOutput,
                'ListaBloquesH30_{}.csv'.format(self.MAINusuario)
            )
            self.GLBLficheroListaBloquesH29 = os.path.join(
                self.MAINrutaOutput,
                'ListaBloquesH29_{}.csv'.format(self.MAINusuario)
            )
            creaRutaDeFichero(self.GLBLficheroListaBloquesH30)
            self.paramConfigAdicionalesGLBL['GLBLficheroListaBloquesH30'] = [
                self.GLBLficheroListaBloquesH30, 'str', '', 'GrupoDirsFiles', self.GLBLficheroListaBloquesH30,
                self.GLBLficheroListaBloquesH30, self.GLBLficheroListaBloquesH30, self.GLBLficheroListaBloquesH30,
                self.GLBLficheroListaBloquesH30, self.GLBLficheroListaBloquesH30, self.GLBLficheroListaBloquesH30,
            ]
            self.paramConfigAdicionalesGLBL['GLBLficheroListaBloquesH29'] = [
                self.GLBLficheroListaBloquesH29, 'str', '', 'GrupoDirsFiles', self.GLBLficheroListaBloquesH29,
                self.GLBLficheroListaBloquesH29, self.GLBLficheroListaBloquesH29, self.GLBLficheroListaBloquesH29,
                self.GLBLficheroListaBloquesH29, self.GLBLficheroListaBloquesH29, self.GLBLficheroListaBloquesH29,
            ]        # ======================================================================

        if __verbose__:
            print(f'clidconfig-> GLBLficheroDeControlGral: {self.GLBLficheroDeControlGral}')
            print(f'clidconfig-> GLBLficheroDeInfoLasH30:     {self.GLBLficheroDeInfoLasH30}')
            print(f'clidconfig-> GLBLficheroDeInfoLasH29:     {self.GLBLficheroDeInfoLasH29}')
            print(f'clidconfig-> GLBLficheroMuestraPuntosH30: {self.GLBLficheroMuestraPuntosH30}')
            print(f'clidconfig-> GLBLficheroMuestraPuntosH29: {self.GLBLficheroMuestraPuntosH29}')
            print(f'clidconfig-> GLBLficheroListaBloquesH30:  {self.GLBLficheroListaBloquesH30}')
            print(f'clidconfig-> GLBLficheroListaBloquesH29:  {self.GLBLficheroListaBloquesH29}')
            print(f'{"":=^80}')


    # ==========================================================================
    def completarVariablesGlobalesNombresDeModelosConvolucionales(self):

        if __verbose__:
            print(f'\n{"":_^80}')
            print(f'{" clidconfig-> Asignando Variables MAIN_ Nombres De Modelos Convolucionales ":=^80}')
            print(f'{" y guardandolas en el fichero de configuracion cfg ":=^80}')
            print(f'{"":=^80}')
            if self.GLBLpredecirCubiertasSingularesConvolucional or self.GLBLpredecirClasificaMiniSubCelConvolucional:
                print(f'{TB}-> MAIN_COD_... y MAIN_LISTA_...')
            print(f'{"":=^80}')

        # ==========================================================================
        # Nombres de los modelos convolucionales: genero las variables
        # MAIN_COD_... y MAIN_LISTA_..., que guardo en el ichero de configuracion
        # y son las que uso en otros modulos (en vez de GLBLmodeloCartolid...)
        # ==========================================================================
        if self.GLBLpredecirCubiertasSingularesConvolucional or self.GLBLpredecirClasificaMiniSubCelConvolucional:
            if (
                not self.GLBLmodeloCartolidMiniSubCelEntrenado is None
                and self.GLBLmodeloCartolidMiniSubCelEntrenado != ''
            ):
                idCuadranteActual = '_{}'.format((self.MAINcuadrante)[:2].upper())
                indexCuadranteModelo = self.GLBLmodeloCartolidMiniSubCelEntrenado.find('_Png') - 4
                if indexCuadranteModelo > 0:
                    idCuadranteModelo = self.GLBLmodeloCartolidMiniSubCelEntrenado[indexCuadranteModelo: indexCuadranteModelo + 3]
                    self.GLBLmodeloCartolidMiniSubCelEntrenado = self.GLBLmodeloCartolidMiniSubCelEntrenado.replace(
                        idCuadranteModelo,
                        idCuadranteActual
                    )
            if (
                not self.GLBLmodeloCartolidCartoSinguEntrenadoA is None
                and self.GLBLmodeloCartolidCartoSinguEntrenadoA != ''
            ):
                idCuadranteActual = '_{}'.format((self.MAINcuadrante)[:2].upper())
                indexCuadranteModelo = self.GLBLmodeloCartolidCartoSinguEntrenadoA.find('_Png') - 4
                if indexCuadranteModelo > 0:
                    idCuadranteModelo = self.GLBLmodeloCartolidCartoSinguEntrenadoA[indexCuadranteModelo: indexCuadranteModelo + 3]
                    self.GLBLmodeloCartolidCartoSinguEntrenadoA = self.GLBLmodeloCartolidCartoSinguEntrenadoA.replace(
                        idCuadranteModelo,
                        idCuadranteActual
                    )
            if (
                not self.GLBLmodeloCartolidCartoSinguEntrenadoB is None
                and self.GLBLmodeloCartolidCartoSinguEntrenadoB != ''
            ):
                idCuadranteActual = '_{}'.format((self.MAINcuadrante)[:2].upper())
                indexCuadranteModelo = self.GLBLmodeloCartolidCartoSinguEntrenadoB.find('_Png') - 4
                if indexCuadranteModelo > 0:
                    idCuadranteModelo = self.GLBLmodeloCartolidCartoSinguEntrenadoB[indexCuadranteModelo: indexCuadranteModelo + 3]
                    self.GLBLmodeloCartolidCartoSinguEntrenadoB = self.GLBLmodeloCartolidCartoSinguEntrenadoB.replace(
                        idCuadranteModelo,
                        idCuadranteActual
                    )
    
            # Nombre del modelo acumulativo
            if self.MAINcuadrante != '' and self.MAINnInputsModeloNln != 0 and self.MAINcodModeloNln != '':
                # Si se incluyen estos argumentos en linea de comandos, prevalecen sobre el nombre del fichero xls de configuracion
                nInputVars = self.MAINnInputsModeloNln
                if self.MAINcodModeloNln == '64+32+16_resBatchN':
                    self.GLBLnombreFicheroConModeloParaInferencia = 'NNCaleLote{}_i{:03}_h64_h32_h16_o10_dropoutPriUlt_normHL_RNN'.format(
                        self.MAINcuadrante.upper(),
                        nInputVars
                    )
                else:
                    print('clidconfig-> ATENCION: codigo de modelo en linea de comandos no implementado: {}'.format(self.MAINcodModeloNln))
                    sys.exit(0)
                print('clidconfig-> Nombre del modelo: {}'.format(self.GLBLnombreFicheroConModeloParaInferencia))
            elif (
                not self.GLBLnombreFicheroConModeloParaInferencia is None
                and self.GLBLnombreFicheroConModeloParaInferencia != ''
            ):
                idCuadranteActual = 'Lote{}'.format((self.MAINcuadrante)[:2].upper())
                indexCuadranteModelo = self.GLBLnombreFicheroConModeloParaInferencia.find('_i') - 6
                if indexCuadranteModelo > 0:
                    idCuadranteModelo = self.GLBLnombreFicheroConModeloParaInferencia[indexCuadranteModelo: indexCuadranteModelo + 6]
                    self.GLBLnombreFicheroConModeloParaInferencia = self.GLBLnombreFicheroConModeloParaInferencia.replace(
                        idCuadranteModelo,
                        idCuadranteActual
                    )
    
            # self.GLBLmodeloCartolidMiniSubCelEntrenado = 'clidGen_cale_LasClass_2_345_6_reDepurada_Png6_012345_20210605v0.h5'
            if self.GLBLmodeloCartolidMiniSubCelEntrenado:
                iniDataset = 13
                if '__Png' in self.GLBLmodeloCartolidMiniSubCelEntrenado:
                    finDataset = self.GLBLmodeloCartolidMiniSubCelEntrenado.find('__Png')
                elif '_Png' in self.GLBLmodeloCartolidMiniSubCelEntrenado:
                    finDataset = self.GLBLmodeloCartolidMiniSubCelEntrenado.find('_Png')
                else:
                    finDataset = self.GLBLmodeloCartolidMiniSubCelEntrenado.find('.h5')
                txtDataset = self.GLBLmodeloCartolidMiniSubCelEntrenado[iniDataset: finDataset]
                if '_Png' in self.GLBLmodeloCartolidMiniSubCelEntrenado: # Incluye '__Png'
                    iniNumPngs = self.GLBLmodeloCartolidMiniSubCelEntrenado.find('_Png') + 1
                    finNumPngs = iniNumPngs + 4
                    txtNumPngs = self.GLBLmodeloCartolidMiniSubCelEntrenado[iniNumPngs: finNumPngs]
                    intNumPngs = int(txtNumPngs[-1])
                    iniLstPngs = finNumPngs + 1
                    finLstPngs = iniLstPngs + intNumPngs
                    txtLstPngs = self.GLBLmodeloCartolidMiniSubCelEntrenado[iniLstPngs: finLstPngs]
                else:
                    print('clidaux-> ATENCION: revisar el nombre del modelo entrenado (no incluye _PngX): {}'.format(self.GLBLmodeloCartolidMiniSubCelEntrenado))
                    intNumPngs = 0
                    txtLstPngs = ''
    
                MAIN_COD_18_16N_04_MODELOENTRENADO_MINI = 'cartolidMiniSubCel{}{}'.format(txtDataset, txtNumPngs)
                MAIN_LISTA_PNGS_MODELOENTRENADO_MINI = 'X{}'.format(txtLstPngs)
                # print('clidaux-> MAIN_COD_18_16N_04_MODELOENTRENADO_MINI:', MAIN_COD_18_16N_04_MODELOENTRENADO_MINI)
                # print('txtDataset:', txtDataset)
                # print('txtNumPngs:', txtNumPngs)
                # print('txtLstPngs:', txtLstPngs)
    
            # self.GLBLmodeloCartolidCartoSinguEntrenadoA = 'clidGen_cale_UsosDisp12_45678_Png6_012345_20210612v0.h5.h5'
            # self.GLBLmodeloCartolidCartoSinguEntrenadoA = 'clidGen_cale_UsosDisp_2_45678_Png6_012345_20210202.h5.h5'
            if self.GLBLmodeloCartolidCartoSinguEntrenadoA:
                if '__Png' in self.GLBLmodeloCartolidCartoSinguEntrenadoA:
                    finDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoA.find('__Png')
                elif '_Png' in self.GLBLmodeloCartolidCartoSinguEntrenadoA:
                    finDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoA.find('_Png')
                else:
                    finDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoA.find('.h5')
                txtDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoA[iniDataset: finDataset] # Puede ser: UsosDisp12_45678, UsosDisp_2_45678
                if '_Png' in self.GLBLmodeloCartolidCartoSinguEntrenadoA:
                    iniNumPngs = self.GLBLmodeloCartolidCartoSinguEntrenadoA.find('_Png') + 1
                    finNumPngs = iniNumPngs + 4
                    txtNumPngs = self.GLBLmodeloCartolidCartoSinguEntrenadoA[iniNumPngs: finNumPngs]
                    intNumPngs = int(txtNumPngs[-1])
                    iniLstPngs = finNumPngs + 1
                    finLstPngs = iniLstPngs + intNumPngs
                    txtLstPngs = self.GLBLmodeloCartolidCartoSinguEntrenadoA[iniLstPngs: finLstPngs]
                else:
                    print('clidaux-> ATENCION: revisar el nombre del modelo entrenadoA (no incluye _PngX): {}'.format(self.GLBLmodeloCartolidCartoSinguEntrenadoA))
                    intNumPngs = 0
                    txtLstPngs = ''
                MAIN_COD_18_16N_04_MODELOENTRENADO_CART_ = 'cartolidCartoSingu{}{}'.format(txtDataset, txtNumPngs)
                MAIN_COD_18_16N_04_MODELOENTRENADO_CARTA = 'cartolidCartoSingu{}{}'.format(txtDataset, txtNumPngs)
                MAIN_LISTA_PNGS_MODELOENTRENADO_CART_ = 'X{}'.format(txtLstPngs)
                MAIN_LISTA_PNGS_MODELOENTRENADO_CARTA = 'X{}'.format(txtLstPngs)
            else:
                MAIN_COD_18_16N_04_MODELOENTRENADO_CART_ = 'cartolidCartoSingu_'
                MAIN_COD_18_16N_04_MODELOENTRENADO_CARTA = 'cartolidCartoSingu_'
                MAIN_LISTA_PNGS_MODELOENTRENADO_CART_ = 'X123456'
                MAIN_LISTA_PNGS_MODELOENTRENADO_CARTA = 'X123456'
            # print('clidaux-> MAIN_COD_18_16N_04_MODELOENTRENADO_CARTA:', MAIN_COD_18_16N_04_MODELOENTRENADO_CARTA)
            # print('txtDataset:', txtDataset)
            # print('txtNumPngs:', txtNumPngs)
            # print('txtLstPngs:', txtLstPngs)
        
            # self.GLBLmodeloCartolidCartoSinguEntrenadoB = 'clidGen_cale_UsosDisp___45678_Png6_012345_20210202.h5.h5'
            if self.GLBLmodeloCartolidCartoSinguEntrenadoB:
                if '__Png' in self.GLBLmodeloCartolidCartoSinguEntrenadoB:
                    finDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoB.find('__Png')
                elif '_Png' in self.GLBLmodeloCartolidCartoSinguEntrenadoB:
                    finDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoB.find('_Png')
                else:
                    finDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoB.find('.h5')
                txtDataset = self.GLBLmodeloCartolidCartoSinguEntrenadoB[iniDataset: finDataset] # Puede ser: UsosDisp12_45678, UsosDisp_2_45678
                if '_Png' in self.GLBLmodeloCartolidCartoSinguEntrenadoB:
                    iniNumPngs = self.GLBLmodeloCartolidCartoSinguEntrenadoB.find('_Png') + 1
                    finNumPngs = iniNumPngs + 4
                    txtNumPngs = self.GLBLmodeloCartolidCartoSinguEntrenadoB[iniNumPngs: finNumPngs]
                    intNumPngs = int(txtNumPngs[-1])
                    iniLstPngs = finNumPngs + 1
                    finLstPngs = iniLstPngs + intNumPngs
                    txtLstPngs = self.GLBLmodeloCartolidCartoSinguEntrenadoB[iniLstPngs: finLstPngs]
                else:
                    print('clidaux-> ATENCION: revisar el nombre del modelo entrenadoB (no incluye _PngX): {}'.format(self.GLBLmodeloCartolidCartoSinguEntrenadoA))
                    intNumPngs = 0
                    txtLstPngs = ''
                MAIN_COD_18_16N_04_MODELOENTRENADO_CARTB = 'cartolidCartoSingu{}{}'.format(txtDataset, txtNumPngs)
                MAIN_LISTA_PNGS_MODELOENTRENADO_CARTB = 'X{}'.format(txtLstPngs)
            else:
                MAIN_COD_18_16N_04_MODELOENTRENADO_CARTB = 'cartolidCartoSingu_'
                MAIN_LISTA_PNGS_MODELOENTRENADO_CARTB = 'X123456'
            # print('clidaux-> MAIN_COD_18_16N_04_MODELOENTRENADO_CARTB:', MAIN_COD_18_16N_04_MODELOENTRENADO_CARTB)
            # print('txtDataset:', txtDataset)
            # print('txtNumPngs:', txtNumPngs)
            # print('txtLstPngs:', txtLstPngs)
        
            # self.GLBLmodeloCartolid128PixelesEntrenado = 'modeloPix2PixGeneratEntrenado_calendula_Pn6_012345_20201224.h5'
            txtDataset = 'SingUse128pixel'
            txtNumPngs = 'Png6'
            txtLstPngs = '012345'
            MAIN_COD_18_16N_04_MODELOENTRENADO_128P = 'cartolid128Pixeles{}{}'.format(txtDataset, txtNumPngs)
            MAIN_LISTA_PNGS_MODELOENTRENADO_128P = 'X{}'.format(txtLstPngs)
            # ======================================================================
    
            # ======================================================================
            if (
                (
                    not self.GLBLmodeloCartolidMiniSubCelEntrenado is None
                    and not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLmodeloCartolidMiniSubCelEntrenado
                )
                or (
                    not self.GLBLmodeloCartolidCartoSinguEntrenadoA is None
                    and not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLmodeloCartolidCartoSinguEntrenadoA
                )
                or (
                    not self.GLBLmodeloCartolidCartoSinguEntrenadoB is None
                    and not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLmodeloCartolidCartoSinguEntrenadoB
                )
                or (
                    not self.GLBLnombreFicheroConModeloParaInferencia is None
                    and not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLnombreFicheroConModeloParaInferencia
                )
        
                or (
                    not MAIN_COD_18_16N_04_MODELOENTRENADO_MINI is None
                    and not '{}_'.format((self.MAINcuadrante[:2]).upper()) in MAIN_COD_18_16N_04_MODELOENTRENADO_MINI
                )
                or (
                    not MAIN_COD_18_16N_04_MODELOENTRENADO_CART_ is None
                    and not '{}_'.format((self.MAINcuadrante[:2]).upper()) in MAIN_COD_18_16N_04_MODELOENTRENADO_CART_
                )
            ):
                print(f'\n{"":_^80}')
                print('clidaux-> Verificando modelos entrenados disponibles para el cuadrante {} (identificador completo: {}):'.format((self.MAINcuadrante[:2]).upper(), self.MAINcuadrante))
                print('\t-> GLBLmodeloCartolidMiniSubCelEntrenado:    {}'.format(self.GLBLmodeloCartolidMiniSubCelEntrenado))
                print('\t-> GLBLmodeloCartolidCartoSinguEntrenadoA:   {}'.format(self.GLBLmodeloCartolidCartoSinguEntrenadoA))
                print('\t-> GLBLmodeloCartolidCartoSinguEntrenadoB:   {}'.format(self.GLBLmodeloCartolidCartoSinguEntrenadoB))
                print('\t-> GLBLnombreFicheroConModeloParaInferencia: {}'.format(self.GLBLnombreFicheroConModeloParaInferencia))
                print('\t-> MAIN_COD_18_16N_04_MODELOENTRENADO_MINI:  {}'.format(MAIN_COD_18_16N_04_MODELOENTRENADO_MINI))
                print('\t-> MAIN_COD_18_16N_04_MODELOENTRENADO_CART_: {}'.format(MAIN_COD_18_16N_04_MODELOENTRENADO_CART_))
                print('\t-> cuadrante: {}_'.format((self.MAINcuadrante[:2]).upper()), '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLmodeloCartolidMiniSubCelEntrenado)
                if not self.GLBLmodeloCartolidMiniSubCelEntrenado is None:
                    if not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLmodeloCartolidMiniSubCelEntrenado:
                        print(
                            'clidaux-> ATENCION 1: el modelo <{}> no esta disponible para el cuadrante {}'.format(
                                self.GLBLmodeloCartolidMiniSubCelEntrenado,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                    else:
                        print(
                            '\tModelo1 <{}> entrenado OK para el cuadrante {}'.format(
                                self.GLBLmodeloCartolidMiniSubCelEntrenado,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                if not self.GLBLmodeloCartolidCartoSinguEntrenadoA is None:
                    if not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLmodeloCartolidCartoSinguEntrenadoA:
                        print(
                            'clidaux-> ATENCION 2: el modelo <{}> no esta disponible para el cuadrante {}'.format(
                                self.GLBLmodeloCartolidCartoSinguEntrenadoA,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                    else:
                        print(
                            '\tModelo2 <{}> entrenado OK para el cuadrante {}'.format(
                                self.GLBLmodeloCartolidCartoSinguEntrenadoA,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                if not self.GLBLmodeloCartolidCartoSinguEntrenadoB is None:
                    if not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLmodeloCartolidCartoSinguEntrenadoB:
                        print(
                            'Modelo3: el modelo <{}> no esta disponible para el cuadrante {} (no esencial)'.format(
                                self.GLBLmodeloCartolidCartoSinguEntrenadoB,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                    else:
                        print(
                            '\tModelo3 <{}> entrenado OK para el cuadrante {}'.format(
                                self.GLBLmodeloCartolidCartoSinguEntrenadoB,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                else:
                    print(
                        '\tModelo3 <{}> no se usa para el cuadrante {}'.format(
                            self.GLBLmodeloCartolidCartoSinguEntrenadoB,
                            '{}_'.format((self.MAINcuadrante[:2]).upper()),
                        )
                    )
        
                if (
                    self.MAINobjetivoEjecucion == 'CREAR_LAZ'
                    and not self.GLBLnombreFicheroConModeloParaInferencia is None
                ):
                    if not '{}_'.format((self.MAINcuadrante[:2]).upper()) in self.GLBLnombreFicheroConModeloParaInferencia:
                        # Solo necesito el modelo para inferencia si voy a CREAR_LAZ
                        print(
                            'clidaux-> ATENCION 4: el modelo <{}> no esta disponible para el cuadrante {}'.format(
                                self.GLBLnombreFicheroConModeloParaInferencia,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                    else:
                        print(
                            '\tModelo4 <{}> entrenado OK para el cuadrante {}'.format(
                                self.GLBLnombreFicheroConModeloParaInferencia,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                if not MAIN_COD_18_16N_04_MODELOENTRENADO_MINI is None:
                    if not '{}_'.format((self.MAINcuadrante[:2]).upper()) in MAIN_COD_18_16N_04_MODELOENTRENADO_MINI:
                        print(
                            '\tModelo5 <{}> no disponible para el cuadrante {} (no esencial)'.format(
                                MAIN_COD_18_16N_04_MODELOENTRENADO_MINI,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                    else:
                        print(
                            '\tModelo5 <{}> entrenado OK para el cuadrante {}'.format(
                                MAIN_COD_18_16N_04_MODELOENTRENADO_MINI,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
        
                if not MAIN_COD_18_16N_04_MODELOENTRENADO_CART_ is None:
                    if not '{}_'.format((self.MAINcuadrante[:2]).upper()) in MAIN_COD_18_16N_04_MODELOENTRENADO_CART_:
                        print(
                            '\tModelo6 <{}> no disponible para el cuadrante {} (no esencial)'.format(
                                MAIN_COD_18_16N_04_MODELOENTRENADO_CART_,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                    else:
                        print(
                            '\tModelo6 <{}> entrenado OK para el cuadrante {}'.format(
                                MAIN_COD_18_16N_04_MODELOENTRENADO_CART_,
                                '{}_'.format((self.MAINcuadrante[:2]).upper()),
                            )
                        )
                print(f'{"":=^80}')
        else:
            MAIN_COD_18_16N_04_MODELOENTRENADO_MINI = ''
            MAIN_COD_18_16N_04_MODELOENTRENADO_CART_ = ''
            MAIN_COD_18_16N_04_MODELOENTRENADO_CARTA = ''
            MAIN_COD_18_16N_04_MODELOENTRENADO_CARTB = ''
            MAIN_COD_18_16N_04_MODELOENTRENADO_128P = ''
        
            MAIN_LISTA_PNGS_MODELOENTRENADO_MINI = ''
            MAIN_LISTA_PNGS_MODELOENTRENADO_CART_ = ''
            MAIN_LISTA_PNGS_MODELOENTRENADO_CARTA = ''
            MAIN_LISTA_PNGS_MODELOENTRENADO_CARTB = ''
            MAIN_LISTA_PNGS_MODELOENTRENADO_128P = ''
        # ======================================================================
        self.paramConfigAdicionalesGLBL['GLBLmodeloCartolidMiniSubCelEntrenado'] = [self.GLBLmodeloCartolidMiniSubCelEntrenado, 'str', '', 'GrupoPredConvolucional']
        self.paramConfigAdicionalesGLBL['GLBLmodeloCartolidCartoSinguEntrenadoA'] = [self.GLBLmodeloCartolidCartoSinguEntrenadoA, 'str', '', 'GrupoPredConvolucional']
        self.paramConfigAdicionalesGLBL['GLBLmodeloCartolidCartoSinguEntrenadoB'] = [self.GLBLmodeloCartolidCartoSinguEntrenadoB, 'str', '', 'GrupoPredConvolucional']
        self.paramConfigAdicionalesGLBL['GLBLnombreFicheroConModeloParaInferencia'] = [self.GLBLnombreFicheroConModeloParaInferencia, 'str', '', 'GrupoPredConvolucional']
    
        self.paramConfigAdicionalesGLBL['MAIN_COD_18_16N_04_MODELOENTRENADO_MINI'] = [MAIN_COD_18_16N_04_MODELOENTRENADO_MINI, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_COD_18_16N_04_MODELOENTRENADO_CART_'] = [MAIN_COD_18_16N_04_MODELOENTRENADO_CART_, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_COD_18_16N_04_MODELOENTRENADO_CARTA'] = [MAIN_COD_18_16N_04_MODELOENTRENADO_CARTA, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_COD_18_16N_04_MODELOENTRENADO_CARTB'] = [MAIN_COD_18_16N_04_MODELOENTRENADO_CARTB, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_COD_18_16N_04_MODELOENTRENADO_128P'] = [MAIN_COD_18_16N_04_MODELOENTRENADO_128P, 'str', '', 'GrupoModeloEntrenado']
    
        self.paramConfigAdicionalesGLBL['MAIN_LISTA_PNGS_MODELOENTRENADO_MINI'] = [MAIN_LISTA_PNGS_MODELOENTRENADO_MINI, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_LISTA_PNGS_MODELOENTRENADO_CART_'] = [MAIN_LISTA_PNGS_MODELOENTRENADO_CART_, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_LISTA_PNGS_MODELOENTRENADO_CARTA'] = [MAIN_LISTA_PNGS_MODELOENTRENADO_CARTA, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_LISTA_PNGS_MODELOENTRENADO_CARTB'] = [MAIN_LISTA_PNGS_MODELOENTRENADO_CARTB, 'str', '', 'GrupoModeloEntrenado']
        self.paramConfigAdicionalesGLBL['MAIN_LISTA_PNGS_MODELOENTRENADO_128P'] = [MAIN_LISTA_PNGS_MODELOENTRENADO_128P, 'str', '', 'GrupoModeloEntrenado']
    

    # ==========================================================================
    def completarVariablesGlobalesResto(self):

        # ==========================================================================
        # ======================== Procesados especiales ===========================
        # ================ RENOMBRAR_FICHEROS, MERGEAR, GEOINTEGRAR ================
        # ==================== COMPRIMIR_LAS, DESCOMPRIMIR_LAZ ===================== 
        # ==== CREAR_SHAPE, CREAR_CAPA_CON_UNA_PROPIEDAD_DE_LOS_FICHEROS_LIDAR =====
        # ==========================================================================
    
        # ==========================================================================
        if __verbose__:
            print(f'{"":_^80}')
            print(f'{" clidaux-> Configuracion final ":_^80}')
            print(f'{"":_^80}')
            print(f'{TB}{"MAINobjetivoEjecucion":.<21}: {self.MAINobjetivoEjecucion}')
            print(f'{TB}{"MAINobjetivoSiReglado":.<21}: {self.MAINobjetivoSiReglado}')
            print(f'{TB}{"MAINprocedimiento":.<21}: {self.MAINprocedimiento}')
            print(f'{TB}{"MAINrutaRaizHome":.<21}: {self.MAINrutaRaizHome}')
            print(f'{TB}{"MAINmiRutaProyecto":.<21}: {self.MAINmiRutaProyecto}')
            print(f'{TB}{"MAINrutaCarto":.<21}: {self.MAINrutaCarto}')
            print(f'{TB}{"MAINrutaLaz":.<21}: {self.MAINrutaLaz}')
            print(f'{TB}{"MAINrutaOutput":.<21}: {self.MAINrutaOutput}')
            print(f'{TB}{"MAINcuadrante":.<21}: {self.MAINcuadrante}')
            if 'MAIN_RAIZ_DIR' in dir(GLO):
                print(f'{TB}{"MAIN_RAIZ_DIR":.<21}: {self.MAIN_RAIZ_DIR}')
            else:
                print(f'{TB}{"MAIN_RAIZ_DIR":.<21}: {"no en clidbase.xls"}')
            if 'MAINrutaRaizData' in dir(GLO):
                print(f'{TB}{"MAINrutaRaizData":.<21}: {self.MAINrutaRaizData}')
            else:
                print(f'{TB}{"MAINrutaRaizData":.<21}: {"no en clidbase.xls"}')
            if 'MAIN_MDLS_DIR' in dir(GLO):
                print(f'{TB}{"MAIN_MDLS_DIR":.<21}: {self.MAIN_MDLS_DIR}')
            else:
                print(f'{TB}{"MAIN_MDLS_DIR":.<21}: {"no en clidbase.xls"}')
            if 'GLBL_TRAIN_DIR' in dir(GLO):
                print(f'{TB}{"GLBL_TRAIN_DIR":.<21}: {self.GLBL_TRAIN_DIR}')
            else:
                print(f'{TB}{"GLBL_TRAIN_DIR":.<21}: {"no en clidbase.xls"}')
            print(f'{"":=^80}')
        # ==========================================================================
    
    
        # ==========================================================================
        self.paramConfigAdicionalesGLBL['MAINusuario'] = [self.MAINusuario, 'str', '', 'GrupoMAIN', self.MAINusuario]
        self.paramConfigAdicionalesGLBL['MAIN_ENTORNO'] = [MAIN_ENTORNO, 'str', '', 'GrupoMAIN', MAIN_ENTORNO]
        self.paramConfigAdicionalesGLBL['MAIN_PC'] = [MAIN_PC, 'str', '', 'GrupoMAIN', MAIN_PC]
    
        self.paramConfigAdicionalesGLBL['MAIN_DRIVE'] = [MAIN_DRIVE, 'str', '', 'GrupoDirsFiles', MAIN_DRIVE]
        self.paramConfigAdicionalesGLBL['MAIN_HOME_DIR'] = [MAIN_HOME_DIR, 'str', '', 'GrupoDirsFiles', MAIN_HOME_DIR]
        self.paramConfigAdicionalesGLBL['MAIN_FILE_DIR'] = [MAIN_FILE_DIR, 'str', '', 'GrupoDirsFiles', MAIN_FILE_DIR]
        self.paramConfigAdicionalesGLBL['MAIN_PROJ_DIR'] = [MAIN_PROJ_DIR, 'str', '', 'GrupoDirsFiles', MAIN_PROJ_DIR]
        self.paramConfigAdicionalesGLBL['MAIN_RAIZ_DIR'] = [MAIN_RAIZ_DIR, 'str', '', 'GrupoDirsFiles', MAIN_RAIZ_DIR]
        self.paramConfigAdicionalesGLBL['MAIN_THIS_DIR'] = [MAIN_THIS_DIR, 'str', '', 'GrupoDirsFiles', MAIN_THIS_DIR]
        self.paramConfigAdicionalesGLBL['MAIN_WORK_DIR'] = [MAIN_WORK_DIR, 'str', '', 'GrupoDirsFiles', MAIN_WORK_DIR]
        self.paramConfigAdicionalesGLBL['MAIN_RAIZ_DIR'] = [self.MAIN_RAIZ_DIR, 'str', '', 'GrupoDirsFiles', MAIN_RAIZ_DIR]
        self.paramConfigAdicionalesGLBL['MAINrutaRaizData'] = [self.MAINrutaRaizData, 'str', '', 'GrupoDirsFiles', MAIN_RAIZ_DIR]
#ççç    
        self.paramConfigAdicionalesGLBL['MAINmiRutaProyecto'] = [MAIN_PROJ_DIR, 'str', '', 'GrupoDirsFiles', MAIN_PROJ_DIR]
        self.paramConfigAdicionalesGLBL['MAINrutaLaz'] = [self.MAINrutaLaz, 'str', '', 'GrupoDirsFiles']
        # self.paramConfigAdicionalesGLBL['MAINrutaLazFinal'] = [
        #     self.MAINrutaLaz,
        #     'str',
        #     'Ruta en la que estan los ficheros laz o las (la usada finalmente, que puede ser distinta del establecido en cartolid.xml)',
        #     'GrupoDirsFiles',
        #     ]
        self.paramConfigAdicionalesGLBL['MAINrutaOutput'] = [self.MAINrutaOutput, 'str', '', 'GrupoDirsFiles']
        self.paramConfigAdicionalesGLBL['MAINrutaCarto'] = [self.MAINrutaCarto, 'str', '', 'GrupoDirsFiles']
    
        self.paramConfigAdicionalesGLBL['MAINprocedimiento'] = [self.MAINprocedimiento, 'str', '', 'GrupoMAIN']
#ççç    
        self.paramConfigAdicionalesGLBL['GLBLsoloCuadradoDeEjemplo'] = [self.GLBLsoloCuadradoDeEjemplo, 'bool', '', 'GrupoGestionDeFicheros']
        self.paramConfigAdicionalesGLBL['GLBLreprocesarFallidosUsandoMenosRAM'] = [self.GLBLreprocesarFallidosUsandoMenosRAM, 'bool', '', 'GrupoGestionDeFicheros']
        self.paramConfigAdicionalesGLBL['GLBLficheroLasTemporal'] = [self.GLBLficheroLasTemporal, 'bool', '', 'GrupoGestionDeFicheros']
        self.paramConfigAdicionalesGLBL['GLBLprocesarComprimidosLaz'] = [self.GLBLprocesarComprimidosLaz, 'bool', '', 'GrupoGestionDeFicheros']
    
        self.paramConfigAdicionalesGLBL['GLBLshapeNumPoints'] = [self.GLBLshapeNumPoints, 'int', '', 'GrupoShape']
        self.paramConfigAdicionalesGLBL['GLBLshapeFilter'] = [self.GLBLshapeFilter, 'str', '', 'GrupoShape']
    
        self.paramConfigAdicionalesGLBL['GLBLtipoLectura'] = [self.GLBLtipoLectura, 'str', '', 'GrupoManejoMemoria']
        self.paramConfigAdicionalesGLBL['GLBLusoDeRAM'] = [self.GLBLusoDeRAM, 'str', '', 'GrupoManejoMemoria']
    
        self.paramConfigAdicionalesGLBL['GLBLnumeroDePuntosAleer'] = [self.GLBLnumeroDePuntosAleer, 'int', '', 'GrupoLecturaPuntosPasadas']
        self.paramConfigAdicionalesGLBL['GLBLgrabarPercentilesRelativos'] = [self.GLBLgrabarPercentilesRelativos, 'bool', '', 'GrupoDasoLidar']
        self.paramConfigAdicionalesGLBL['GLBLmetrosCelda'] = [self.GLBLmetrosCelda, 'float', '', 'GrupoDimensionCeldasBloques']
    
        self.paramConfigAdicionalesGLBL['MAINprocedimientoFinal'] = [
            self.MAINprocedimiento,
            'str',
            'Procedimiento finalmente ejecutado (puede ser distinto del establecido en cartolid.xml)',
            'GrupoMAIN',
        ]
    
        if LCLverbose:
            print('clidaux-> paramConfigAdicionalesGLBL:')
            for nuevoParametro in self.paramConfigAdicionalesGLBL.keys():
                print('\t{:>40}: {}'.format(nuevoParametro, self.paramConfigAdicionalesGLBL[nuevoParametro]))
    
    
    # ==============================================================================
    def asignarMAINrutaCarto(
            self
        ):
        if MAIN_ENTORNO == 'calendula':
    #         MAINrutaRaizCarto =  LCLmiRutaRais
            MAINrutaRaizCarto =  self.MAINrutaRaizHome
            print(f'{TB}{TV}{TV}-> GLO.MAINrutaCarto en calendula es GLO.MAINrutaRaizHome: {self.MAINrutaRaizHome}')
        elif 'MAINrutaRaizHome' in dir(self):
            MAINrutaRaizCarto =  self.MAINrutaRaizHome
        else:
            if 'cartolidar' in MAIN_RAIZ_DIR:
                MAINrutaRaizCarto = os.path.abspath(os.path.join(MAIN_RAIZ_DIR, '..'))
            else:
                MAINrutaRaizCarto = os.path.abspath(MAIN_RAIZ_DIR)
        print(f'{TB}-> MAINrutaRaizCarto: {MAINrutaRaizCarto}')
    
        # Primera opcion:
        MAINrutaCarto1 = os.path.abspath(
            os.path.join(
                MAINrutaRaizCarto,
                'data/carto/'
            )
        )
        if os.path.isdir(MAINrutaCarto1):
            LCLrutaCarto = MAINrutaCarto1
        else:
            # Segunda opcion:
            MAINrutaCarto2 = os.path.abspath(
                os.path.join(
                    MAINrutaRaizCarto,
                    '../data/carto/'
                )
            )
            if os.path.isdir(MAINrutaCarto2):
                LCLrutaCarto = MAINrutaCarto2
            else:
                # myLog.warning(f'{"":+^80}')
                # myLog.warning(f'clidaux-> ATENCION: No se ha localizado el directorio data/carto con informacion cartografica de apoyo.')
                # myLog.warning(f'{TB}Directorios buscados:')
                # myLog.warning(f'{TB}{TV}{MAINrutaCarto1}')
                # myLog.warning(f'{TB}{TV}{MAINrutaCarto2}')
                # myLog.warning(f'{"":+^80}')
                sys.stderr.write(f'{"":+^80}\n')
                sys.stderr.write(f'clidaux-> ATENCION: No se ha localizado el directorio data/carto con informacion cartografica de apoyo.\n')
                sys.stderr.write(f'{TB}Directorios buscados:\n')
                sys.stderr.write(f'{TB}{TV}{MAINrutaCarto1}\n')
                sys.stderr.write(f'{TB}{TV}{MAINrutaCarto2}\n')
                sys.stderr.write(f'{"":+^80}\n')
                LCLrutaCarto = ''

        return LCLrutaCarto


    # ==============================================================================
    def asignarMAINrutaOutput(
            self,
            LCLprocedimiento,
            LCLrutaOutput,
            LCLmiRutaRais,
            LCLobjetivoSiReglado,
            LCLcuadrante,
        ):
    
        if MAIN_ENTORNO == 'calendula':
            MAINrutaRaizOutput =  LCLmiRutaRais
        elif 'MAINrutaRaizHome' in dir(self):
            MAINrutaRaizOutput =  self.MAINrutaRaizHome
        else:
            if 'cartolidar' in MAIN_RAIZ_DIR:
                MAINrutaRaizOutput = os.path.abspath(os.path.join(MAIN_RAIZ_DIR, '..'))
            else:
                MAINrutaRaizOutput = os.path.abspath(MAIN_RAIZ_DIR)

        print(f'{TB}{TV}-> MAINrutaRaizOutput (b): {MAINrutaRaizOutput}')
        # ==========================================================================
        # Ruta por defecto
        self.MAINrutaOutput = os.path.abspath(os.path.join(
            MAINrutaRaizOutput,
            'cartolidout'
        ))
        print(f'{TB}{TV}-> MAINrutaOutput por defecto: {self.MAINrutaOutput}')
        # ==========================================================================
        # Casos especiales
        if (
            LCLobjetivoSiReglado == 'GENERAL'
            or LCLobjetivoSiReglado == 'CREAR_TILES_TRAIN'
            or LCLobjetivoSiReglado == 'PREPROCESADO_EN_CALENDULA'
            or LCLobjetivoSiReglado == 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ'
            or LCLobjetivoSiReglado == 'CREAR_PUNTOS_TRAIN_ROQUEDOS'
            or LCLobjetivoSiReglado == 'CREAR_LAZ'
            or LCLobjetivoSiReglado == 'MALLA_DESFASADA'
            or LCLobjetivoSiReglado == 'EXTRA3'
        ):
            print(f'{TB}{TV}-> Se obtiene valor customizado de MAINrutaOutput para casos especiales:')
            print(f'{TB}{TV}{TW}-> LCLobjetivoSiReglado: {LCLobjetivoSiReglado}')
            print(f'{TB}{TV}{TW}-> LCLprocedimiento:     {LCLprocedimiento}')
            if (
                LCLprocedimiento.startswith('AUTOMATICO_EN_CALENDULA_SCRATCH')
                or 'COLOREAR_RGBI' in LCLprocedimiento
                or 'LAS_INFO' in LCLprocedimiento
                or 'LAS_EDIT' in LCLprocedimiento
            ):
                if 'COLOREAR_RGBI' in LCLprocedimiento:
                    LCLrutaOutput = os.path.abspath(os.path.join(
                        MAINrutaRaizOutput,
                        # '..',
                        'cartolidout_{}_{}_{}'.format(
                            (LCLcuadrante)[:2].upper(),
                            'COLOREAR_RGBI',
                            'completo'
                        )
                    ))
                elif 'LAS_INFO' in LCLprocedimiento:
                    LCLrutaOutput = os.path.abspath(os.path.join(
                        MAINrutaRaizOutput,
                        # '..',
                        'cartolidout_{}_{}_{}'.format(
                            (LCLcuadrante)[:2].upper(),
                            'LAS_INFO',
                            'completo'
                        )
                    ))
                elif 'LAS_EDIT' in LCLprocedimiento:
                    LCLrutaOutput = os.path.abspath(os.path.join(
                        MAINrutaRaizOutput,
                        # '..',
                        'cartolidout_{}_{}_{}'.format(
                            (LCLcuadrante)[:2].upper(),
                            'LAS_EDIT',
                            GLO.GLBLeditarLasFile,
                            # 'completo'
                        )
                    ))
                else:
                    LCLrutaOutput = os.path.abspath(os.path.join(
                        MAINrutaRaizOutput,
                        # '..',
                        'cartolidout_{}_{}_{}'.format(
                            (LCLcuadrante)[:2].upper(),
                            LCLobjetivoSiReglado,
                            'completo'
                        )
                    ))
            elif (
                LCLprocedimiento == 'AUTOMATICO_EN_CALENDULA_SELECT'
                or LCLprocedimiento == 'AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA'
                or LCLprocedimiento == 'AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA_SELECT'
            ):
                LCLrutaOutput = os.path.abspath(os.path.join(
                    MAINrutaRaizOutput,
                    # '..',
                    'cartolidout_{}_{}'.format(
                        (LCLcuadrante)[:2].upper(),
                        LCLobjetivoSiReglado,
                    )
                ))
            elif not self.MAINrutaOutput is None:
                LCLrutaOutput = self.MAINrutaOutput
                print(f'{TB}{TV}-> 6 Asignando LCLrutaOutput: {LCLrutaOutput}')
            elif not LCLcuadrante is None:
                LCLrutaOutput = os.path.abspath(os.path.join(
                    MAINrutaRaizOutput,
                    # '..',
                    'cartolidout_{}_{}'.format(
                        (LCLcuadrante)[:2].upper(),
                        LCLobjetivoSiReglado
                    )
                ))
                print(f'{TB}{TV}-> 7 Asignando LCLrutaOutput: {LCLrutaOutput}')
            elif LCLobjetivoSiReglado == 'GENERAL':
                LCLrutaOutput = os.path.abspath(os.path.join(
                    MAINrutaRaizOutput,
                    # '..',
                    'cartolidout'
                ))
                print(f'{TB}{TV}-> 8 Asignando LCLrutaOutput: {LCLrutaOutput}')
            else:
                LCLrutaOutput = self.MAINrutaOutput
                print(f'{TB}{TV}-> 9 Asignando LCLrutaOutput: {LCLrutaOutput}')
            if not LCLcuadrante is None:
                LCLrutaOutput = (LCLrutaOutput).replace('_XX', '_{}'.format((LCLcuadrante)[:2].upper()))
            # print('\t{:.<25}: {}'.format('MAINrutaOutput (adaptado)', LCLrutaOutput))
        elif (
            LCLrutaOutput is None
            or LCLrutaOutput == 'None'
            or LCLrutaOutput == ''
        ):
            LCLrutaOutput = os.path.abspath(os.path.join(
                MAINrutaRaizOutput,
                # '..',
                'cartolidout_{}'.format(
                    (LCLcuadrante)[:2].upper()
                )
            ))
            LCLrutaOutput = (LCLrutaOutput).replace('_XX', '_{}'.format((LCLcuadrante)[:2].upper()))
            print('\t{:.<25}: {}'.format('MAINrutaOutput (adaptado)', LCLrutaOutput))
        else:
            print('\t{:.<25}: {}'.format('MAINrutaOutput (original)', LCLrutaOutput))

        if LCLprocedimiento.startswith('AUTOMATICO_DISCOEXTERNO'):
            # Incluye 'AUTOMATICO_DISCOEXTERNO_UNIFICAR_RGBI_E'
            unidadLaz = LCLprocedimiento[-1:]
            # LCLrutaOutput = unidadLaz + ':/CE/cartolidout/'
            LCLrutaOutput = unidadLaz + ':/lidardata_2017_2021/Result/SE/cartolidout/'
        elif LCLprocedimiento[:20] == 'AUTOMATICO_CUADRANTE':
            # Guardo los resultados en una ruta que cuelga de la ruta de los laz (no de la ruta de cartolid)
            bloqueElegido = 0
            miCuadrante = LCLprocedimiento[21:23]
            miMarco = LCLprocedimiento[-8:]
            if LCLrutaOutput == '':
                LCLrutaOutput = os.path.join(self.MAIN_RAIZ_DIR, 'cartolidout_%s_%s' % (miCuadrante, miMarco))
            else:
                # LCLrutaOutput = quitarContrabarrasAgregarBarraFinal(LCLrutaOutput)
                LCLrutaOutput = (LCLrutaOutput).replace(os.sep, '/')
    
        if LCLrutaOutput == '' or LCLrutaOutput == None:
            miCarpetaResultadosInicial = 'cartolidout'
            print(f'\n{"":_^80}')
            print('Escribir ruta en la que se guardan los resultados:')
            print('Se interpreta como ruta completa si empieza por "/" o por letra de unidad ("C:", "D:", etc.).')
            print('y como subdirectorio de "%s" en el resto de los casos' % self.MAIN_RAIZ_DIR)
            selec = input('Introduce un valor (pulsa [enter] para valor por defecto: %s): ' %
                          os.path.join(self.MAIN_RAIZ_DIR, miCarpetaResultadosInicial))
            if selec[1:2] == ':' or selec[:1] == '/':
                # Es una ruta completa en vez de un subdirectorio
                LCLrutaOutput = selec
            elif selec != '':
                LCLrutaOutput = os.path.join(self.MAIN_RAIZ_DIR, selec)
            else:
                LCLrutaOutput = os.path.join(self.MAIN_RAIZ_DIR, miCarpetaResultadosInicial)
            print('{:^80}'.format(''))
    
        return LCLrutaOutput


# ==============================================================================
def leerCambiarVariablesGlobales(
        nuevosParametroConfiguracion={},
        LCL_idProceso=MAIN_idProceso,
        inspect_stack=inspect.stack(),
        verbose=False,
    ):
    # Lectura del config (cfg) generado especificamente para esta ejecucion.
    # print('LCL_idProceso:', type(LCL_idProceso), LCL_idProceso)
    # print('sys.argv:', sys.argv)
    configFileNameCfg = getConfigFileNameCfg(LCL_idProceso, LOCL_verbose=0)

    if CONFIGverbose or verbose:
        print(f'\n{"":_^80}')

    if CONFIGverbose and verbose:
        print(f'clidconfig-> Leo los paramConfig del cfg (lo actualizo si tengo nuevosParametroConfiguracion) con leerCambiarVariablesGlobales<>')
        print(f'{TB}Reviso la pila de llamadas para ver desde que modulo estoy cargando los paramConfig del cfg')
        callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
        print(f'{TB}-> callingModulePrevio: {callingModulePrevio}')

    config = RawConfigParser()
    config.optionxform = str  # Avoid change to lowercase

    if not os.path.exists(configFileNameCfg):
        print(f'\nclidconfig-> ATENCION: fichero de configuracion no encontrado: {configFileNameCfg}')
        print(f'{TB}-> Revisar la linea ~2523 de clidconfig.py, para que se cree el .cfg si callingModuleInicial == "clidbase" or ...')
        callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=True)
        print(f'{TB}-> callingModulePrevio: {callingModulePrevio} callingModuleInicial: {callingModuleInicial}')
        sys.exit(0)
        # return False

    if CONFIGverbose or verbose:
        print(f'clidconfig-> Leyendo cfg: {configFileNameCfg}')
    numObjetivosExtraMax = 0
    LOCALconfigDict = {}
    # El primer parametro de configuracion es el propio nombre del fichero 
    # de configuracion cfg (en realidad es redundante y no lo uso):
    LOCALconfigDict['configFileNameCfg'] = [
        configFileNameCfg,
        'GrupoMAIN',
        'Nombre del fichero de configuracion (este fichero).',
        'str',
    ]
    # if True:
    try:
        config.read(configFileNameCfg)
        mostrarFrupoMAIN = False
        if mostrarFrupoMAIN:
            print(f'clidconfig-> Parametros del GrupoMAIN del fichero de configuracion ({configFileNameCfg}):')
        for grupoParametroConfiguracion in config.sections():
            for nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion):
                strParametroConfiguracion = config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion)
                listaParametroConfiguracion = strParametroConfiguracion.split('|+|')
                valorParametroConfiguracion = valorConfig(
                    listaParametroConfiguracion[0],
                    nombreParametro=nombreParametroDeConfiguracion,
                    tipoVariable=listaParametroConfiguracion[1]
                )
                if len(listaParametroConfiguracion) > 1:
                    tipoParametroConfiguracion = listaParametroConfiguracion[1]
                else:
                    tipoParametroConfiguracion = 'str'
                if len(listaParametroConfiguracion) > 2:
                    descripcionParametroConfiguracion = listaParametroConfiguracion[2]
                else:
                    descripcionParametroConfiguracion = ''
                if nombreParametroDeConfiguracion[:1] == '_':
                    grupoParametroConfiguracion_new = '_%s' % grupoParametroConfiguracion
                else:
                    grupoParametroConfiguracion_new = grupoParametroConfiguracion
                LOCALconfigDict[nombreParametroDeConfiguracion] = [
                    valorParametroConfiguracion,
                    grupoParametroConfiguracion_new,
                    descripcionParametroConfiguracion,
                    tipoParametroConfiguracion,
                ]
                numObjetivosExtra = max(len(listaParametroConfiguracion) - 3, 0)
                numObjetivosExtraMax = max(numObjetivosExtra, numObjetivosExtraMax)
                if numObjetivosExtraMax:
                    valObjetivosExtra = []
                    for nObjetivoExtra in range(numObjetivosExtraMax):
                        if 3 + nObjetivoExtra < len(listaParametroConfiguracion):
                            valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[3 + nObjetivoExtra], tipoVariable=listaParametroConfiguracion[1]))
                        else:
                            valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[0], tipoVariable=listaParametroConfiguracion[1]))
                    LOCALconfigDict[nombreParametroDeConfiguracion].extend(valObjetivosExtra)

                if mostrarFrupoMAIN and grupoParametroConfiguracion == 'GrupoMAIN':
                    print(f'{TB}-> >>>5 numObjetivosExtra: {numObjetivosExtra}, Max: {numObjetivosExtraMax}, >>> {nombreParametroDeConfiguracion}, {LOCALconfigDict[nombreParametroDeConfiguracion]}')

        if CONFIGverbose or verbose:
            print(f'{TB}Parametros leidos ok del fichero cfg: {LOCALconfigDict["configFileNameCfg"][0]}')

    except Exception as excpt:
        program_name = 'clidconfig.py'
        print(f'\n{program_name}-> Error Exception en clidconfig-> {excpt}\n')
        # https://stackoverflow.com/questions/1278705/when-i-catch-an-exception-how-do-i-get-the-type-file-and-line-number
        exc_type, exc_obj, exc_tb = sys.exc_info()
        if verbose > 1:
            print()
            # print(f'exc_obj ({type(exc_obj)}): <<{str(exc_obj)}>>')
            # print(dir(exc_obj)) # 'args', 'characters_written', 'errno', 'filename', 'filename2', 'strerror', 'winerror', 'with_traceback'
            try:
                numeroError = exc_obj.errno
            except:
                numeroError = -1
            print(f'numError:  {numeroError}')      # 13
            print(f'filename:  {exc_obj.filename}')
            print(f'filename2: {exc_obj.filename2}')
            try:
                print(f'strerror:  {exc_obj.strerror}')
            except:
                print(f'strerror_: {exc_obj}')
            # print(f'with_traceback: {exc_obj.with_traceback}')  # <built-in method
            print()
            print(f'filename {exc_tb.tb_frame.f_code.co_filename}')
            print(f'lineno   {exc_tb.tb_lineno}')
            print(f'function {exc_tb.tb_frame.f_code.co_name}')
            print(f'type     {exc_type.__name__}')
            # print(f'message  {exc_obj.message}')  # or see traceback._some_str()
            print()
    
        fileNameError = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        funcError = os.path.split(exc_tb.tb_frame.f_code.co_name)[1]
        lineError = exc_tb.tb_lineno
        typeError = exc_type.__name__
        try:
            descError = exc_obj.strerror
        except:
            descError = exc_obj
        sys.stderr.write(f'\nPuede contribuir a este programa remitiendo este error a cartolidar@gmail.com:\n')
        sys.stderr.write(f'{TB}Error en     {fileNameError}\n')
        sys.stderr.write(f'{TB}Funcion:     {funcError}\n')
        sys.stderr.write(f'{TB}Linea:       {lineError}\n')
        sys.stderr.write(f'{TB}Descripcion: {descError}\n') # = {exc_obj}
        sys.stderr.write(f'{TB}Tipo:        {typeError}\n')
    
        sys.stderr.write(f'clidconfig-> Error al leer la configuracion del fichero: {configFileNameCfg}\n')
        sys.exit(0)


    # Estos parametros llegan como dict de listas de valores (no como listas de textos, que es lo que ocurre con la listaParametroConfiguracion leida del cfg)        
    if nuevosParametroConfiguracion != {}:
        if CONFIGverbose or verbose:
            print(f'clidconfig-> nuevosParametroConfiguracion: {nuevosParametroConfiguracion}')
        for nombreParametroDeConfiguracion in nuevosParametroConfiguracion.keys():
            listaNuevosParametroConfiguracion = nuevosParametroConfiguracion[nombreParametroDeConfiguracion]
            valorParametroConfiguracion = listaNuevosParametroConfiguracion[0]
            if valorParametroConfiguracion is None or valorParametroConfiguracion == 'None':
                continue
            try:
                tipoParametroConfiguracion = listaNuevosParametroConfiguracion[1]
            except:
                tipoParametroConfiguracion = 'str'
            try:
                descripcionParametroConfiguracion = listaNuevosParametroConfiguracion[2]
            except:
                descripcionParametroConfiguracion = ''
            try:
                grupoParametroConfiguracion = listaNuevosParametroConfiguracion[3]
            except:
                grupoParametroConfiguracion = 'GrupoDESC'
            if not grupoParametroConfiguracion in config.sections():
                # print('clidconfig-> ->grupoParametros nuevo', grupoParametroConfiguracion)
                config.add_section(grupoParametroConfiguracion)

            # Si el parametro ya tiene descripcion en clidbase.xml, se mantiene esa descripcion
            # print('clidconfig-> Revisando valores previos-> grupo: {} parametro: {}'.format(grupoParametroConfiguracion, nombreParametroDeConfiguracion))
            # print('\t', nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion))
            if nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion):
                listaParametroConfiguracionOriginal = config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion).split('|+|')
                descParametroConfiguracion = listaParametroConfiguracionOriginal[2]
                if descParametroConfiguracion != '':
                    descripcionParametroConfiguracion = descParametroConfiguracion
                # print('clidconfig-> valor previo de: {} -> {}'.format(nombreParametroDeConfiguracion, listaParametroConfiguracionOriginal))

            listaConcatenada = '{}|+|{}|+|{}'.format(
                str(valorParametroConfiguracion),
                str(tipoParametroConfiguracion),
                str(descripcionParametroConfiguracion)
                )

            if numObjetivosExtraMax:
                valObjetivosExtra = []
                for nObjetivoExtra in range(numObjetivosExtraMax):
                    if 4 + nObjetivoExtra < len(listaNuevosParametroConfiguracion):
                        valObjetivosExtra.append(listaNuevosParametroConfiguracion[4 + nObjetivoExtra])
                        listaConcatenada += '|+|{}'.format(listaNuevosParametroConfiguracion[4 + nObjetivoExtra])
                    else:
                        valObjetivosExtra.append(listaNuevosParametroConfiguracion[0])
                        listaConcatenada += '|+|{}'.format(listaNuevosParametroConfiguracion[0])

            config.remove_option(
                grupoParametroConfiguracion,
                nombreParametroDeConfiguracion,
            )
            config.set(
                grupoParametroConfiguracion,
                nombreParametroDeConfiguracion,
                listaConcatenada
            )

            LOCALconfigDict[nombreParametroDeConfiguracion] = [
                valorParametroConfiguracion,
                grupoParametroConfiguracion,
                descripcionParametroConfiguracion,
                tipoParametroConfiguracion,
            ]
            if numObjetivosExtra:
                LOCALconfigDict[nombreParametroDeConfiguracion].extend(valObjetivosExtra)

            if verbose:
                # print(f'{TB}Nuevo Valor de: {nombreParametroDeConfiguracion} -> {listaConcatenada}')
                print(f'{TB}-> Nuevo Valor ok: {nombreParametroDeConfiguracion} -> {config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion).split("|+|")}')

            if (CONFIGverbose or verbose) and grupoParametroConfiguracion == 'GrupoMAIN':
                # print(f'{TB}{TV}clidconfig-> >>>6 {configFileNameCfg}, {nombreParametroDeConfiguracion}, {listaNuevosParametroConfiguracion}')
                print(
                    f'{TB}-> >>>6 Nuevo parametro del grupo',
                    grupoParametroConfiguracion,
                    '->',
                    nombreParametroDeConfiguracion,
                    'con valor:',
                    valorParametroConfiguracion,
                    'y valores extras:',
                    valObjetivosExtra,
                    'listaNuevosParametroConfiguracion',
                    listaNuevosParametroConfiguracion
                )

        if CONFIGverbose or verbose:
            print(f'clidconfig-> Guardando los nuevos parametros en: {configFileNameCfg}')

        nIntentos = 0
        while True:
            nIntentos += 1
            try:
                if os.path.exists(configFileNameCfg):
                    os.remove(configFileNameCfg)
                break
            except:
                if nIntentos > 5:
                    print(f'clidconfig-> ATENCION: no se ha podio eliminar {configFileNameCfg} para crear uno nuevo con los nuevos parametros de configuracion')
                    sys.exit(0)
                time.sleep(60)
        with open(configFileNameCfg, mode='w+') as configfile:
            config.write(configfile)

        if CONFIGverbose or verbose:
            if os.path.exists(configFileNameCfg):
                print(f'{TB}-> Ok fichero cfg actualizado con los nuevos parametros.')
            else:
                print(f'{TB}-> AVISO: error al crear {configFileNameCfg} con los nuevos parametros.')

    if CONFIGverbose or verbose:
        print(f'{"":=^80}')

    return LOCALconfigDict


# ==============================================================================
def guardarVariablesGlobales(
        LOCALconfigDict,
        LCL_idProceso=MAIN_idProceso,
        inspect_stack=inspect.stack(),
    ):
    # runnigFileName = sys.argv[0].replace('.py', '.running')
    # if not os.path.exists(runnigFileName):
    #     open(runnigFileName, mode='w+')
    #     time.sleep(2)
    # else:
    #     time.sleep(5)

    configFileNameCfg = getConfigFileNameCfg(LCL_idProceso, LOCL_verbose=0)
    if os.path.exists(configFileNameCfg):
        try:
            os.remove(configFileNameCfg)
        except:
            print('No se elimina el fichero', configFileNameCfg, 'porque debe haberse eliminado por otro proceso')
            time.sleep(5)

    if CONFIGverbose or LCLverbose:
        print(f'\n{"":_^80}')
        print(f'clidconfig-> Guardo los paramConfig en fichero cfg (inicial) con guardarVariablesGlobales:')
        print(f'{TB}{configFileNameCfg}')
        print(f'{TB}-> Reviso la pila de llamadas por si llamara a esta funcion de nuevo:')
        callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect_stack, verbose=CONFIGverbose)
        print(f'{TB}{TV}-> callingModulePrevio: {callingModulePrevio}; callingModuleInicial: {callingModuleInicial}')

    config = RawConfigParser()
    config.optionxform = str  # Avoid change to lowercase

    # When adding sections or items, add them in the reverse order of
    # how you want them to be displayed in the actual file.
    # In addition, please note that using RawConfigParser's and the raw
    # mode of ConfigParser's respective set functions, you can assign
    # non-string values to keys internally, but will receive an error
    # when attempting to write to a file or when you get it in non-raw
    # mode. SafeConfigParser does not allow such assignments to take place.

    # Creo los grupos de variables
    for nombreParametroDeConfiguracion in LOCALconfigDict.keys():
        grupoParametroConfiguracion = LOCALconfigDict[nombreParametroDeConfiguracion][1]
        if not grupoParametroConfiguracion in config.sections():
            # print('clidconfig-> ->grupoParametros nuevo', grupoParametroConfiguracion)
            config.add_section(grupoParametroConfiguracion)
    config.add_section('ChangedConfig')
    config.add_section('UserConfig')
    config.add_section('ControlFiles')

    numObjetivosExtraMax = 0
    for nombreParametroDeConfiguracion in LOCALconfigDict.keys():
        # LOCALconfigDict[nombreParametroDeConfiguracion] = [valorParametroDeConfiguracion, grupoParametros, tipoVariable, descripcionParametro]
        listaParametroConfiguracion = LOCALconfigDict[nombreParametroDeConfiguracion]
        valorParametroConfiguracion = listaParametroConfiguracion[0]
        grupoParametroConfiguracion = listaParametroConfiguracion[1]
        descripcionParametroConfiguracion = listaParametroConfiguracion[2]
        tipoParametroConfiguracion = listaParametroConfiguracion[3]

        # config.set(grupoParametroConfiguracion, nombreParametroDeConfiguracion, [str(valorParametroConfiguracion), tipoParametroConfiguracion])
        if not descripcionParametroConfiguracion is None:
            if (
                'á' in descripcionParametroConfiguracion
                or 'é' in descripcionParametroConfiguracion
                or 'í' in descripcionParametroConfiguracion
                or 'ó' in descripcionParametroConfiguracion
                or 'ú' in descripcionParametroConfiguracion
                or 'ñ' in descripcionParametroConfiguracion
                or 'ç' in descripcionParametroConfiguracion
            ):
                descripcionParametroConfiguracion = ''.join(normalize(c) for c in str(descripcionParametroConfiguracion))
            if (descripcionParametroConfiguracion.encode('utf-8')).decode('cp1252') != descripcionParametroConfiguracion:
                descripcionParametroConfiguracion = ''

        listaConcatenada = '{}|+|{}|+|{}'.format(
            str(valorParametroConfiguracion),
            str(tipoParametroConfiguracion),
            str(descripcionParametroConfiguracion)
            )

        numObjetivosExtra = max(len(listaParametroConfiguracion) - 4, 0)
        numObjetivosExtraMax = max(numObjetivosExtra, numObjetivosExtraMax)
        if numObjetivosExtraMax:
            valObjetivosExtra = []
            for nObjetivoExtra in range(numObjetivosExtraMax):
                if len(listaParametroConfiguracion) > 4 + nObjetivoExtra:
                    valObjetivosExtra.append(listaParametroConfiguracion[4 + nObjetivoExtra])
                    listaConcatenada += '|+|{}'.format(listaParametroConfiguracion[4 + nObjetivoExtra])
                else:
                    valObjetivosExtra.append(listaParametroConfiguracion[0])
                    listaConcatenada += '|+|{}'.format(listaParametroConfiguracion[0])
        config.set(
            grupoParametroConfiguracion,
            nombreParametroDeConfiguracion,
            listaConcatenada
        )
        # print(str(valorParametroConfiguracion), tipoParametroConfiguracion, descripcionParametroConfiguracion)

    # print('clidconfig-> configFileNameCfg:', configFileNameCfg)
    # with open(configFileNameCfg, 'wb') as configfile:
    try:
        with open(configFileNameCfg, mode='w+') as configfile:
            config.write(configfile)
    except:
        print(f'\nclidconfig-> ATENCION, revisar caracteres no admitidos en el fichero de configuracion: {configFileNameCfg}')
        print(f'{TB}Ejemplos: vocales acentuadas, ennes, cedillas, flecha dchea (->), etc.')

    if CONFIGverbose or LCLverbose:
        print(f'{TB}-> Ok creado fichero: {configFileNameCfg}')
        print(f'{"":=^80}')


# oooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# A este modulo se le llamaba solo desde clidbase.py
# Estaba inicialmente destinado a poder acceder a las variables globales
# como propiedades de este modulo en vez de como propiedades de GLO,
# que es la opcion finalmente elegida
# oooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# def registrarVariablesModuleControlFiles(
#         ctrlFileLasName='',
#         ctrlFileLasObj='',
#         ctrlFileGralObj='',
#         GLBLficheroDeCtrlGral='',
#     ):
#     # global controlsInitiated
#     # controlsInitiated = True
#     global controlFileLasName
#     controlFileLasName = ctrlFileLasName
#     global controlFileLasObj
#     controlFileLasObj = ctrlFileLasObj
#     global controlFileGralObj
#     controlFileGralObj = ctrlFileGralObj
#
#     global GLBLficheroDeControlGral
#     GLBLficheroDeControlGral = GLBLficheroDeCtrlGral



# ==============================================================================
def casosEspecialesParaMAINrutaLaz(
        LCLprocedimiento,
        LCLrutaLaz,
        LCLlistaDirsLaz,
        LCLlistaSubDirsLaz,
        LCLcuadrante,
        LCLverbose = False,
    ):

    # Ya no uso esto para elegir la lista de subdirectorios a explorar
    # porque para eso uso _MAINlistaSubDirsLaz y MAINincluirRaizEnListaSubdirsLaz

    # ==========================================================================
    # ============== Casos especiales para MAINrutaLaz =========================
    # ==========================================================================
    # if MAIN_ENTORNO == 'calendula' and False:
    #     # Solo los laz se ubican en scratch, el resto en home
    #     # GLO.MAINrutaCarto = '/scratch/jcyl_spi_1/jcyl_spi_1_1/data/carto'
    #     # GLO.MAINrutaCarto = '/LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1/data/carto'
    #     # MAIN_RAIZ_DIR = '/scratch/jcyl_spi_1/jcyl_spi_1_1/'
    #
    #     if (
    #         LCLprocedimiento.startswith('COMPRIMIR_LAS')
    #         or LCLprocedimiento.startswith('DESCOMPRIMIR_LAZ')
    #     ):
    #         if 'RGBI_cartolid' in LCLprocedimiento:
    #             LCLlistaSubDirsLaz = ['RGBI_cartolid']
    #         elif 'RGB_cartolid' in LCLprocedimiento:
    #             LCLlistaSubDirsLaz = ['RGB_cartolid']
    #         elif 'IRG_cartolid' in LCLprocedimiento:
    #             LCLlistaSubDirsLaz = ['IRC_cartolid']
    #         elif 'IRG_cartolid' in LCLprocedimiento:
    #             LCLlistaSubDirsLaz = ['IRG_cartolid']
    #         else:
    #             LCLlistaSubDirsLaz = ['RGBI']
    #     elif (
    #         LCLprocedimiento.endswith('UNIFICAR_RGBI')
    #         or LCLprocedimiento.endswith('COLOREAR_RGBI')
    #     ):
    #         LCLlistaSubDirsLaz = ['IRC']
    #         # LCLlistaSubDirsLaz = ['IRC/_corregidos_versionOk_preORT_preCOLOR'] # Lo traslado temporalmente a SW y lo proceso ahi
    #     else:
    #         if GLO.MAINobjetivoEjecucion == 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ':
    #             # LCLlistaSubDirsLaz = ['lazNew']
    #             LCLlistaSubDirsLaz = ['lazNewCLR']
    #         else:
    #             if LCLcuadrante[:2].upper() == 'SE':
    #                 if 'SELECT' in LCLprocedimiento:
    #                     LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI', 'RGBI_laz_H29', 'RGBI_laz']
    #                 else:
    #                     LCLlistaSubDirsLaz = ['RGBI_laz_cartolid_20220316']
    #             elif LCLcuadrante[:2].upper() == 'NE':
    #                     LCLlistaSubDirsLaz = ['', '000_laz1']  # , '000_laz2
    #             elif LCLcuadrante[:2].upper() == 'NW':
    #                 if 'SELECT' in LCLprocedimiento:
    #                     primeraVersionDeLasFiles = False
    #                     if primeraVersionDeLasFiles:
    #                         # La primera vez proceso una muestra (SELECT) de bloques sin clasicacion
    #                         # para tener una primera version de lasFiles clasificados provisionalmente en lasNew
    #                         # Los lasFiles clasificados los muevo a lazNewCLR para trabajar preferencialmente con ellos en lo sucesivo
    #                         # La segunda vez trabajo sobre esos lasFiles clasificados para mejorarlos
    #                         rutaLazCompleta = os.path.join(GLO.MAINrutaLaz, LCLlistaDirsLaz[0], 'lazNewCLR')
    #                         if len(LCLlistaDirsLaz) == 1 and os.path.isdir(rutaLazCompleta):
    #                             print('clidaux-> Aviso: Uso lazNewCLR para usar lasFiles provisionalmente clasificados y disponer de clase del miniSubCel')
    #                             LCLlistaSubDirsLaz = ['lazNewCLR']
    #                         else:
    #                             print('clidaux-> Aviso: No se ha encontrado la ruta de lasFiles pro-clasificados: {}'.format(rutaLazCompleta))
    #                             print('\t-> LCLlistaDirsLaz: {}'.format(LCLlistaDirsLaz))
    #                             # LCLlistaSubDirsLaz = ['RGBI_H29']
    #                             # LCLlistaSubDirsLaz = ['RGBI_laz_H29', 'RGBI_laz']
    #                             # LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI']
    #                             LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI', 'RGBI_laz_H29', 'RGBI_laz']
    #                     else:
    #                         LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI', 'RGBI_laz_H29', 'RGBI_laz']
    #                 else:
    #                     LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI']
    #                     LCLlistaSubDirsLaz = ['lazNewCompletoAll_RGBI']
    #                     LCLlistaSubDirsLaz = ['lazNewCompleto_RGBI_laz_20220302']
    #                     LCLlistaSubDirsLaz = ['lazNewCompleto_RGBI_laz_20220316']
    #
    #             elif LCLcuadrante[:2].upper() == 'CE':
    #                 if 'SELECT' in LCLprocedimiento:
    #                     LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI', 'RGBI_laz_H29', 'RGBI_laz']
    #                 else:
    #                     LCLlistaSubDirsLaz = ['RGBI_laz_cartolid_20220316']
    #             elif LCLcuadrante[:2].upper() == 'SE':
    #                 if 'SELECT' in LCLprocedimiento:
    #                     LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI', 'RGBI_laz_H29', 'RGBI_laz']
    #                 else:
    #                     LCLlistaSubDirsLaz = ['RGBI_laz_cartolid_20220316']
    #
    #             elif LCLcuadrante[:2].upper() == 'SW':
    #                 # LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI', 'RGBI_laz_H29', 'RGBI_laz']
    #                 # LCLlistaSubDirsLaz = ['RGBI_H29']
    #                 LCLlistaSubDirsLaz = ['RGBI_laz_H29', 'RGBI_laz']
    #                 # LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI']
    #             elif LCLcuadrante[:2].upper() == 'XX':
    #                 LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI_H30']
    #             else:
    #                 # LCLlistaSubDirsLaz = ['RGBI_laz']
    #                 LCLlistaSubDirsLaz = ['RGBI']
    #
    #     if '_H29_' in GLO.MAINprocedimiento:
    #         # Por el momento, esto solo lo uso para AUTOMATICO_EN_CALENDULA_SCRATCH_H29_COLOREAR_RGBI
    #         for nSubDir, subDirLaz in enumerate(LCLlistaSubDirsLaz):
    #             if '_H29' in subDirLaz.upper():
    #                 print(f'\nclidaux-> ATENCION: revisar este codigo para adaptarlo a la lista de directorios que quiero recorrer:')
    #                 print(f'\t-> GLO.MAINprocedimiento: {GLO.MAINprocedimiento}')
    #                 print(f'\t-> LCLlistaSubDirsLaz:       {LCLlistaSubDirsLaz}')
    #                 sys.exit(0)
    #             LCLlistaSubDirsLaz[nSubDir] = subDirLaz + '_H29'
    #
    # elif (
    #     LCLprocedimiento.startswith('DESCOMPRIMIR_LAZ')
    #     or LCLprocedimiento.startswith('COMPRIMIR_LAS')
    # ):
    #     # Rutas ad-hoc
    #     LCLlistaDirsLaz = ['lasfile-nw']
    #     LCLlistaSubDirsLaz = ['RGBI_laz_H29']
    # elif LCLprocedimiento.startswith('LAS_INFO_ASK_LASDIR'):
    #     LCLrutaLaz = preguntarRutaLaz(GLO.MAINrutaLaz)
    # elif LCLprocedimiento == 'LAS_INFO_BASE_LASDIR':
    #     LCLrutaLaz = GLO.MAINrutaLaz
    # elif (
    #     'COLOREAR_RGBI' in LCLprocedimiento
    #     or 'CORREGIR_ELIPSOIDALES' in LCLprocedimiento
    # ) and not LCLrutaLaz.endswith('laz')  and not LCLrutaLaz.endswith('laz2'):
    #     print(f'clidaux-> Se explora la ruraLaz sin subdirectorios: {LCLrutaLaz}')
    #     print(f'{TB}-> GLO.MAINrutaLaz: {GLO.MAINrutaLaz}')
    #     LCLlistaDirsLaz = ['']
    '''
    # Esto es lo estandar no es caso especial
    if (
        LCLprocedimiento.startswith('AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA')
        or LCLprocedimiento.startswith('AUTOMATICO_EN_CALENDULA')
        or LCLprocedimiento.startswith('LAS_INFO')
    ):
        # Se usan los valores establecidos por defecto
        if (LCLcuadrante)[:2].upper() == 'CE':
            LCLlistaDirsLaz = ['lasfile-ce']
        elif (LCLcuadrante)[:2].upper() == 'NE':
            LCLlistaDirsLaz = ['lasfile-ne']
        elif (LCLcuadrante)[:2].upper() == 'NW':
            LCLlistaDirsLaz = ['lasfile-nw']
        elif (LCLcuadrante)[:2].upper() == 'SE':
            LCLlistaDirsLaz = ['lasfile-se']
        elif (LCLcuadrante)[:2].upper() == 'SW':
            LCLlistaDirsLaz = ['lasfile-sw']
        elif (LCLcuadrante)[:2].upper() == 'XX':
            LCLlistaDirsLaz = ['lasfile-ce', 'lasfile-nw', 'lasfile-ne', 'lasfile-se', 'lasfile-sw']
            LCLlistaDirsLaz = ['lasfile-yy']
        else:
            LCLlistaDirsLaz = ['', 'lasfile-ce', 'lasfile-nw', 'lasfile-ne', 'lasfile-se', 'lasfile-sw', 'roquedos']
        LCLlistaSubDirsLaz = ['', 'RGBI_H29', 'RGBI', 'RGBI_laz_H29', 'RGBI_laz']
        print(f'\n{"":_^80}')
        print(f'clidaux-> ATENCION: ASIGNACION PROVISIONAL DE listaDirsLaz Y LCLlistaSubDirsLaz EN FUNCION DE LCLcuadrante:')
        print(f'{TB}cuando el procedimiento es AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA, AUTOMATICO_EN_CALENDULA o LAS_INFO')
        print(f'{TB}Lista de directorios y subdirectorios de {GLO.MAINrutaLaz} que se exploran:')
        print(f'{TB}-> listaDirsLaz:       {LCLlistaDirsLaz}')
        print(f'{TB}-> LCLlistaSubDirsLaz: {LCLlistaSubDirsLaz}')
        print(f'{"":=^80}')
    '''
    if LCLprocedimiento == 'PRECONFIGURADO_SINRUTA':
        bloqueElegido = 0
        if not LCLrutaLaz:
            LCLrutaLaz = preguntarRutaLaz(r'../laz/')
    elif LCLprocedimiento.startswith('AUTOMATICO_DISCOEXTERNO'):
        # Incluye 'AUTOMATICO_DISCOEXTERNO_UNIFICAR_RGBI_E'
        unidadLaz = LCLprocedimiento[-1:]
        # LCLrutaLaz = unidadLaz + ':/laz1/'
        # LCLrutaLaz = unidadLaz + ':/CE/CIR/LAS/Huso_30/'
        LCLrutaLaz = unidadLaz + ':/lidardata_2017_2021/lasfile_SE/IRC/'
    elif LCLprocedimiento.startswith('AUTOMATICO_SIGMENA'):
        rutaSigmenaIntercam = LCLprocedimiento[19:]
        LCLrutaLaz = os.path.join('O:Sigmena/intercam/', rutaSigmenaIntercam)
    elif LCLprocedimiento[:20] == 'AUTOMATICO_CUADRANTE':
        bloqueElegido = 0
        miCuadrante = LCLprocedimiento[21:23]
        miMarco = LCLprocedimiento[-8:]
        miUbicacionLaz = LCLprocedimiento[24:31]
        if LCLrutaLaz is None or LCLrutaLaz == '':
            if miUbicacionLaz == 'SIGMENA':
                LCLrutaLaz = 'O:/Sigmena/Intercam/laz/'
            else:
                input('ATENCION: nombre de procedimiento incorrecto: si empieza con AUTOMATIZADO_CUADRANTE solo admite SIGMENA en las posiciones 24:31')
                sys.exit(0)
        print('  MAINrutaLaz     ', LCLrutaLaz)

        if miCuadrante == 'XX':
            LCLlistaDirsLaz = ['2010_NW', '2010_NE', '2014_SW', '2010_SE']
        elif miCuadrante in ['NW', 'NE', 'SW', 'SE']:
            LCLlistaDirsLaz = ['2010_%s' % miCuadrante]
        else:
            input('ATENCION: nombre de procedimiento incorrecto: si empieza con AUTOMATIZADO_CUADRANTE solo admite cuadrantes NW, NE, SW o SE')
            sys.exit(0)
        if miMarco == 'SINMARCO':
            GLO.GLBLsoloCuadradoDeEjemplo = False
        elif miMarco == 'CONMARCO':
            GLO.GLBLsoloCuadradoDeEjemplo = True
        else:
            input('ATENCION: nombre de procedimiento incorrecto: si empieza con AUTOMATIZADO_CUADRANTE debe terminar con SINMARCO o CONMARCO')
            sys.exit(0)

    elif (
        LCLprocedimiento == 'PRECONFIGURADO_CONRUTA_4CUADRANTES_SIGMENA_CONOSINMARCO'
        or LCLprocedimiento == 'PRECONFIGURADO_CONRUTA_4CUADRANTES_DISCOEXTERNO_CONOSINMARCO'
    ):
        if LCLprocedimiento == 'PRECONFIGURADO_CONRUTA_4CUADRANTES_DISCOEXTERNO_CONOSINMARCO':
            selec = input('Para trabajar en disco externo (x:/LAZ/) pulsa la letra de unidad (C)')
            try:
                LCLrutaLaz = selec[:1] + ':/LAZ/'
            except:
                print('Entrada incorrecta')
                LCLrutaLaz = 'C:/LAZ/'
        else:
            #LCLrutaLaz = 'O:/Sigmena/Intercam/laz/'
            pass

        LCLlistaDirsLaz = elegirSubcarpetas(LCLrutaLaz)
        bloqueElegido = 0

        selec = input('\nProcesar solo determinados cuadrados de cada cuadrante? (n/s)')
        GLO.GLBLsoloCuadradoDeEjemplo = True if selec.upper() == 'S' else False

        if GLO.GLBLsoloCuadradoDeEjemplo:
            xMin_NW, yMin_NW, ladoMarco_NW = 200000, 4700000, 40000  # Zona generica
            xMin_NE, yMin_NE, ladoMarco_NE = 450000, 4650000, 20000  # Zona generica
            xMin_NE, yMin_NE, ladoMarco_NE = 458000, 4672000, 20000  # Cuadricula de ensayo Burgos
            xMin_NE, yMin_NE, ladoMarco_NE = 436000, 4760000, 20000  # Soncillo - Machorras
            xMin_SW, yMin_SW, ladoMarco_SW = 240000, 4584000, 22000  # Cuadricula de ensayo Zamora
            xMin_SE, yMin_SE, ladoMarco_SE = 490000, 4600000, 20000  # Zona generica
            coordenadasDeMarcos = {
                'NW': [xMin_NW, yMin_NW, ladoMarco_NW],
                'NE': [xMin_NE, yMin_NE, ladoMarco_NE],
                'SW': [xMin_SW, yMin_SW, ladoMarco_SW],
                'SE': [xMin_SE, yMin_SE, ladoMarco_SE],
            }
            listaValores = ['x min (inclusive)', 'y Min (inclusive)', 'lado del marco']
            listaCuadrantes = ['NW', 'NE', 'SW', 'SE']
            selec = input('\nEditar coordenadas para cada cuadrante? (n/s)')
            if selec.upper() == 'S':
                for cuadrante in listaCuadrantes:
                    selec = input('\nEditar coordenadas para el cuadrante %s? (n/s/x) (x: no procesarlo)' % cuadrante)
                    if selec.upper() == 'X':
                        coordenadasDeMarcos[cuadrante] = [0, 0, 0]
                    elif selec.upper() == 'S':
                        nuevosValores = []
                        for nValor, nombreValor in enumerate(listaValores):
                            strNuevoValor = input('Cuadrante %s -> %s (%i):' % (cuadrante, nombreValor, coordenadasDeMarcos[cuadrante][nValor]))
                            if strNuevoValor == '':
                                nuevosValores.append(coordenadasDeMarcos[cuadrante][nValor])
                            else:
                                nuevosValores.append(int(strNuevoValor))
                        coordenadasDeMarcos[cuadrante] = nuevosValores
        else:
            coordenadasDeMarcos = {}

    elif LCLprocedimiento == 'MANUAL':
        selec = input('Nombre de este PC (%s) -> ' % (GLO.MAINusuario))
        if selec != '':
            GLO.MAINusuario = selec
        print('Los ficheros de control empiezan con una linea %s\n' % (GLO.MAINusuario))

        selec = input('Ancho del pixel: (10 m)')
        try:
            GLO.GLBLmetrosCelda = float(selec)
        except:
            GLO.GLBLmetrosCelda = 10.0
        print('Ancho del pixel: %i m\n' % (GLO.GLBLmetrosCelda))

        LCLrutaLaz = preguntarRutaLaz(r'../laz/')

        selec = input('\nProcesar solo los que quedaron incompletos en anteriores sesiones? (n/s)')
        GLO.GLBLprocesarSoloIncompletos = True if selec.upper() == 'S' else False
        print(
            'Procesado normal'
            if GLO.GLBLprocesarSoloIncompletos
            else 'Se procesan los que quedaron incompletos en la primera vuelta (por falta de RAM disponible)'
        )
        GLO.GLBLreprocesarFallidosUsandoMenosRAM = False

        print('1. Uso de RAM standard')
        print('2. Uso de RAM alternativo')
        selec = input('Selecciona opcion 1-2 (1):')

        try:
            if int(selec) == 1:
                GLO.GLBLusoDeRAM = 'standard'
            elif int(selec) == 2:
                GLO.GLBLusoDeRAM = 'alternativo'
            else:
                sys.exit()
        except:
            GLO.GLBLusoDeRAM = 'standard'
        print('Opcion elegida: %s' % (GLO.GLBLusoDeRAM))

        selec = input('\nNumero de registros a procesar: (por defecto: todos -> Escribir 0 o [enter] directamente)')
        try:
            GLO.GLBLnumeroDePuntosAleer = int(selec)
        except:
            GLO.GLBLnumeroDePuntosAleer = 0
        print('Numero de puntos: %s' % (str(GLO.GLBLnumeroDePuntosAleer) if GLO.GLBLnumeroDePuntosAleer != 0 else 'todos'))

        selec = input('\nCalcular percentiles de valores absolutos (altitud o cota absoluta): (s/n)')
        GLO.GLBLgrabarPercentilesAbsolutos = False if selec.upper() == 'N' else True
        print('Calcular percentiles absolutos:', GLO.GLBLgrabarPercentilesAbsolutos)

        selec = input('\nCalcular percentiles de valores relativos (altura sobre el plano-suelo o plano-basal): (s/n)')
        GLO.GLBLgrabarPercentilesRelativos = False if selec.upper() == 'N' else True
        print('Calcular percentiles relativos:', GLO.GLBLgrabarPercentilesRelativos)
        # ==========================================================================

        exploraDirectorios = False
        # selec = input('Explorar directorios de '+LCLrutaLaz+ ' (n/s)')
        # exploraDirectorios = True if selec.upper() == 'S' else False

        if exploraDirectorios:
            LCLlistaDirsLaz = []
            for (_, dirnames, _) in os.walk(LCLrutaLaz):
                LCLlistaDirsLaz.extend(dirnames)
                # files.extend(filenames)
                break
            print(
                'Directorios en %s:' % (LCLrutaLaz),
            )
            print(LCLlistaDirsLaz)
            if len(LCLlistaDirsLaz) == 0:
                print('No se rastrean los directorios')
                LCLlistaDirsLaz = ['']
            # bloqueElegido = 5
        else:
            print('\n0. Procesar ficheros de %s' % (LCLrutaLaz))
            print('1. Procesar los ficheros laz de 2010_NW (%s)' % (LCLrutaLaz + '2010_NW/'))
            print('2. Procesar los ficheros laz de 2010_NW (%s)' % (LCLrutaLaz + '2010_NW_las/'))
            print('3. Procesar los ficheros las de 2010_NE (%s)' % (LCLrutaLaz + '2010_NE/'))
            print('4. Procesar los ficheros laz de 2010_NE (%s)' % (LCLrutaLaz + '2010_NE_las/'))
            print('5. Procesar los ficheros laz de 2014_SW (%s)' % (LCLrutaLaz + '2014_SW/'))
            print('6. Procesar los ficheros las de 2014_SW (%s)' % (LCLrutaLaz + '2014_SW_las/'))
            print('7. Procesar los ficheros laz de 2010_SE (%s)' % (LCLrutaLaz + '2010_SE/'))
            print('8. Procesar los ficheros las de 2010_SE (%s)' % (LCLrutaLaz + '2010_SE_las/'))
            try:
                bloqueElegido = int(input('Selecciona opcion 1-9 (0):'))
            except:
                bloqueElegido = 0
            if bloqueElegido == 1:
                LCLlistaDirsLaz = ['2010_NW/']
            elif bloqueElegido == 2:
                LCLlistaDirsLaz = ['2010_NW_las/']
            elif bloqueElegido == 3:
                LCLlistaDirsLaz = ['2010_NE/']
            elif bloqueElegido == 4:
                LCLlistaDirsLaz = ['2010_NE_las/']
            elif bloqueElegido == 5:
                LCLlistaDirsLaz = ['2014_SW/']
            elif bloqueElegido == 6:
                LCLlistaDirsLaz = ['2014_SW_las/']
            elif bloqueElegido == 7:
                LCLlistaDirsLaz = ['2010_SE/']
            elif bloqueElegido == 8:
                LCLlistaDirsLaz = ['2010_SE_las/']
            else:
                LCLlistaDirsLaz = ['']

        selec = input('Crear ficheros las de forma permanente (no se borran una vez procesados) (S/n)')
        GLO.GLBLficheroLasTemporal = True if selec.upper() == 'N' else False
        if GLO.GLBLficheroLasTemporal:
            print('Los ficheros las creados son temporales (se borran tras procesarlos)\n')
        else:
            print('Los ficheros las creados son permanentes\n')

    elif LCLprocedimiento.startswith('CREAR_CAPA_CON_UNA_PROPIEDAD_DE_LOS_FICHEROS_LIDAR'):
        # Rutas por defecto, o bien:
        if not LCLrutaLaz:
            # LCLrutaLaz = 'E:/lidardata_2017_2021/lasfile_SE/IRC'
            LCLrutaLaz = 'E:/lidardata_2017_2021/lasfile_CE/IRC'
        pass
    elif LCLprocedimiento == 'CREAR_SHAPE':
        LCLlistaDirsLaz = ['']
        bloqueElegido = 0
        LCLrutaLaz = preguntarRutaLaz(r'../laz/')
        selec = input('\nNumero de puntos del shape: (1000 por defecto; indicar 0 para todos)')
        try:
            GLO.GLBLshapeNumPoints = int(selec)
        except:
            GLO.GLBLshapeNumPoints = 1000
        print('\nNumero de puntos: %s' % (str(GLO.GLBLshapeNumPoints) if GLO.GLBLshapeNumPoints != 0 else 'todos'))
        if GLO.GLBLshapeNumPoints != 0:
            GLO.GLBLtipoLectura = 'registrosPorLotes'
        else:
            GLO.GLBLtipoLectura = 'registrosIndividuales'

        print('\n1. Todas las clases')
        print('2. Todas las clases menos la 12 (relleno)')
        print('3. Solo puntos de borde de escaneo')
        selec = input('Selecciona opcion (1):')
        try:
            if int(selec) == 2:
                GLO.GLBLshapeFilter = 'sinClase12'
            elif int(selec) == 3:
                GLO.GLBLshapeFilter = 'soloEdge'
            else:
                GLO.GLBLshapeFilter = 'noFilter'
        except:
            GLO.GLBLshapeFilter = 'noFilter'
        print('Tipo de filtro: %s\n' % (GLO.GLBLshapeFilter))


    elif LCLprocedimiento == 'RENOMBRAR_FICHEROS':
        LCLrutaLaz = 'O:/Sigmena/Intercam/laz/'

        LCLlistaDirsLaz = ['2010_NW', '2010_NE', '2014_SW', '2010_SE']
        LCLlistaDirsLaz = ['2010_NW_Huso_29', '2010_NW_las']
        GLO.GLBLprocesarComprimidosLaz = True
        bloqueElegido = 0

    elif LCLprocedimiento == 'MERGEAR':
        input('Lo desarrollo en raster2vector de clidgis.py (rescatado de copiaSeg/2017/lidasMerge.py)')
        sys.exit()

    elif LCLprocedimiento == 'GEOINTEGRAR':
        input('Lo desarrollo en raster2vector de clidcluster.py (antes GIS/cluster.py, renombrado a cluster_old_VerCartolid_clidax.py')
        sys.exit()

    else:
        print('\nRevisar el nombre del procedimiento en cartolid.xlm. MAINprocedimiento:', LCLprocedimiento)
        sys.exit()

    # print('clidaux-> 3b LCLrutaLaz:', LCLrutaLaz)

    return LCLrutaLaz, LCLlistaDirsLaz, LCLlistaSubDirsLaz


# ==============================================================================
def preguntarRutaLaz(rutaInicial):
    rutaNormalizada = os.path.abspath(rutaInicial)
    # rutaNormalizada = os.path.normpath(rutaInicial)
    # rutaNormalizada = quitarContrabarrasAgregarBarraFinal(rutaNormalizada)
    rutaNormalizada = rutaNormalizada.replace(os.sep, '/')

    print('\nSe solicitan datos:')
    selec = input('\nEscribir la ruta base de los ficheros lidar (.laz o las; p. ej. C:/laz/): (por defecto: %s)' % rutaNormalizada)
    if selec == '':
        rutaFinal = rutaNormalizada
    else:
        # rutaFinal = quitarContrabarrasAgregarBarraFinal(selec)
        rutaFinal = selec.replace(os.sep, '/')
    print('Ruta elegida: %s' % (rutaFinal))
    return rutaFinal


# ==============================================================================
def elegirSubcarpetas(USERmiRutaRaiz):
    print('\nElegir directorios:')
    print('        (0) Solo los ficheros que cuelgan directamente de %s' % USERmiRutaRaiz)
    print('        (1) Las subcarpetas "2010_NW", "2010_NE", "2014_SW" y "2010_SE"')
    print('        (2) Las subcarpetas "2010_NW_las", "2010_NE_las", "2014_SW_las" y "2010_SE_las"')
    print('        (3) Solo la subcarpeta "2014_SW"')
    print('        (4) Solo la subcarpeta "2010_SE"')
    print('        (6) Ficheros de subcarpetas Zona1 y Zona2, que cuelgan de %s2009' % USERmiRutaRaiz)
    print('        (7) Ficheros de subcarpetas Zona1 y Zona2, que cuelgan de %s2017' % USERmiRutaRaiz)
    print('        (8) Ficheros de subcarpetas Zona1 y Zona2, que cuelgan de %s2009 y %s2017' % (USERmiRutaRaiz, USERmiRutaRaiz))
    print('        (9) Todos los ficheros de todas las subcarpetas que cuelgan de %s (sin implementar)' % USERmiRutaRaiz)
    print('(1) y (3) tienen todos los ficheros y (2) solo ficheros ya descomprimidos.')
    selec = input('\nSelecciona las subcarpetas a procesar (0)')
    if selec.upper() == '1':
        listaDirsLaz = ['2010_NW', '2010_NE', '2014_SW', '2010_SE']
    elif selec.upper() == '2':
        GLO.GLBLdescomprimirLaz = False
        listaDirsLaz = ['2010_NW_las', '2010_NE_las', '2014_SW_las', '2010_SE_las']
    elif selec.upper() == '3':
        listaDirsLaz = ['2014_SW']
    elif selec.upper() == '4':
        listaDirsLaz = ['2010_SE']
    elif selec.upper() == '6':
        listaDirsLaz = ['2009/Zona1', '2009/Zona2']
    elif selec.upper() == '7':
        listaDirsLaz = ['2017/Zona1', '2017/Zona2']
    elif selec.upper() == '8':
        listaDirsLaz = ['2009/Zona1', '2009/Zona2', '2017/Zona1', '2017/Zona2']
    elif selec.upper() == '9':
        listaDirsLaz = ['']
    else:
        listaDirsLaz = ['']
    return listaDirsLaz





#ççç













# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# He modificado la funcion para que el tipo se deduzca del valor;
# Ya no es necesario poner el tipo de variable en el fichero de configuracion,
# salvo que quiera que un numero se interprete como texto, por ejemplo: 0000
# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def valorConfig(valorPrincipalTxt, valorAlternativoTxt='', usarAlternativo=False, nombreParametro='SinNombre', tipoVariable=''):
    # if nombreParametro == 'GLBLmetrosCelda':
    #     print('clidconfig->Borrar-> nombreParametro', nombreParametro, valorPrincipalTxt, tipoVariable)
    if valorPrincipalTxt == None:
        valorPrincipalOk = None
        tipoVariable = 'NoneType'
    elif valorPrincipalTxt == 'None':
        valorPrincipalOk = None
    elif tipoVariable == 'str':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. 0000 puede ser str
        valorPrincipalOk = valorPrincipalTxt
    elif tipoVariable == 'dict_str':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. '[0000, 1]' puede ser list_str
        try:
            valorPrincipalOk = {valorPrincipalTxt.split(',')[0]: valorPrincipalTxt.split(',')[1:]}
        except:
            valorPrincipalOk = {}
    elif tipoVariable == 'dict_int':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. '[0000, 1]' puede ser list_str
        try:
            valorPrincipalOk = {valorPrincipalTxt.split(',')[0]: [int(item) for item in valorPrincipalTxt.split(',')[1:]]}
        except:
            valorPrincipalOk = {}
    elif tipoVariable == 'list_str':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. '[0000, 1]' puede ser list_str
        valorPrincipalOk = valorPrincipalTxt.split(',')
        # print('-->>', nombreParametro, tipoVariable, type(valorPrincipalOk), valorPrincipalOk)
    elif tipoVariable == 'list_int':
        valorPrincipalList = valorPrincipalTxt.split(',')
        valorPrincipalOk = []
        try:
            valorPrincipalOk = [int(item) for item in valorPrincipalList]
        except:
            print('clidconfig-> Error en parametro', nombreParametro, 'tiene formato distinto al previsto:', tipoVariable)
        # print('-->>', nombreParametro, tipoVariable, type(valorPrincipalOk), valorPrincipalOk)
    elif tipoVariable == 'list_float':
        valorPrincipalList = valorPrincipalTxt.split(',')
        valorPrincipalOk = []
        try:
            valorPrincipalOk = [float(item) for item in valorPrincipalList]
        except:
            print('clidconfig-> Error en parametro', nombreParametro, 'tiene formato distinto al previsto.')
    else:
        # Incluye el tipoVariable == 'desconocido'
        try:
            valorPrincipalFloat = float(valorPrincipalTxt)
            esNumero = True
        except:
            esNumero = False
        if esNumero:
            try:
                if valorPrincipalFloat == int(valorPrincipalFloat):
                    valorPrincipalOk = int(valorPrincipalFloat)
                    tipoVariable = 'int'
                else:
                    valorPrincipalOk = valorPrincipalFloat
                    tipoVariable = 'float'
            except:
                valorPrincipalOk = ''
                tipoVariable = 'error'
        else:
            if valorPrincipalTxt == 'True':
                valorPrincipalOk = True
                tipoVariable = 'bool'
            elif valorPrincipalTxt == 'False':
                valorPrincipalOk = False
                tipoVariable = 'bool'
            elif valorPrincipalTxt == 'None':
                valorPrincipalOk = None
                # No se el tipo de variable, pero todas las None de lisdas.xml son str
                tipoVariable = 'str'
            else:
                valorPrincipalOk = valorPrincipalTxt
                tipoVariable = 'str'
        # print('-->>', nombreParametro, tipoVariable, type(valorPrincipalOk), valorPrincipalOk)
        # if nombreParametro == 'GLBLmetrosCelda':
        #     print('clidconfig->Borrar-> int/float valorPrincipalOk', valorPrincipalOk, type(valorPrincipalOk))

    if usarAlternativo and valorAlternativoTxt != '':
        if tipoVariable == 'NoneType':
            valorAlternativoOk = None
        elif tipoVariable == 'str':
            valorAlternativoOk = valorAlternativoTxt
        elif tipoVariable == 'bool':
            if valorAlternativoTxt == 'True':
                valorAlternativoOk = True
            elif valorAlternativoTxt == 'False':
                valorAlternativoOk = False
            else:
                print('clidconfig-> Valor alternativo del parametro', nombreParametro, 'tiene formato distinto al principal; se usa el valor principal')
                valorAlternativoOk = valorPrincipalOk
        elif tipoVariable == 'int':
            try:
                valorAlternativoOk = int(valorAlternativoTxt)
            except:
                print('clidconfig-> Valor alternativo del parametro', nombreParametro, 'tiene formato distinto al principal; se usa el valor principal')
                valorAlternativoOk = valorPrincipalOk
        elif tipoVariable == 'float':
            try:
                valorAlternativoOk = float(valorAlternativoTxt)
            except:
                print('clidconfig-> Valor alternativo del parametro', nombreParametro, 'tiene formato distinto al principal; se usa el valor principal')
                valorAlternativoOk = valorPrincipalOk
        else:
            valorPrincipalOk = None
            valorAlternativoOk = None
        valorElegido = valorAlternativoOk
    else:
        valorElegido = valorPrincipalOk
    # if nombreParametro == 'GLBLmetrosCelda':
    #     print('clidconfig->Borrar-> valorElegido', nombreParametro, valorElegido, type(valorElegido))
    return valorElegido


# ==============================================================================
def get_MAIN_CFG_DIR(cfgSubDir='data/cfg'):
    # Orden de preferencia para el fichero de configuracion:
    #  1-> El del directorio actual (desde el que se lanza la aplicacion) si no incluye site-packages
    #      y no es calendula para no acumular .cfg en /LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1 y no crear un /LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1/cfg/
    #  2-> El del modulo inicial si no incluye site-packages
    #  3-> El del modulo actual si no incluye site-packages
    #  4-> El user/documents
    MAIN_WORK_DIR = os.getcwd()
    if not 'site-packages' in MAIN_WORK_DIR and MAIN_ENTORNO != 'calendula':
        MAIN_CFG_DIR = os.path.join(MAIN_WORK_DIR, cfgSubDir)
    else:
        if not 'site-packages' in MAIN_BASE_DIR:
            MAIN_CFG_DIR = os.path.join(MAIN_BASE_DIR, cfgSubDir)
        else:
            if not 'site-packages' in MAIN_FILE_DIR:
                MAIN_CFG_DIR = os.path.join(MAIN_FILE_DIR, cfgSubDir)
            else:
                MAIN_HOME_DIR = str(pathlib.Path.home())
                MAIN_CFG_DIR = os.path.join(MAIN_BASE_DIR, 'Documents', cfgSubDir)
    if not os.path.isdir(MAIN_CFG_DIR):
        try:
            os.makedirs(MAIN_CFG_DIR)
        except:
            print(f'clidconfig-> No se ha podido crear el directorio {MAIN_CFG_DIR}. Revisar derechos de escritura en esa ruta.')
            sys.exit(0)
    return MAIN_CFG_DIR


# ==============================================================================
def getConfigFileName():
    if len(sys.argv) == 0:
        print(f'\nclidconfig-> Revisar esta forma de ejecucion. sys.argv: <{sys.argv}>')
        sys.exit(0)
    elif (sys.argv[0].endswith('__main__.py') and 'cartolidar' in sys.argv[0]):
        # print('\nqlidtwins.py se ejecuta lanzando el paquete cartolidar desde linea de comandos:')
        # print('\t  python -m cartolidar')
        configFileNameSinExt = 'clidbase{:006}'.format(int(MAIN_idProceso))
    elif sys.argv[0].endswith('qlidtwins.py'):
        # print('\nqlidtwins.py se ha lanzado desde linea de comandos:')
        # print('\t  python qlidtwins.py')
        configFileNameSinExt = 'qlidtwins'
    elif sys.argv[0] == '':
        # print('\nqlidtwins se esta importando desde el interprete interactivo:')
        configFileNameSinExt = 'clidbase{:006}'.format(int(MAIN_idProceso))
    elif sys.argv[0].endswith('clidtools.py'):
        # print('\nclidtools.py se ejecuta lanzandolo desde linea de comandos:')
        # print('\t  python clidtools.py')
        configFileNameSinExt = 'clidtools{:006}'.format(int(MAIN_idProceso))
    else:
        # print(f'\nqlidtwins.py se esta importando desde el modulo: {sys.argv[0]}')
        if MAIN_idProceso:
            if not type(MAIN_idProceso) == int and not type(MAIN_idProceso) == str:
                print('\nclidconfig-> AVISO: Revisar asignacion de idProceso (no es int ni str):')
                print('idProceso:   <{}> type: {}'.format(MAIN_idProceso, type(MAIN_idProceso)))
                print('sys.argv[0]: <{}>'.format(sys.argv[0]))
            try:
                if sys.argv[0].endswith('.py'):
                    configFileNameSinExt = os.path.basename(sys.argv[0]).replace('.py', '{:006}'.format(int(MAIN_idProceso)))
                elif sys.argv[0].endswith('pytest'):
                    configFileNameSinExt = 'cfgForTest'
                else:
                    configFileNameSinExt = 'unknownLaunch'
            except:
                print('\nclidconfig-> Revisar asignacion de idProceso (b):')
                print('idProceso:   <{}> type: {}'.format(MAIN_idProceso, type(MAIN_idProceso)))
                print('sys.argv[0]: <{}>'.format(sys.argv[0]))
                sys.exit(0)
        else:
            if sys.argv[0].endswith('.py'):
                configFileNameSinExt = os.path.basename(sys.argv[0]).replace('.py', '')
            elif sys.argv[0].endswith('pytest'):
                configFileNameSinExt = 'cfgForTest'
            else:
                configFileNameSinExt = 'unknownLaunch'

    return configFileNameSinExt


# ==============================================================================
def getConfigFileNameCfg(idProceso, LOCL_verbose=__verbose__):
    MAIN_CFG_DIR = get_MAIN_CFG_DIR(cfgSubDir='data/cfg')
    configFileNameSinExt = getConfigFileName()
    configFileNameSinPath = f'{configFileNameSinExt}.cfg'

    configFileNameCfg = os.path.join(MAIN_CFG_DIR, configFileNameSinPath)
    if LOCL_verbose:
        print(f'\n{"":_^80}')
        print(f'clidconfig-> Buscando fichero de configuracion cfg: {configFileNameCfg}')

    try:
        if not os.path.exists(configFileNameCfg):
            controlConfigFile = open(configFileNameCfg, mode='w')
            controlConfigFile.close()
            os.remove(configFileNameCfg)
            if LOCL_verbose:
                print(f'{TB}-> No hay de configuracion de tipo cfg previamente creado-> se usa el xlsx.')
        else:
            controlConfigFile = open(configFileNameCfg, mode='r+')
            controlConfigFile.close()
            if LOCL_verbose:
                print(f'{TB}-> Este fichero de configuracion ya existe previamente.')
    except:
        if LOCL_verbose:
            # print(f'\n{"":_^80}')
            print(f'clidconfig-> AVISO:')
            print(f'{TB}-> No se puede guardar el fichero de configuracion:  {configFileNameCfg}')
            print(f'{TB}-> Es posible que no tenga permisos de escritura en: {MAIN_CFG_DIR}')
            print(f'{TB}-> O que exista el fichero {configFileNameCfg} y este bloqueado.')
            print(f'{TB}-> callingModulePrevio: {callingModulePrevio}')
        try:
            MAIN_HOME_DIR = str(pathlib.Path.home())
            MAIN_CFG_DIR = os.path.join(MAIN_HOME_DIR, 'Documents')
            if LOCL_verbose:
                print(f'{TB}Se intenta la ruta alternativa: {MAIN_CFG_DIR}')
            configFileNameCfg = os.path.join(MAIN_CFG_DIR, configFileNameSinPath)
            if not os.path.exists(configFileNameCfg):
                controlConfigFile = open(configFileNameCfg, mode='w')
                controlConfigFile.close()
                os.remove(configFileNameCfg)
            else:
                controlConfigFile = open(configFileNameCfg, mode='r+')
                controlConfigFile.close()
            if LOCL_verbose:
                print(f'{TB}Ok log file: {configFileNameCfg}')
        except:
            if LOCL_verbose:
                print(f'{TB}Tampoco se puede escribir en la ruta {MAIN_CFG_DIR}')
            MAIN_HOME_DIR = str(pathlib.Path.home())
            configFileNameCfg = os.path.join(MAIN_HOME_DIR, os.path.basename(sys.argv[0]).replace('.py', '.cfg'))
            if LOCL_verbose:
                print(f'clidconfig-> Ok cfg file (b): {configFileNameCfg}')
        if LOCL_verbose:
            if os.path.exists(configFileNameCfg):
                print(f'{TB}-> Este fichero de configuracion ya existe previamente.')

    if LOCL_verbose:
        print(f'{"":=^80}')
    return configFileNameCfg


# ==============================================================================
def getConfigFileNameXls(
        configFileNameCfg,
        LOCL_verbose=__verbose__,
    ):
    MAIN_XLS_DIR = get_MAIN_CFG_DIR(cfgSubDir='')
    MAIN_CFG_DIR = get_MAIN_CFG_DIR(cfgSubDir='data/cfg')
    moduloBaseXls = 'clidbase.xlsx'
    moduloInicXls = (os.path.basename(sys.argv[0])).replace('.py', '.xlsx')
    directorioActual = (os.getcwd()).replace(os.sep, '/')
    directorioProyecto = os.path.dirname(sys.argv[0])
    directorioCfg = os.path.dirname(configFileNameCfg)

    # Opcion 1: fichero de configuracion con el nombre clidbase.xls
    #           en el directorio inicialmente propuesto para el cfg:
    # Orden de preferencia para el fichero de configuracion:
    #  1-> El del directorio actual (desde el que se lanza la aplicacion) si no incluye site-packages
    #      y no es calendula para no acumular .cfg en /LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1 y no crear un /LUSTRE/HOME/jcyl_spi_1/jcyl_spi_1_1/cfg/
    #  2-> El del modulo inicial si no incluye site-packages
    #  3-> El del modulo actual si no incluye site-packages
    #  4-> El user/documents
    filenameXlsCfg1Clidbase = os.path.join(MAIN_XLS_DIR, moduloBaseXls)
    # No verifico si hay derechos de escritura en este directorio (cosa que si hago para cfg)

    # Opcion 2: fichero de configuracion con el nombre clidbase.xls
    #           en el directorio actual (desde el que se lanza la aplicacion)
    #           Puede coincidir con el anterior si no incluye site-packages
    filenameXlsCwdClidbase = os.path.join(directorioActual, moduloBaseXls)

    # Opcion 3: fichero de configuracion con el nombre del modulo que se lanza inicialmente (__init__.py, __main__.py, clidbase.py, etc.)
    #           en el directorio actual (desde el que se lanza la aplicacion)
    #           Puede coincidir con los anteriores si el modulo inicial es clidbase.py
    filenameXlsCwdModulo = os.path.join(directorioActual, moduloInicXls)

    # Opcion 4: fichero de configuracion con el nombre clidbase.xls
    #           en el directorio finalmente propuesto para el cfg:
    filenameXlsCfg2Clidbase = os.path.join(directorioCfg, moduloBaseXls)

    # Opcion 5: fichero de configuracion instalado junto al modulo que lanza la aplicacion (clidbase.py)
    filenameXlsProjClidbase = os.path.abspath(os.path.join(directorioProyecto, moduloBaseXls))

    return (
        filenameXlsCfg1Clidbase,
        filenameXlsCwdClidbase,
        filenameXlsCwdModulo,
        filenameXlsCfg2Clidbase,
        filenameXlsProjClidbase,
        MAIN_CFG_DIR,
    )








#ççç








# ==============================================================================
def creaDirectorio(rutaDirectorio):
    # Parecido a os.makedirs(), pero que este no crea todo el arbol de directorios,
    # sino solo intenta crear el directorio y su padre.
    if not os.path.exists(rutaDirectorio):
        try:
            os.mkdir(rutaDirectorio)
        except:
            rutaPadre = os.path.abspath(os.path.join(rutaDirectorio, '..'))
            try:
                os.mkdir(rutaPadre)
                os.mkdir(rutaDirectorio)
                print('clidaux-> Se ha creado el directorio %s despues de crear su dir padre: %s' % (rutaDirectorio, rutaPadre))
            except:
                print('clidaux-> No se ha podido crear el directorio %s ni su dir padre %s' % (rutaDirectorio, rutaPadre))
            sys.exit(0)


# ==============================================================================
def creaDirectorios(GLOBAL_rutaResultados, listaSubdirectorios=[]):
    if not os.path.exists(GLOBAL_rutaResultados):
        print('No existe el directorio %s -->> Se crea automaticamente...' % (GLOBAL_rutaResultados))
    listaDirectorios = [
        GLOBAL_rutaResultados,
        GLOBAL_rutaResultados + 'Ajustes/',
        GLOBAL_rutaResultados + 'Ajustes/Basal/',
        GLOBAL_rutaResultados + 'Ajustes/Suelo/',
        GLOBAL_rutaResultados + 'Alt/',
        GLOBAL_rutaResultados + 'AltClases/',
        GLOBAL_rutaResultados + 'Clasificacion/',
        GLOBAL_rutaResultados + 'CobClases/',
        GLOBAL_rutaResultados + 'Fcc/',
        GLOBAL_rutaResultados + 'Fcc/RptoAzMin_MasDe/',
        GLOBAL_rutaResultados + 'Fcc/RptoAsuelo_MasDe/',
        GLOBAL_rutaResultados + 'Fcc/RptoAmds_MasDe/',
        GLOBAL_rutaResultados + 'Fcc/RptoAmds/',
        GLOBAL_rutaResultados + 'Fcc/RptoAmdb/',
        GLOBAL_rutaResultados + 'FormasUsos/',
        GLOBAL_rutaResultados + 'NumPtosPasadas/',
        GLOBAL_rutaResultados + 'NumPtosPasadas/PorRetornos/',
        GLOBAL_rutaResultados + 'NumPtosPasadas/PorClases/',
        GLOBAL_rutaResultados + 'OrientPte/',
        GLOBAL_rutaResultados + 'Varios/',
        GLOBAL_rutaResultados + 'z/',
    ]
    for directorio in listaDirectorios:
        if not os.path.exists(directorio):
            try:
                os.makedirs(directorio)
            except:
                print('No se ha podido crear el directorio %s' % (directorio))
                sys.exit()


# ==============================================================================
def creaRutaDeFichero(rutaFichero):
    rutaDirectorio = os.path.dirname(os.path.realpath(rutaFichero))
    miFileNameSinPath = os.path.basename(os.path.realpath(rutaFichero))
    if not os.path.exists(rutaDirectorio):
        print(f'clidaux-> Creando ruta {rutaDirectorio} para {miFileNameSinPath}')
        try:
            os.makedirs(rutaDirectorio)
        except:
            print(f'\nclidaux-> ATENCION: No se ha podido crear el directorio {rutaDirectorio}')
            sys.exit()


# ==============================================================================
def cerrarFicheroLasPorErrorInvalidante(
        controlFileGralName,
        controlFileLasName,
        controlFileLasObj,
        mensajeError,
    ):
    print(f'clidconfig-> clidconfig-> {mensajeError}')

    try:
        controlFileGralObj = open(controlFileGralName, mode='a+')
        controlFileGralObj.write(f'{mensajeError}\n')
        controlFileGralObj.write('Ver fichero de control correspondiente.\n')
        controlFileGralObj.close()
    except:
        pass

    if not controlFileLasName is None:
        try:
            controlFileLasObj = open(controlFileLasName, mode='a+')
        except:
            pass
    if not controlFileLasObj is None:
        try:
            controlFileLasObj.write(f'{mensajeError}\n')
            controlFileLasObj.close()
        except:
            pass

    # liberaArraysVuelta1(interrumpidoPorError=True, completo=False)
    return


# ==============================================================================
def crearListaVariablesEntrenamientoAcumulativo(nInputVars=0):
    dictVariablesEndogenasTodas = {

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> Requieren RGBI
        # miPto['red']..miPto['nir'] extraidos de self_aCeldasListaDePtosTlcPralPF99[] -> clidnv0->numbaMainVuelta0{}
        # RGBI de cada punto.
        # ATENCION: se usan distintas escalas en distintos ficheros. Ver clidflow -> #RangoDeValoresRGBI
        0: 'red',
        1: 'green',
        2: 'blue',
        3: 'nir',

        # miPto['intensity'] extraido de self_aCeldasListaDePtosTlcPralPF99[] -> clidnv0->numbaMainVuelta0{}
        # Intensidad de cada punto.
        # ATENCION: se usan distintas escalas en distintos ficheros. Ver clidflow -> #RangoDeValoresRGBI
        4: 'intensity',

        # self_aMetricoIntSRet[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # Valor medio de la intensidad de los puntos con un solo retorno en cada metro cuadrado (mas buffer).
        # ATENCION: se usan distintas escalas en distintos ficheros.  Ver clidflow -> #RangoDeValoresRGBI
        5: 'metroIntensitySRet',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> No uso estas variables
        # Indices calculadas con las propiedades RGBI de miPto[] extraidas de self_aCeldasListaDePtosTlcPralPF99[] -> clidnv0->numbaMainVuelta0{}
        # Re-escalado de puntoEVI2calculado = 2.5 * (miPtoNewNir - miPtoNewRed) / np.float32(miPtoNewNir + (2.4 * miPtoNewRed) + 1)
        6: 'puntoEVI2ReEscalado',
        # Re-escalado de puntoNDVIcalculado = (miPtoNewNir255 - miPtoNewRed255) / sumaNirRed
        7: 'puntoNDVIReEscalado',
        # Re-escalado de puntoNDWIcalculado = (miPtoNewNir255 - miPtoNewRed255) / sumaGreenNir
        8: 'puntoNDWIReEscalado',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # self_aMetricoEVI2Med[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # self_aMetricoEVI2Med[nMetX, nMetY] += int(100 * (2.5 * (float(miPtoNewNir - miPtoNewRed) / (float(miPtoNewNir) + (2.4 * float(miPtoNewRed)) + 1))))
        9: 'metroEVI2ReEscalado',
        # self_aMetricoNDVIMed[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # self_aMetricoNDVIMed[nMetX, nMetY] += int(100 * (float(miPtoNewNir - miPtoNewRed) / (float(miPtoNewNir) + float(miPtoNewRed))))
        10: 'metroNDVIReEscalado',
        # self_aMetricoNDWIMed[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # self_aMetricoNDWIMed[nMetX, nMetY] += int(100 * ((miPtoNewGreen - miPtoNewNir) / (float(miPtoNewGreen) + float(miPtoNewNir))))
        11: 'metroNDWIReEscalado',

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> No uso estas variables
        # Esta variable es menor de 100 cuando no es lastReturn (% rpto a un total de retornos del punto) y
        # superior a 100 si es last return, tanto mayor cuantos mas retornos haya (120->~200)
        # miPtoRetNumRel = (100.0 * miPtoReturnNumber / miPtoReturnTotal) + (20 * numLastReturn)
        # Multiplico el numero total de retornos x 50 para ampliar el rango hasta 250 (en caso de 5 retornos)
        12: 'retNumRel250',
        # Numero total de retornos x 50: miPtoReturnTotal * 50
        13: 'retNumTotX50',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        14: 'miPtoZdecametros', # Cota del punto en decametros

        # Estas variables se calculan en clidnv1.procesaCeldasVuelta1b<>->clidnv1.buscarGuardarPuntosMiniSubCel<> 
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        # mseMx50MicroPlanoNubePuntual = miPtoNpExtrVar['mseMx50MicroPlanoNubePuntual']
        # 15: 'mseMx50MicroPlanoNubePuntual',
        # mseCmMicroPlanoNubePuntual = miPtoNpExtrVar['mseMx50MicroPlanoNubePuntual']
        15: 'mseCmMicroPlanoNubePuntual',
        # ptePctjMicroPlanoNubePuntual = miPtoNpExtrVar['ptePctjMicroPlanoNubePuntual']
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        16: 'ptePctjMicroPlanoNubePuntual',
        # cotaRelDmMinNubePuntual = miPtoNpExtrVar['cotaRelDmMinNubePuntual']
        # Es la diferencia (abs) de cotas entre este punto y el mas bajo de la nube puntual calculado
        #  en el mini enjambre de puntos con radio: GLBLradioNubeDeCadaPunto y un maximo de puntos: GLBLmaxPtNubeDeCadaPunto)
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        17: 'cotaRelDmMinNubePuntual',
        # cotaRelDmPlanoNubePuntual = miPtoNpExtrVar['cotaRelDmPlanoNubePuntual']
        # Es la diferencia (abs) de cotas entre este punto y ?el mas alto? de la nube puntual calculado
        #  en el mini enjambre de puntos con radio: GLBLradioNubeDeCadaPunto y un maximo de puntos: GLBLmaxPtNubeDeCadaPunto)
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        18: 'cotaRelDmPlanoNubePuntual',

        19: 'random0a255_1',
        20: 'random0a255_2',

        # ///////////////////////////////////////////////// variables opcionales-> si GLBLcalcularHiperFormas
        # Fuente: self_aMetricoRugosidadMacroInterCeldillas[metrX, metrY] -> calculado en clidnv4.procesaCeldasVuelta4{}
        21: 'metricoRugoMacroReEscalado',
        # self_aMetricoRugosidadMesosInterCeldillas[metrX, metrY] -> clidnv4.procesaCeldasVuelta4{}
        22: 'metricoRugoMesosReEscalado',
        # self_aMetricoRugosidadMicroInterCeldillas[metrX, metrY] -> clidnv4.procesaCeldasVuelta4{}
        23: 'metricoRugoMicroReEscalado',
        # self_aSubCeldasRugosidadMacroInterCeldillas -> clidnv4.procesaCeldasVuelta4{}
        24: 'subCeldasRugoMacroReEscalado',
        # self_aSubCeldasRugosidadMesosInterCeldillas -> clidnv4.procesaCeldasVuelta4{}
        25: 'subCeldasRugoMesosReEscalado',
        # self_aSubCeldasRugosidadMicroInterCeldillas -> clidnv4.procesaCeldasVuelta4{}
        26: 'subCeldasRugoMicroReEscalado',
        # self_aMetricoPlanoTejado[metrX, metrY] -> clidnv4.procesaCeldasVuelta4{}
        27: 'metricoPlanoTejadoReEscalado',
        # //////////////////////////////////////////////////////////////////////

        # >>>>>>>>>>>>>>>>>>>>> Calculadas a partir de MDS cartolid aproximativo
        # Esta variable la uso si GLBLcalcularMdb and GLBLincluirVarDeMdbEnElModeloAcumulativo
        # CotaEnDmPlus20SobreMdb calculada a partir de cotaSobreMdbEnMetros, calculada a partir de cotaCmSobreMdb32bits
        # cotaCmSobreMdb32bits -> calculado a partir de self_aCeldasCoeficientesMdb_[] con las coordenadas de punto de aCeldasListaDePtosTlcPralPF99[]
        28: 'CotaEnDmPlus20SobreMdb',
        # Esta variable la uso si GLBLcalcularMdp and GLBLincluirVarDeMdpEnElModeloAcumulativo
        # CotaEnDmPlus20SobreMdf calculada a partir de cotaSobreMdfEnMetrosElegida, calculada a partir de cotaCmSobreMdf16bitsElegida
        # cotaCmSobreMdf16bitsElegida = aCeldasListaDePtosTlcPralPF99['cotaCmSobreMdfManual16bits'] o aCeldasListaDePtosTlcPralPF99['cotaCmSobreMdfConvol16bits'] o aCeldasListaDePtosTlcPralPF99['cotaCmSobreMdfConual16bits']
        # Extraido de aCeldasListaDePtosTlcPralPF99 = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        29: 'CotaEnDmPlus20SobreMdf',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> No uso estas variables
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> Dependientes del modelo convolucional
        # El miniSubcel de la celda en la que cae el punto esta clasificado como suelo segun el modelo convolucional 
        30: 'esSueloMiniSubCel',
        # El miniSubcel de la celda en la que cae el punto esta clasificado como vegetacion segun el modelo convolucional 
        31: 'esVegetaMiniSubCel',
        # El miniSubcel de la celda en la que cae el punto esta clasificado como edificio segun el modelo convolucional 
        32: 'esEdifiMiniSubCel',

        # Esta en una subcelda clasificada como edificio segun el modelo convolucional entrenado con cartoSingu
        33: 'esEdiSubCeldaUsoSingularPredicho',
        # Esta en una subcelda clasificada como via segun el modelo convolucional entrenado con cartoSingu
        34: 'esViaSubCeldaUsoSingularPredicho',
        # Esta en una subcelda clasificada como puente segun el modelo convolucional entrenado con cartoSingu
        35: 'esPteSubCeldaUsoSingularPredicho',

        # Esta en una subcelda clasificada como agua segun el modelo convolucional entrenado con cartoSingu
        100: 'esAguSubCeldaUsoSingularPredicho',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # Distancia en x2cm hasta el punto miniSubCel mas cercano (en dm x5 y truncada a 255)
        36: 'distDmX5HastaMiniSubCel0',
        # Cota en dm sobre el punto miniSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        37: 'cotaDmPlus20SobreMiniSubCel0',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        38: 'pteReEscaladaSobreMiniSubCel0',
        # Es suelo el punto miniSubCel mas cercano
        39: 'esSueMiniSubCel0',
        # Es vegetacion el punto miniSubCel mas cercano
        40: 'esVegMiniSubCel0',
        # Es edificio el punto miniSubCel mas cercano
        41: 'esEdiMiniSubCel0',
        # Es agua el punto miniSubCel mas cercano
        42: 'esAguMiniSubCel0',
        # Es via de comunicacion el punto miniSubCel mas cercano
        43: 'esViaMiniSubCel0',

        # Distancia en x2cm hasta el segundo punto miniSubCel mas cercano (en dm x5 y truncada a 255)
        44: 'distDmX5HastaMiniSubCel1',
        # Cota en dm sobre el segundo punto miniSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        45: 'cotaDmPlus20SobreMiniSubCel1',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        46: 'pteReEscaladaSobreMiniSubCel1',
        # LasClass original del segundo punto miniSubCel mas cercano
        47: 'esSueMiniSubCel1',
        # LasClass predicha convolucionalmente del segundo punto miniSubCel mas cercano
        48: 'esVegMiniSubCel1',
        # Es edificio el punto miniSubCel mas cercano
        49: 'esEdiMiniSubCel1',
        # Es agua el punto miniSubCel mas cercano
        50: 'esAguMiniSubCel1',
        # Es via de comunicacion el punto miniSubCel mas cercano
        51: 'esViaMiniSubCel1',

        # Distancia en x2cm hasta el tercer punto miniSubCel mas cercano (en dm x5 y truncada a 255)
        52: 'distDmX5HastaMiniSubCel2',
        # Cota en dm sobre el tercer punto miniSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        53: 'cotaDmPlus20SobreMiniSubCel2',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        54: 'pteReEscaladaSobreMiniSubCel2',
        # LasClass original del tercer punto miniSubCel mas cercano
        55: 'esSueMiniSubCel2',
        # LasClass predicha convolucionalmente del tercer punto miniSubCel mas cercano
        56: 'esVegMiniSubCel2',
        # Es edificio el punto miniSubCel mas cercano
        57: 'esEdiMiniSubCel2',
        # Es agua el punto miniSubCel mas cercano
        58: 'esAguMiniSubCel2',
        # Es via de comunicacion el punto miniSubCel mas cercano
        59: 'esViaMiniSubCel2',

        # Distancia en x2cm hasta el punto maxiSubCel mas cercano (en dm x5 y truncada a 255)
        60: 'distDmX5HastaMaxiSubCel0',
        # Cota en dm sobre el punto maxiSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        61: 'cotaDmPlus50BajoMaxiSubCel0',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        62: 'pteReEscaladaBajoMaxiSubCel0',
        # Distancia en x2cm hasta el punto maxiSubCel mas cercano (en dm x5 y truncada a 255)
        63: 'distDmX5HastaMaxiSubCel1',
        # Cota en dm sobre el punto maxiSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        64: 'cotaDmPlus50BajoMaxiSubCel1',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        65: 'pteReEscaladaBajoMaxiSubCel1',
        # Distancia en x2cm hasta el punto maxiSubCel mas cercano (en dm x5 y truncada a 255)
        66: 'distDmX5HastaMaxiSubCel2',
        # Cota en dm sobre el punto maxiSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        67: 'cotaDmPlus50BajoMaxiSubCel2',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        68: 'pteReEscaladaBajoMaxiSubCel2',

        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        69: 'distMtX2HastaMiniCel0',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        70: 'cotaDmPlus20SobreMiniCel0',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        71: 'pteReEscaladaSobreMiniCel0',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        72: 'distMtX2HastaMiniCel1',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        73: 'cotaDmPlus20SobreMiniCel1',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        74: 'pteReEscaladaSobreMiniCel1',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        75: 'distMtX2HastaMiniCel2',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        76: 'cotaDmPlus20SobreMiniCel2',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        77: 'pteReEscaladaSobreMiniCel2',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        78: 'distMtX2HastaMiniCel3',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        79: 'cotaDmPlus20SobreMiniCel3',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        80: 'pteReEscaladaSobreMiniCel3',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        81: 'distMtX2HastaMiniCel4',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        82: 'cotaDmPlus20SobreMiniCel4',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        83: 'pteReEscaladaSobreMiniCel4',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        84: 'distMtX2HastaMiniCel5',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        85: 'cotaDmPlus20SobreMiniCel5',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        86: 'pteReEscaladaSobreMiniCel5',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        87: 'distMtX2HastaMiniCel6',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        88: 'cotaDmPlus20SobreMiniCel6',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        89: 'pteReEscaladaSobreMiniCel6',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        90: 'distMtX2HastaMiniCel7',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        91: 'cotaDmPlus20SobreMiniCel7',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        92: 'pteReEscaladaSobreMiniCel7',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        93: 'distMtX2HastaMiniCel8',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        94: 'cotaDmPlus20SobreMiniCel8',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        95: 'pteReEscaladaSobreMiniCel8',

        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minAnisotropy = miPtoNpExtrVar['minAnisotropy']
        96: 'minAnisotropy',
        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minPlanarity = miPtoNpExtrVar['minPlanarity']
        97: 'minPlanarity',
        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minSphericity = miPtoNpExtrVar['minSphericity']
        98: 'minSphericity',
        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minLinearity = miPtoNpExtrVar['minLinearity']
        99: 'minLinearity',

        101: 'tSNEx',
        102: 'tSNEy',
        103: 'tSNEz',
        104: 'tSNElabel',
    }

    # ==========================================================================
    # Review of Input Variable Selection Methods for Artificial Neural Networks 
    # https://www.researchgate.net/profile/Mah-Parsa/post/What-is-the-best-approach-to-select-input-variables-in-Artificial-Neural-Networks-ANNs/attachment/59d63c46c49f478072ea7c27/AS%3A273748995837952%401442278254924/download/input+selection.pdf
    # ==========================================================================
    
    # Variables obligadas (0-1-3-4-8)
    if nInputVars == 0 or nInputVars == 1 or nInputVars == 3:
        listaNumerosVariablesRGBI = [] # +0 variables
    elif nInputVars == 2:
        listaNumerosVariablesRGBI = [
            10, # NDVI metrico
        ] # +1 variables
    elif nInputVars in [4, 5, 6, 7, 8]:
        listaNumerosVariablesRGBI = [
            5, # Intensity Single Return
            10, 11, # NDVI, NDWI metricos
        ] # +3 variables
    elif nInputVars == 9:
        listaNumerosVariablesRGBI = [
            0, 1, 2, 3, # RGBI
        ] # +4 variables
    elif nInputVars == 58:
        listaNumerosVariablesRGBI = [
            5, # Intensity Single Return
        ] # +1 variables
    else:
        listaNumerosVariablesRGBI = [
            0, 1, 2, 3, # RGBI
            5, # Intensity Single Return
            9, 10, 11, # EVI2, NDVI, NDWI metricos
        ] # +8 variables

    # Variables obligadas (0-1-2-3-4-7)
    if nInputVars <= 2 or nInputVars == 4:
        listaNumerosVariablesGeometriaBasica = [] # +0 variables
    elif nInputVars == 5:
        listaNumerosVariablesGeometriaBasica = [
            15, # Geometria de la nube puntual
        ] # +1 variables
    elif nInputVars == 3 or nInputVars == 6:
        listaNumerosVariablesGeometriaBasica = [
            15, 18, # Geometria de la nube puntual
        ] # +2 variables
    elif nInputVars == 7:
        listaNumerosVariablesGeometriaBasica = [
            15, 17, 18, # Geometria de la nube puntual
        ] # +3 variables
    elif nInputVars == 8 or nInputVars == 9:
        listaNumerosVariablesGeometriaBasica = [
            15, 16, 17, 18, # Geometria de la nube puntual
        ] # +4 variables
    else:
        listaNumerosVariablesGeometriaBasica = [
            12, 13, # Posicion relativa del retorno
            14, # miPtoZdecametros
            15, 16, 17, 18, # Geometria de la nube puntual
        ] # +7 variables

    if nInputVars <= 9:
        GLO.GLBLincluirVarUsoSingPredichoEnElModeloAcumulativo = False
        GLO.GLBLincluirVarClaseMiniSubCelPredichaEnElModeloAcumulativo = False
        GLO.GLBLincluirVarGeometriaMiniMaxiSubCelEnElModeloAcumulativo = False
        GLO.GLBLincluirVarGeometriaMiniCelEnElModeloAcumulativo = False
        GLO.GLBLincluirVarDeMdbEnElModeloAcumulativo = False
        GLO.GLBLincluirVarDeHiperformasEnElModeloAcumulativo = False

    # Variables opcionales
    if nInputVars == 0:
        # Si entreno un modelo nuevo, lo hago conforme a las variables:
        # GLBLincluirVarUsoSingPredichoEnElModeloAcumulativo
        # GLBLincluirVarClaseMiniSubCelPredichaEnElModeloAcumulativo
        # GLBLincluirVarGeometriaMiniMaxiSubCelEnElModeloAcumulativo
        # GLBLincluirVarGeometriaMiniCelEnElModeloAcumulativo
        # GLBLincluirVarDeMdbEnElModeloAcumulativo
        # GLBLincluirVarDeMdpEnElModeloAcumulativo
        # GLBLincluirVarDeHiperformasEnElModeloAcumulativo
        if GLO.GLBLpredecirCubiertasSingularesConvolucional and GLO.GLBLincluirVarUsoSingPredichoEnElModeloAcumulativo:
            listaNumerosVariablesUsoSingularPredichoConvolucional = [
                33, 34, 35, 100 # Uso singular predicho de la subCel (esEdi, esVia, esPte, esAgu)
            ] # +4 variables
        else:
            listaNumerosVariablesUsoSingularPredichoConvolucional = []
        if GLO.GLBLpredecirClasificaMiniSubCelConvolucional and GLO.GLBLincluirVarClaseMiniSubCelPredichaEnElModeloAcumulativo:
            listaNumerosVariablesMiniSubCelClasePredicha = [
                30, 31, 32, # Uso predicho del miniSubCel (esSue, esVeg, esEdi)
                39, 40, 41, # Uso predicho del miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                47, 48, 49, # Uso predicho del segundo miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                55, 56, 57, # Uso predicho del tercer miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
            ]
        else: # +12 variables
            listaNumerosVariablesMiniSubCelClasePredicha = []
        if GLO.GLBLincluirVarGeometriaMiniMaxiSubCelEnElModeloAcumulativo:
            listaNumerosVariablesMiniMaxiSubCelGeometria = [
                36, 37, 38, # Distancia, cota y pte hasta el punto miniSubCel mas cercano
                44, 45, 46, # Distancia, cota y pte hasta el segundo punto miniSubCel mas cercano
                52, 53, 54, # Distancia, cota y pte hasta el tercer punto miniSubCel mas cercano
                60, 61, 62, # Distancia, cota y pte hasta el punto maxiSubCel mas cercano
                63, 64, 65, # Distancia, cota y pte hasta el segundo punto maxiSubCel mas cercano
                66, 67, 68, # Distancia, cota y pte hasta el tercer punto maxiSubCel mas cercano
            ]
        else: # +18 variables
            listaNumerosVariablesMiniMaxiSubCelGeometria = []
        if GLO.GLBLincluirVarGeometriaMiniCelEnElModeloAcumulativo:
            listaNumerosVariablesMiniCelGeometria = [
                69, 70, 71, # Distancia, cota y pte hasta el punto miniCel 0 mas cercano
                72, 73, 74, # Distancia, cota y pte hasta el punto miniCel 1 mas cercano
                75, 76, 77, # Distancia, cota y pte hasta el punto miniCel 2 mas cercano
                78, 79, 80, # Distancia, cota y pte hasta el punto miniCel 3 mas cercano
                81, 82, 83, # Distancia, cota y pte hasta el punto miniCel 4 mas cercano
                84, 85, 86, # Distancia, cota y pte hasta el punto miniCel 5 mas cercano
                87, 88, 89, # Distancia, cota y pte hasta el punto miniCel 6 mas cercano
                90, 91, 92, # Distancia, cota y pte hasta el punto miniCel 7 mas cercano
                93, 94, 95, # Distancia, cota y pte hasta el punto miniCel 8 mas cercano
            ]
        else: # +27 variables
            listaNumerosVariablesMiniCelGeometria = []
        if GLO.GLBLcalcularMdb and GLO.GLBLincluirVarDeMdbEnElModeloAcumulativo:
            listaNumerosVariablesMdb = [
                28, # Cota sobre Mdb
            ] # +1 variables
        else:
            listaNumerosVariablesMdb = []
        if GLO.GLBLcalcularMdp and GLO.GLBLincluirVarDeMdpEnElModeloAcumulativo:
            listaNumerosVariablesMdp = [
                29, # Cota sobre Mdf
            ] # +1 variables
        else:
            listaNumerosVariablesMdp = []
        if GLO.GLBLcalcularHiperFormas and GLO.GLBLincluirVarDeHiperformasEnElModeloAcumulativo:
            listaNumerosVariablesHiperFormas = [
                21, 22, 23, 24, 25, 26, 27, # Usarlos si los calculo -> Inconveniente: lo que tarda en calcular el hipercubo
            ] # +7 variables
        else:
            listaNumerosVariablesHiperFormas = []

    else:
        # Si entreno un modelo preentrenado, leo el numero de inputVars del nombre
        #  y elijo la lista de variables a la vista del numero
        # Opcion elegida por el momento: 65 variables, que incluye:
        #  Obligatorias RGBI (0-4-8)
        #  Obligatorias GeometriaBasica (2-3-7)
        #  CotaSobreMdb (1)
        #  UsoSingularPredichoConvolucional (4)
        #  MiniMaxiSubCelGeometria (18)
        #  MiniCelGeometria (27)
        # Y excluye:
        #  CotaSobreMdp (1)
        #  HiperFormas (7)
        #  MiniSubCelClasePredicha (12)
        if nInputVars in [
            1, 2, 3, 4, 5, 6, 7, 8, 9,
            20, 21, 38, 39, 58, 65, 66, 72, 73, 84, 85
        ]:
            listaNumerosVariablesMdb = [
                28, # Cota sobre Mdb
            ] # +1 variables
        else:
            listaNumerosVariablesMdb = []
        if nInputVars in [19, 20, 38, 58, 65, 72, 84, 85]:
            listaNumerosVariablesUsoSingularPredichoConvolucional = [
                33, 34, 35, 100 # Uso singular predicho de la subCel (esEdi, esVia, esPte, esAgu)
            ] # +4 variables
        else:
            listaNumerosVariablesUsoSingularPredichoConvolucional = []
        if nInputVars in [21, 39, 66, 73, 85]:
            listaNumerosVariablesMdp = [
                29, # Cota sobre Mdf
            ] # +1 variables
        else:
            listaNumerosVariablesMdp = []
        # Para las siguientes, de cada 3 numeros consecutivos:
        #  El primero excluye cotaSobreMdb y cotaSobreMdp
        #  El segundo incluye cotaSobreMdb y excluye cotaSobreMdp
        #  El tercero incluye cotaSobreMdb y cotaSobreMdp
        # Por el momento uso el segundo para evitar calcular Mdp
        if nInputVars in [37, 38, 39, 58, 64, 65, 66, 71, 72, 73, 83, 84, 85]:
            listaNumerosVariablesMiniMaxiSubCelGeometria = [
                36, 37, 38, # Distancia, cota y pte hasta el punto miniSubCel mas cercano
                44, 45, 46, # Distancia, cota y pte hasta el segundo punto miniSubCel mas cercano
                52, 53, 54, # Distancia, cota y pte hasta el tercer punto miniSubCel mas cercano
                60, 61, 62, # Distancia, cota y pte hasta el punto maxiSubCel mas cercano
                63, 64, 65, # Distancia, cota y pte hasta el segundo punto maxiSubCel mas cercano
                66, 67, 68, # Distancia, cota y pte hasta el tercer punto maxiSubCel mas cercano
            ]
        else: # +18 variables
            listaNumerosVariablesMiniMaxiSubCelGeometria = []
        if nInputVars in [58, 64, 65, 66, 71, 72, 73, 83, 84, 85]:
            listaNumerosVariablesMiniCelGeometria = [
                69, 70, 71, # Distancia, cota y pte hasta el punto miniCel 0 mas cercano
                72, 73, 74, # Distancia, cota y pte hasta el punto miniCel 1 mas cercano
                75, 76, 77, # Distancia, cota y pte hasta el punto miniCel 2 mas cercano
                78, 79, 80, # Distancia, cota y pte hasta el punto miniCel 3 mas cercano
                81, 82, 83, # Distancia, cota y pte hasta el punto miniCel 4 mas cercano
                84, 85, 86, # Distancia, cota y pte hasta el punto miniCel 5 mas cercano
                87, 88, 89, # Distancia, cota y pte hasta el punto miniCel 6 mas cercano
                90, 91, 92, # Distancia, cota y pte hasta el punto miniCel 7 mas cercano
                93, 94, 95, # Distancia, cota y pte hasta el punto miniCel 8 mas cercano
            ]
        else: # +27 variables
            listaNumerosVariablesMiniCelGeometria = []
        if nInputVars in [71, 72, 73, 83, 84, 85]:
            listaNumerosVariablesHiperFormas = [
                21, 22, 23, 24, 25, 26, 27, # Usarlos si los calculo -> Inconveniente: lo que tarda en calcular el hipercubo
            ] # +7 variables
        else:
            listaNumerosVariablesHiperFormas = []
        if nInputVars in [83, 84, 85]:
            listaNumerosVariablesMiniSubCelClasePredicha = [
                30, 31, 32, # Uso predicho del miniSubCel (esSue, esVeg, esEdi)
                39, 40, 41, # Uso predicho del miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                47, 48, 49, # Uso predicho del segundo miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                55, 56, 57, # Uso predicho del tercer miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
            ]
        else: # +12 variables
            listaNumerosVariablesMiniSubCelClasePredicha = []

    # Variables obligatorias: 15

    # Variables descartadas (20 variables):
    # listaNumerosVariablesMiniSubCelClaseNoPredicha = [
    #     42, 43, # Uso predicho del miniSubCel mas cercano (esAgu, esVia)
    #     50, 51, # Uso predicho del segundo miniSubCel mas cercano (esAgu, esVia)
    #     58, 59, # Uso predicho del tercer miniSubCel mas cercano (esAgu, esVia)
    # ] # +6 variables
    # listaNumerosVariablesDescartados = [
    #     4, # intensity
    #     6, 7, 8, # puntoEVI2ReEscalado, puntoNDVIReEscalado, puntoNDWIReEscalado
    #     19, 20, # Disponibles (random)
    #     96, 97, 98, 99, # anisotropy
    #     101, 102, 103, 104, # tSNE
    # ] # +14 variables
    listaNumerosVariablesForTrain = (
        listaNumerosVariablesRGBI # (0-3-4-8) variables
        + listaNumerosVariablesGeometriaBasica # (1-2-3-4-7) variables
        + listaNumerosVariablesMdb # 1 variables
        + listaNumerosVariablesUsoSingularPredichoConvolucional # 0-4 variables
        + listaNumerosVariablesMiniMaxiSubCelGeometria # 0-18 variables
        + listaNumerosVariablesMiniCelGeometria # 0-27 variables
        + listaNumerosVariablesHiperFormas # 0-7 variables
        + listaNumerosVariablesMiniSubCelClasePredicha # 0-12 variables
        + listaNumerosVariablesMdp # 0-1 variables
    ) # maximo: 85 variables
    listaNumerosVariablesForTrain.sort()
    print('clidconfig-> Numero de Variables seleccionadas forTrain: {}'.format(nInputVars))
    print(f'{TB}-> listaNumerosVariablesForTrain: {len(listaNumerosVariablesForTrain)}->{listaNumerosVariablesForTrain}')
    for numVariableForTrain in listaNumerosVariablesForTrain:
        print(f'{TB}{TV}-> Variable {numVariableForTrain}: {dictVariablesEndogenasTodas[numVariableForTrain]}')
    print(f'{"":=^80}')

    return dictVariablesEndogenasTodas, listaNumerosVariablesForTrain


def foo1():
    pass

# ==============================================================================
# Dejo esto por si quiero ver como funciona la pila de llamadas segun desde donde se cargue este modulo
if False:
    mostrarModuloInicial = True
    if mostrarModuloInicial:
        if len(inspect.stack()) > 1:
            try:
                callingModuleActual = inspect.getmodulename(inspect.stack()[0][1])
                callingModuleInicial = inspect.getmodulename(inspect.stack()[-1][1])
                callingModulePrevio = 'desconocido'
                for nOrden, llamada in enumerate(inspect.stack()[1:]):
                    #print('\t->', nOrden, llamada[1])
                    if not llamada[1].startswith('<frozen'):
                        callingModulePrevio = inspect.getmodulename(llamada[1])
                        break
            except:
                print('Error al identificar el modulo')
        else:
            callingModuleActual = 'Este modulo'
            callingModuleInicial = 'Este modulo'
            callingModulePrevio = 'Modulo no importado desde otro modulo'
        # print('clidconfig->  Este modulo:                   ', callingModuleActual)
        # print('clidconfig->  Modulo desde el que se importa:', callingModulePrevio)
        # print('clidconfig->  Modulo ejecutado inicialmente: ', callingModuleInicial)
# ==============================================================================


def foo2():
    pass

# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# Si clidconfig estuviera pensado solo para cuando se arranca desde clidbase.py
# esto lo podria quitar de aqui y ejecutarlo en el propio clidbase.py.
# Pero como hay otras alternativas de ejecucion de cartolidar (p. ej. cuando lo instalo)
# dejo la creacion del cfg con los parametros de configuracion en este modulo
# Este cfg tiene los parametros con un primer chequeo de confirmarExistenciaValidezDeAlgunosParametrosDeConfiguracion<>
# Luego, los puedo retocar desde donde quiera, p. ej. tras leer los argumentos en linea de comandos
if (
    callingModuleInicial == 'clidbase'
    or callingModuleInicial == 'clidflow'
    or callingModuleInicial == 'clidtools'
    or callingModuleInicial == 'clidclas'
    or callingModuleInicial == 'clidtry'
    or callingModuleInicial == 'clidgis'
):
    # controlFileLasName = None
    # controlFileLasObj = None
    # controlFileGralObj = None

    if CONFIGverbose:
        # Ver tb: clidbase-> Se van a cargar clidconfig y clidaux. Al cargar clidconfig se ejecutan:
        print('''
clidconfig-> Secuencia de carga de variables de configuracion:
            # Esto solo se hace una vez, aqui en clidconfig, para crear el .cfg  (si no existe ya)
            # Ese .cfg es lo que leeran todos los modulos con leerCambiarVariablesGlobales<> para crear su GLOBALconfigDict
             1. Se lanza initConfigDicts<> para:
                 a. Buscar y, si existe, leer el fichero cfg
                    -> Si hay fichero cfg de anteriores ejecucuiones
                       Se leen los parametros del cfg y se guardan en GLOBALconfigDict
                    -> Si no hay fichero cfg se pasa a la opcion b
                       (se tira del fichero xlsx de configuracion).
                 b. Buscar y leer el fichero xlsx
                    Esta funcion devuelve la variable GLOBALconfigDict (globales de clidconfig)
                    Ademas hace una copia del xlsx que dejo solo como testimonio (no la uso)
                    AVISO: aqui se eligen las columnas a usar para las configuraciones extra
                      Por el momento solo la columna I, con valor CREA_TILES (eso queda guardado en GLOBALconfigDict)
             2. Creo el fichero de configuracion .cfg correspondiente con guardarVariablesGlobales<> a partir de GLOBALconfigDict.
                De esta forma esta disponible para todos los modulos 
                  Lo puedo actualizar si hace falta y esta disponible para todos,
                  cosa que es mas dificil con GLO salvo que fuerce la recarga de clidconfig desde cada modulo.
                  En principio solo se carga la primera vez que lo llamo
                  (cosa que ocurre desde clidaux y no desde clidbase, que llama primero a clidaux).
                Incluyo todas las configuraciones extra, ademas de la general.''')
    LCL_idProceso=MAIN_idProceso
    # Leo los parametros de configuracion de clidbase.xml y los cargo en diccionarios
    if CONFIGverbose:
        print('clidconfig-> Para leer el fichero de configuracion xlsx lanzo initConfigDicts<> (se hace copia con LCL_idProceso: {})'.format(LCL_idProceso))

    # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    # Esto solo se hace una vez, aqui en clidconfig, para crear el .cfg
    # Ese .cfg es lo que leeran todos los modulos para crear su GLOBALconfigDict
    # Eso lo hacen con leerCambiarVariablesGlobales<>
    GLOBALconfigDict = initConfigDicts(LCL_idProceso, LOCL_verbose=__verbose__)
    # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
    if usarXLS and CONFIGverbose:
        print(f'\nclidconfig-> Algunos valores tras tras lanzar initConfigDicts<> y leer el fichero de configuracion xls:')
        print(f'{TB}-> MAINobjetivoEjecucion1: {GLOBALconfigDict["MAINobjetivoEjecucion"][0]}')
        print(f'{TB}-> MAINprocedimiento1:     {GLOBALconfigDict["MAINprocedimiento"][0]}')
        print(f'{TB}-> MAINcuadrante1:         {GLOBALconfigDict["MAINcuadrante"][0]}')
        print(f'{TB}-> GLBLverbose1:           {GLOBALconfigDict["GLBLverbose"][0]}')
        print(f'{TB}-> Numero de parametros:   {len(GLOBALconfigDict)}')
        print(f'\nclidconfig-> Esto sale del xlsx o cfg antes de tener en cuenta los argumentos en linea de comandos')
        print(f'{TB}-> os args se leen e interpretan en clidbase, despues de importar clidconfig (despues de pasar por aqui).')
    # Creo el objeto GLO (de la clase VariablesGlobales) y le asigno los paramConfig como propiedades
    #  Uso el parametro MAINobjetivoEjecucion para elegir la columna del excel que le corresponde (self.numObjetivoEjecucion)
    # Ademas reviso los paramConfig cambiando o ampliando los que proceda
    if CONFIGverbose:
        print(f'clidconfig-> Voy a crear el objeto GLO (objeto global de clidconfig de la clase GLO_CLASS<>)')
        print(f'\t-> al que puedo acceder desde el resto de los modulos importando este modulo')
    # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    GLO = GLO_CLASS(GLOBALconfigDict, LCLverbose=__verbose__)
    # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

    if usarXLS:
        if CONFIGverbose:
            print(f'clidconfig-> Revisando la coherencia y validez de la configuracion')
        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
        ok1 = GLO.revisarCompletarVariablesMAINdelConfigVarsDict()
        ok2 = GLO.revisarCompletarVariablesGLBLdelConfigVarsDict()
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
        if not ok1 or not ok2:
            print(f'clidconfig-> Configuracion no permitida: cambiarla')
            sys.exit(0)

    # ==========================================================================
    GLOBALconfigDict = GLO.configVarsDict
    # ==========================================================================

    # ==========================================================================
    if usarXLS: 
        if CONFIGverbose:
            print(f'clidconfig-> Chequeo de algunos parametros antes de tener en cuenta los argumentos en linea de comandos:')
            print(f'{TB}-> MAINobjetivoEjecucion GLOBALconfigDict (1): {GLOBALconfigDict["MAINobjetivoEjecucion"][GLO.numObjetivoEjecucion]}')
            print(f'{TB}-> MAINobjetivoEjecucion GLO (1):              {GLO.MAINobjetivoEjecucion}')
            print(f'{TB}-> numObjetivoEjecucion GLO (1):               {GLO.numObjetivoEjecucion}')
            print(f'{TB}-> MAINprocedimiento GLOBALconfigDict (1):     {GLOBALconfigDict["MAINprocedimiento"][GLO.numObjetivoEjecucion]}')
            print(f'{TB}-> MAINprocedimiento GLO (1):                  {GLO.MAINprocedimiento}')
            # input('Pulsa algo 1')
        # ==========================================================================
    
        # ==============================================================================
        # Esto ya no deberia ser necesario, porque no puede ocurrir
        if GLOBALconfigDict["MAINprocedimiento"][GLO.numObjetivoEjecucion] != GLO.MAINprocedimiento:
            print(f'clidconfig-> AVISO: revisar la asignacion de MAINprocedimiento.')
            print(f'{TB}-> MAINprocedimiento GLOBALconfigDict (1): {GLOBALconfigDict["MAINprocedimiento"][0]}')
            print(f'{TB}-> MAINprocedimiento GLO              (1): {GLO.MAINprocedimiento}')
            print(f'clidconfig-> Esto se debia a que en GLO.revisarCompletarVariablesMAINdelConfigVarsDict<>')
            print(f'{TB}-> Se cambia el AUTOMATICO_EN_CALENDULA_... cuando no esta en calendula')
            print(f'{TB}-> Y no se actualizaba la propiedad en GLO')
            print(f'{TB}-> Pero ya esta solucionado')
        # ==============================================================================

    # Vuelvo a guardarlos en el fichero clidbase*.cfg:
    if CONFIGverbose:
        print(f'clidconfig-> paramConfig tras cargarlos en GLO-> GLOBALconfigDict[GLBLverbose]: {len(GLOBALconfigDict["GLBLverbose"])}, {GLOBALconfigDict["GLBLverbose"]}')
        # print('clidconfig-> sys.argv:', len(sys.argv), '->', sys.argv)
        print('clidconfig-> Guardo los paramConfig en el fichero cfg')
    guardarVariablesGlobales(
        GLOBALconfigDict,
        LCL_idProceso=MAIN_idProceso,
    )

    # Despues de guardarlos en cfg puedo leer los paramtros de configuracion desde cualquier funcion con:
    # GLOBALconfigDict = leerCambiarVariablesGlobales(LCL_idProceso=MAIN_idProceso)
    # Tambien puedo cambiar los paramtros que quiera de forma permanente o agregar nuevos.
    # Ejemplo de como cambiar o agregar parametros de configuracion
    # print('clidconfig->\n\t1.- GLOBALconfigDict', GLOBALconfigDict)
    # nuevosParametroConfiguracion={}
    # nuevosParametroConfiguracion['MAINusarValorALTER'] = [True, 'GrupoMAIN', 'Nueva descripcion 1', 'bool']
    # GLOBALconfigDict = leerCambiarVariablesGlobales(LCL_idProceso=MAIN_idProceso, nuevosParametroConfiguracion=nuevosParametroConfiguracion)
    # print('clidconfig->\n\t2.- GLOBALconfigDict', GLOBALconfigDict)

else:
    # print('clidconfig-> No se cargan las variables globales de clidbase')
    pass

